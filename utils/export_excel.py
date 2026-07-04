import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_invoice_excel(invoices, summary):
    wb = Workbook()
    
    font_family = "Calibri"
    font_regular = Font(name=font_family, size=11)
    font_bold = Font(name=font_family, size=11, bold=True)
    font_title = Font(name=font_family, size=16, bold=True)
    
    fill_header = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_total = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    double_bottom_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='double', color='000000')
    )
    
    ws = wb.active
    ws.title = "Danh sách hóa đơn"
    ws.views.sheetView[0].showGridLines = True
    
    # Title
    ws['A1'] = "DANH SÁCH HÓA ĐƠN"
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center
    ws.merge_cells('A1:I1')
    ws.row_dimensions[1].height = 40
    
    # Table headers
    headers = [
        "STT", "Mã hóa đơn", "Ngày lập", "Khách hàng", 
        "Số điện thoại", "Tổng tiền", "Giảm giá", "Thanh toán", "Ghi chú"
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[3].height = 25
    
    current_row = 4
    for idx, inv in enumerate(invoices, start=1):
        # STT
        c_stt = ws.cell(row=current_row, column=1, value=idx)
        c_stt.alignment = align_center
        
        # Mã hóa đơn
        c_id = ws.cell(row=current_row, column=2, value=f"HD{inv.id}")
        c_id.alignment = align_center
        
        # Ngày lập
        date_val = inv.invoice_date.strftime('%d/%m/%Y') if inv.invoice_date else 'N/A'
        c_date = ws.cell(row=current_row, column=3, value=date_val)
        c_date.alignment = align_center
        
        # Khách hàng
        c_cust = ws.cell(row=current_row, column=4, value=inv.customer.name if inv.customer else '')
        c_cust.alignment = align_left
        
        # Số điện thoại
        c_phone = ws.cell(row=current_row, column=5, value=inv.customer.phone if inv.customer else '')
        c_phone.alignment = align_center
        c_phone.number_format = '@'
        
        # Tổng tiền
        c_total = ws.cell(row=current_row, column=6, value=inv.total_amount)
        c_total.alignment = align_right
        c_total.number_format = '#,##0" VND"'
        
        # Giảm giá
        c_disc = ws.cell(row=current_row, column=7, value=inv.discount)
        c_disc.alignment = align_right
        c_disc.number_format = '#,##0" VND"'
        
        # Thanh toán
        c_pay = ws.cell(row=current_row, column=8, value=getattr(inv, 'display_payment_method', None) or inv.payment_method or 'Không rõ')
        c_pay.alignment = align_center
        
        # Ghi chú
        c_notes = ws.cell(row=current_row, column=9, value=inv.notes or '')
        c_notes.alignment = align_left
        
        # Apply regular font and thin borders to data cells
        for col_idx in range(1, 10):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
    # Bottom Summary Row: Tổng số hóa đơn, Tổng doanh thu, Tổng giảm giá
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    c_lbl = ws.cell(row=current_row, column=1, value="Tổng cộng")
    c_lbl.font = font_bold
    c_lbl.alignment = align_right
    c_lbl.fill = fill_total
    
    # Total Revenue
    c_tot_rev = ws.cell(row=current_row, column=6, value=summary.get("total_revenue", 0))
    c_tot_rev.font = font_bold
    c_tot_rev.alignment = align_right
    c_tot_rev.number_format = '#,##0" VND"'
    c_tot_rev.fill = fill_total
    
    # Total Discount
    c_tot_disc = ws.cell(row=current_row, column=7, value=summary.get("total_discount", 0))
    c_tot_disc.font = font_bold
    c_tot_disc.alignment = align_right
    c_tot_disc.number_format = '#,##0" VND"'
    c_tot_disc.fill = fill_total
    
    # Fill remaining columns in summary row with styles
    for col_idx in range(1, 10):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.border = double_bottom_border
        cell.fill = fill_total
        if col_idx in [8, 9]:
            cell.value = ""
            
    # Write summary texts below table
    current_row += 2
    ws.cell(row=current_row, column=1, value="Tổng số hóa đơn:").font = font_bold
    ws.cell(row=current_row, column=3, value=summary.get("invoice_count", 0)).font = font_regular
    
    current_row += 1
    ws.cell(row=current_row, column=1, value="Tổng doanh thu:").font = font_bold
    c_sum_rev = ws.cell(row=current_row, column=3, value=summary.get("total_revenue", 0))
    c_sum_rev.font = font_regular
    c_sum_rev.number_format = '#,##0" VND"'
    
    current_row += 1
    ws.cell(row=current_row, column=1, value="Tổng giảm giá:").font = font_bold
    c_sum_disc = ws.cell(row=current_row, column=3, value=summary.get("total_discount", 0))
    c_sum_disc.font = font_regular
    c_sum_disc.number_format = '#,##0" VND"'
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Exclude title cell (row 1) from width calculations
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if cell.number_format and 'VND' in cell.number_format:
                val_str += " VND"
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream


def generate_statistics_excel(summary, customer_stats, service_stats):
    wb = Workbook()
    
    # Define styles
    font_family = "Calibri"
    font_regular = Font(name=font_family, size=11)
    font_bold = Font(name=font_family, size=11, bold=True)
    font_title = Font(name=font_family, size=16, bold=True)
    
    fill_header = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # ------------------ SHEET 1: Tổng quan ------------------
    ws1 = wb.active
    ws1.title = "Tổng quan"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title
    ws1['A1'] = "BÁO CÁO THỐNG KÊ TỔNG QUAN"
    ws1['A1'].font = font_title
    ws1.row_dimensions[1].height = 30
    
    # Table headers
    headers1 = ["Chỉ số", "Giá trị"]
    for col_idx, h in enumerate(headers1, start=1):
        cell = ws1.cell(row=3, column=col_idx, value=h)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws1.row_dimensions[3].height = 24
    
    # Data for summary
    data1 = [
        ("Tổng doanh thu", summary.get("revenue", 0)),
        ("Tổng hóa đơn", summary.get("invoice_count", 0)),
        ("Tổng khách hàng", summary.get("customer_count", 0)),
        ("Tổng lịch hẹn", summary.get("appointment_count", 0))
    ]
    
    for row_idx, (indicator, value) in enumerate(data1, start=4):
        c1 = ws1.cell(row=row_idx, column=1, value=indicator)
        c1.font = font_regular
        c1.alignment = align_left
        c1.border = thin_border
        
        c2 = ws1.cell(row=row_idx, column=2, value=value)
        c2.font = font_regular
        c2.border = thin_border
        
        if indicator == "Tổng doanh thu":
            c2.alignment = align_right
            c2.number_format = '#,##0" VND"'
        else:
            c2.alignment = align_right
            c2.number_format = '#,##0'
            
        ws1.row_dimensions[row_idx].height = 20

    # ------------------ SHEET 2: Thống kê khách hàng ------------------
    ws2 = wb.create_sheet(title="Thống kê khách hàng")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2['A1'] = "THỐNG KÊ KHÁCH HÀNG"
    ws2['A1'].font = font_title
    ws2.row_dimensions[1].height = 30
    
    headers2 = ["STT", "Khách hàng", "SĐT", "Số hóa đơn", "Tổng tiền đã chi"]
    for col_idx, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws2.row_dimensions[3].height = 24
    
    for idx, cust in enumerate(customer_stats, start=1):
        row_idx = idx + 3
        
        c_stt = ws2.cell(row=row_idx, column=1, value=idx)
        c_stt.font = font_regular
        c_stt.alignment = align_center
        c_stt.border = thin_border
        
        c_name = ws2.cell(row=row_idx, column=2, value=cust.get("name", ""))
        c_name.font = font_regular
        c_name.alignment = align_left
        c_name.border = thin_border
        
        c_phone = ws2.cell(row=row_idx, column=3, value=cust.get("phone") or "")
        c_phone.font = font_regular
        c_phone.alignment = align_center
        c_phone.border = thin_border
        c_phone.number_format = '@'
        
        c_inv = ws2.cell(row=row_idx, column=4, value=cust.get("invoice_count", 0))
        c_inv.font = font_regular
        c_inv.alignment = align_right
        c_inv.border = thin_border
        c_inv.number_format = '#,##0'
        
        c_spent = ws2.cell(row=row_idx, column=5, value=cust.get("total_spent", 0))
        c_spent.font = font_regular
        c_spent.alignment = align_right
        c_spent.border = thin_border
        c_spent.number_format = '#,##0" VND"'
        
        ws2.row_dimensions[row_idx].height = 20

    # ------------------ SHEET 3: Thống kê dịch vụ ------------------
    ws3 = wb.create_sheet(title="Thống kê dịch vụ")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3['A1'] = "THỐNG KÊ DỊCH VỤ"
    ws3['A1'].font = font_title
    ws3.row_dimensions[1].height = 30
    
    headers3 = ["STT", "Tên dịch vụ", "Số lượt sử dụng", "Doanh thu"]
    for col_idx, h in enumerate(headers3, start=1):
        cell = ws3.cell(row=3, column=col_idx, value=h)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws3.row_dimensions[3].height = 24
    
    for idx, svc in enumerate(service_stats, start=1):
        row_idx = idx + 3
        
        c_stt = ws3.cell(row=row_idx, column=1, value=idx)
        c_stt.font = font_regular
        c_stt.alignment = align_center
        c_stt.border = thin_border
        
        c_name = ws3.cell(row=row_idx, column=2, value=svc.get("service_name", ""))
        c_name.font = font_regular
        c_name.alignment = align_left
        c_name.border = thin_border
        
        c_qty = ws3.cell(row=row_idx, column=3, value=svc.get("quantity_sold", 0))
        c_qty.font = font_regular
        c_qty.alignment = align_right
        c_qty.border = thin_border
        c_qty.number_format = '#,##0'
        
        c_rev = ws3.cell(row=row_idx, column=4, value=svc.get("revenue", 0))
        c_rev.font = font_regular
        c_rev.alignment = align_right
        c_rev.border = thin_border
        c_rev.number_format = '#,##0" VND"'
        
        ws3.row_dimensions[row_idx].height = 20

    # Auto-adjust column widths for all sheets
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Exclude title cell (row 1) from width calculations
                if cell.row == 1:
                    continue
                val_str = str(cell.value or '')
                if cell.number_format and 'VND' in cell.number_format:
                    val_str += " VND"
                if len(val_str) > max_len:
                    max_len = len(val_str)
            # Add padding
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream
