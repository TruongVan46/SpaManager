import os
import shutil
import tempfile
import unittest
import re
import base64
import zlib
import json
import html as html_module
import uuid
import inspect
import sqlite3
import subprocess
from types import SimpleNamespace
from contextlib import nullcontext
from io import BytesIO
from datetime import datetime, timedelta
from pathlib import Path
from flask import redirect
from unittest.mock import patch

from openpyxl import Workbook
from sqlalchemy import event, text, inspect as sa_inspect
from sqlalchemy.exc import IntegrityError, SQLAlchemyError

TEST_DB_FILE = Path(tempfile.gettempdir()) / "spamanager_owner_seed_test.sqlite"
TEST_MEDIA_ROOT = Path(tempfile.gettempdir()) / "spamanager_media_seed_test"
if TEST_DB_FILE.exists():
    TEST_DB_FILE.unlink()
if TEST_MEDIA_ROOT.exists():
    shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)
os.environ["APP_ENV"] = "testing"
os.environ["TEST_DATABASE_URL"] = f"sqlite:///{TEST_DB_FILE.as_posix()}"
os.environ["PERSISTENT_ROOT"] = TEST_MEDIA_ROOT.as_posix()
os.environ["UPLOAD_ROOT"] = (TEST_MEDIA_ROOT / "uploads").as_posix()
os.environ["LOGO_UPLOAD_FOLDER"] = (TEST_MEDIA_ROOT / "uploads" / "logos").as_posix()
os.environ["AVATAR_UPLOAD_FOLDER"] = (TEST_MEDIA_ROOT / "uploads" / "avatars").as_posix()

from app import app
from config import DevelopmentConfig, ProductionConfig, TestingConfig, _parse_bool_env
from extensions import db
from core.auth.constants import AUTH_SESSION_KEY
from core.exceptions import AuthenticationException, ConflictException
from models.activity_log import ActivityLog
from models.appointment import Appointment
from models.customer import Customer
from models.invoice import Invoice
from models.invoice_detail import InvoiceDetail
from models.service import Service
from models.setting import Setting
from models.user import User
from models.workspace import Workspace, WorkspaceMember
from services.appointment_service import AppointmentService
from services.activity_log_service import ActivityLogService
from services.customer_service import CustomerService
from services.invoice_service import InvoiceService
from services.service_service import ServiceService
from services.auth_service import AuthService
from services.backup_service import BackupService
from services.data_audit_service import run_data_consistency_audit
from services.data_repair_service import run_controlled_repair
from services.performance_profile_service import profile_block, run_performance_profile
from services.operational_diagnostics_service import run_operational_diagnostics
from services.import_service import ImportService
from services.login_rate_limit_service import reset_all_login_attempts
from repositories.backup_repository import BackupRepository
from utils.database_engine import (
    get_database_engine,
    get_postgresql_backup_center_message,
    get_postgresql_restore_guard_message,
    is_postgresql_database,
    is_sqlite_database,
)
from utils.timezone_utils import local_now
from core.auth.permissions import (
    can_manage_backups,
    can_manage_business_data,
    can_manage_settings,
    can_manage_users,
    can_view_activity_logs,
    is_admin,
    is_manager,
    is_owner,
    is_staff,
)
from core.auth.security import PasswordPolicy
from core.activity_log_utils import sanitize_activity_log_value, get_activity_actor_display_name, build_activity_log_entry
from core.auth.google_oauth import init_google_oauth, is_google_auth_available
import core.csrf as csrf_module
import utils.export_pdf as export_pdf_utils


class BasicTestCase(unittest.TestCase):
    @classmethod
    def tearDownClass(cls):
        try:
            with app.app_context():
                db.session.remove()
                db.engine.dispose()
        finally:
            if TEST_DB_FILE.exists():
                TEST_DB_FILE.unlink()
            if TEST_MEDIA_ROOT.exists():
                shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)

    def setUp(self):
        self.app_context = app.app_context()
        self.app_context.push()
        self.client = app.test_client()
        reset_all_login_attempts()
        self.reset_database_schema()

    def tearDown(self):
        db.session.rollback()
        reset_all_login_attempts()
        self.app_context.pop()

    def reset_database_schema(self):
        db.session.rollback()
        db.drop_all()
        db.session.commit()
        with db.engine.begin() as connection:
            connection.execute(text("DROP TABLE IF EXISTS alembic_version"))
        db.create_all()

    def clear_database_schema(self):
        db.session.rollback()
        db.drop_all()
        db.session.commit()
        with db.engine.begin() as connection:
            connection.execute(text("DROP TABLE IF EXISTS alembic_version"))

    def create_user(
        self,
        username,
        password="secret123",
        full_name="Test User",
        role="STAFF",
        is_active=True,
        approval_status="active",
    ):
        user = User(
            username=username,
            full_name=full_name,
            role=role,
            is_active=is_active,
            approval_status=approval_status,
        )
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        return user

    def login_as(self, user):
        with self.client.session_transaction() as sess:
            sess[AUTH_SESSION_KEY] = user.id

    def get_session_user_id(self):
        with self.client.session_transaction() as sess:
            return sess.get(AUTH_SESSION_KEY)

    def get_flashed_messages_from_session(self):
        with self.client.session_transaction() as sess:
            return [message for _, message in sess.get("_flashes", [])]

    def google_oauth_test_config(self, allowed_domain="example.com"):
        return {
            "GOOGLE_AUTH_ENABLED": True,
            "GOOGLE_CLIENT_ID": "google-client-id",
            "GOOGLE_CLIENT_SECRET": "google-client-secret",
            "GOOGLE_REDIRECT_URI": "https://example.com/auth/google/callback",
            "GOOGLE_ALLOWED_DOMAIN": allowed_domain,
            "GOOGLE_SCOPES": ["openid", "email", "profile"],
        }

    def run_google_callback_with_identity(self, identity):
        class FakeGoogleClient:
            def authorize_access_token(self):
                return {"userinfo": identity}

        original_state = dict(app.extensions.get("google_oauth", {}))
        try:
            with patch.dict(app.config, self.google_oauth_test_config(), clear=False):
                init_google_oauth(app)
                app.extensions["google_oauth"]["client"] = FakeGoogleClient()
                return self.client.get("/auth/google/callback?code=fake-code&state=fake-state")
        finally:
            app.extensions["google_oauth"] = original_state

    def get_csrf_token(self, path="/login"):
        response = self.client.get(path, follow_redirects=True)
        html = response.get_data(as_text=True)
        match = re.search(r'name="csrf-token" content="([^"]+)"', html)
        if not match:
            match = re.search(r'name="csrf_token" value="([^"]+)"', html)
        self.assertIsNotNone(match, "CSRF token not found in response HTML")
        return match.group(1)

    def post_with_csrf(self, url, path="/login", **kwargs):
        headers = dict(kwargs.pop("headers", {}) or {})
        headers["X-CSRFToken"] = self.get_csrf_token(path)
        return self.client.post(url, headers=headers, **kwargs)

    def create_customer_record(self, name="Test Customer"):
        customer = Customer(name=name, phone="0900000000", email=f"{name.lower().replace(' ', '')}@example.com")
        db.session.add(customer)
        db.session.commit()
        return customer

    def create_service_record(self, name="Test Service"):
        service = Service(name=name, price=100000, duration=30, description="Test", category="other")
        db.session.add(service)
        db.session.commit()
        return service

    def create_appointment_record(self, customer=None, service=None):
        customer = customer or self.create_customer_record("Appointment Customer")
        service = service or self.create_service_record("Appointment Service")
        appointment_date = local_now().date() + timedelta(days=1)
        appointment = Appointment(
            customer_id=customer.id,
            service_id=service.id,
            appointment_time=datetime.combine(appointment_date, datetime.min.time()).replace(hour=10),
            status="Pending"
        )
        db.session.add(appointment)
        db.session.commit()
        return appointment

    def create_invoice_record(self, customer=None, service=None):
        customer = customer or self.create_customer_record("Invoice Customer")
        service = service or self.create_service_record("Invoice Service")
        invoice = Invoice(
            customer_id=customer.id,
            invoice_date=datetime(2026, 7, 4).date(),
            subtotal=100000,
            discount=0,
            total_amount=100000,
            payment_method="Cash",
            notes="Test invoice"
        )
        db.session.add(invoice)
        db.session.flush()
        detail = InvoiceDetail(
            invoice_id=invoice.id,
            service_id=service.id,
            price=100000,
            quantity=1
        )
        db.session.add(detail)
        db.session.commit()
        return invoice

    def create_media_file(self, relative_path, content=b"media-bytes"):
        absolute_path = TEST_MEDIA_ROOT / relative_path
        absolute_path.parent.mkdir(parents=True, exist_ok=True)
        absolute_path.write_bytes(content)
        return absolute_path

    def execute_raw_sql(self, sql, **params):
        with db.engine.begin() as connection:
            connection.execute(text("PRAGMA foreign_keys=OFF"))
            connection.execute(text(sql), params)

    def create_settings_backup_via_route(self, user, notes="Route backup"):
        self.login_as(user)
        response = self.post_with_csrf(
            "/settings/backup",
            path="/settings",
            data={
                "notes": notes,
                "backup_type": "Manual",
                "format": "json",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["success"])
        backup_id = payload["download_url"].rsplit("/", 1)[-1]
        meta = BackupRepository.get_by_id(app, backup_id)
        self.assertIsNotNone(meta)
        backup_path = Path(BackupService.get_backup_file_path(app, meta["filename"]))
        self.assertTrue(backup_path.exists())
        return backup_id, meta, backup_path

    def create_customer_import_xlsx(self, rows=None):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Khách hàng"
        sheet.append(list(ImportService.CUSTOMER_COLUMNS))
        sample_rows = rows or [[f"Import {uuid.uuid4().hex[:8]}", "0901234567", f"{uuid.uuid4().hex[:8]}@example.com", "123 Đường A"]]
        for row in sample_rows:
            sheet.append(row)

        temp_path = Path(tempfile.gettempdir()) / f"settings-import-{uuid.uuid4().hex}.xlsx"
        workbook.save(temp_path)
        workbook.close()
        return temp_path

    def insert_owner_row(self):
        now = datetime.utcnow()
        with db.engine.begin() as connection:
            connection.execute(
                text(
                    """
                    INSERT INTO users
                        (username, password_hash, full_name, avatar, role, is_active,
                         approval_status, approved_by_id, approved_at,
                         last_login, email, email_verified, auth_provider, oauth_id,
                         created_at, updated_at)
                    VALUES
                        (:username, :password_hash, :full_name, NULL, :role, :is_active,
                         :approval_status, NULL, NULL,
                         NULL, NULL, 0, 'local', NULL, :created_at, :updated_at)
                    """
                ),
                {
                    "username": "owner",
                    "password_hash": "existing-owner-hash",
                    "full_name": "Chá»§ Spa",
                    "role": "OWNER",
                    "is_active": 1,
                    "approval_status": "active",
                    "created_at": now,
                    "updated_at": now,
                },
            )

    def test_app_initialization(self):
        self.assertIsNotNone(app)

    def test_login_page_loads(self):
        response = self.client.get("/login")
        self.assertEqual(response.status_code, 200)
        self.assertIn('csrf-token', response.get_data(as_text=True))

    def test_google_login_button_is_hidden_by_default(self):
        response = self.client.get("/login")
        html = response.get_data(as_text=True)

        self.assertFalse(is_google_auth_available())
        self.assertNotIn("Tiếp tục với Google", html)
        self.assertNotIn("/auth/google/start", html)

    def test_google_login_button_shows_when_flag_enabled_and_client_initializes(self):
        original_available = is_google_auth_available()
        original_state = dict(app.extensions.get("google_oauth", {}))

        try:
            with patch.dict(
                app.config,
                {
                    "GOOGLE_AUTH_ENABLED": True,
                    "GOOGLE_CLIENT_ID": "google-client-id",
                    "GOOGLE_CLIENT_SECRET": "google-client-secret",
                    "GOOGLE_REDIRECT_URI": "https://example.com/auth/google/callback",
                    "GOOGLE_ALLOWED_DOMAIN": "example.com",
                    "GOOGLE_SCOPES": ["openid", "email", "profile"],
                },
                clear=False,
            ):
                init_google_oauth(app)
                self.assertTrue(is_google_auth_available())
                response = self.client.get("/login")
                html = response.get_data(as_text=True)
                self.assertIn("Tiếp tục với Google", html)
                self.assertIn("/auth/google/start", html)
        finally:
            app.extensions["google_oauth"] = original_state
            if not original_available:
                self.assertFalse(is_google_auth_available())

    def test_google_login_start_redirects_safely_when_disabled(self):
        response = self.client.get("/auth/google/start")

        self.assertEqual(response.status_code, 302)
        self.assertIn("/login", response.headers.get("Location", ""))
        follow_up = self.client.get("/login", follow_redirects=True)
        self.assertIn("Đăng nhập Google hiện chưa được bật.", follow_up.get_data(as_text=True))

    def test_google_login_callback_redirects_safely_when_disabled(self):
        before_snapshot = (User.query.count(), self.get_session_user_id())

        response = self.client.get("/auth/google/callback")

        self.assertEqual(response.status_code, 302)
        self.assertIn("/login", response.headers.get("Location", ""))
        follow_up = self.client.get("/login", follow_redirects=True)
        self.assertIn("Đăng nhập Google hiện chưa được bật.", follow_up.get_data(as_text=True))
        self.assertEqual((User.query.count(), self.get_session_user_id()), before_snapshot)

    def test_google_login_missing_config_is_unavailable_and_safe(self):
        original_state = dict(app.extensions.get("google_oauth", {}))
        try:
            with patch.dict(
                app.config,
                {
                    "GOOGLE_AUTH_ENABLED": True,
                    "GOOGLE_CLIENT_ID": "",
                    "GOOGLE_CLIENT_SECRET": "",
                    "GOOGLE_REDIRECT_URI": "https://example.com/auth/google/callback",
                    "GOOGLE_ALLOWED_DOMAIN": "example.com",
                    "GOOGLE_SCOPES": ["openid", "email", "profile"],
                },
                clear=False,
            ):
                init_google_oauth(app)
                self.assertFalse(is_google_auth_available())

                login_response = self.client.get("/login")
                self.assertNotIn("Tiếp tục với Google", login_response.get_data(as_text=True))

                start_response = self.client.get("/auth/google/start")
                self.assertEqual(start_response.status_code, 302)
                self.assertIn("/login", start_response.headers.get("Location", ""))

                callback_response = self.client.get("/auth/google/callback")
                self.assertEqual(callback_response.status_code, 302)
                self.assertIn("/login", callback_response.headers.get("Location", ""))
        finally:
            app.extensions["google_oauth"] = original_state

    def test_google_login_start_uses_mock_client_redirect_when_enabled(self):
        class FakeGoogleClient:
            def __init__(self):
                self.redirect_uri = None

            def authorize_redirect(self, redirect_uri):
                self.redirect_uri = redirect_uri
                return redirect("https://accounts.google.test/oauth")

        original_state = dict(app.extensions.get("google_oauth", {}))
        fake_client = FakeGoogleClient()
        try:
            with patch.dict(
                app.config,
                {
                    "GOOGLE_AUTH_ENABLED": True,
                    "GOOGLE_CLIENT_ID": "google-client-id",
                    "GOOGLE_CLIENT_SECRET": "google-client-secret",
                    "GOOGLE_REDIRECT_URI": "https://example.com/auth/google/callback",
                    "GOOGLE_ALLOWED_DOMAIN": "example.com",
                    "GOOGLE_SCOPES": ["openid", "email", "profile"],
                },
                clear=False,
            ):
                init_google_oauth(app)
                app.extensions["google_oauth"]["client"] = fake_client
                response = self.client.get("/auth/google/start")

        finally:
            app.extensions["google_oauth"] = original_state

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers.get("Location"), "https://accounts.google.test/oauth")
        self.assertEqual(fake_client.redirect_uri, "https://example.com/auth/google/callback")

    def test_google_login_callback_error_redirects_without_side_effects(self):
        before_snapshot = (User.query.count(), self.get_session_user_id())

        response = self.client.get("/auth/google/callback?error=access_denied")

        self.assertEqual(response.status_code, 302)
        self.assertIn("/login", response.headers.get("Location", ""))
        self.assertIn("Đăng nhập Google không hoàn tất.", self.get_flashed_messages_from_session())
        self.assertEqual((User.query.count(), self.get_session_user_id()), before_snapshot)

    def test_google_login_callback_valid_identity_creates_pending_user(self):
        response = self.run_google_callback_with_identity({
            "sub": "google-sub-1",
            "email": "new.google@example.com",
            "email_verified": True,
            "name": "Google Pending",
        })

        self.assertEqual(response.status_code, 302)
        self.assertIn("/auth/pending", response.headers.get("Location", ""))

        user = User.query.filter_by(auth_provider="google", oauth_id="google-sub-1").first()
        self.assertIsNotNone(user)
        self.assertEqual(user.email, "new.google@example.com")
        self.assertTrue(user.email_verified)
        self.assertEqual(user.approval_status, User.APPROVAL_PENDING)
        self.assertFalse(user.is_active)
        self.assertEqual(user.role, "STAFF")
        self.assertIsNone(user.approved_by_id)
        self.assertIsNone(user.approved_at)
        self.assertEqual(self.get_session_user_id(), user.id)
        self.assertFalse(hasattr(user, "access_token"))
        self.assertFalse(hasattr(user, "refresh_token"))

    def test_google_pending_user_cannot_access_dashboard(self):
        response = self.run_google_callback_with_identity({
            "sub": "google-sub-dashboard",
            "email": "pending.dashboard@example.com",
            "email_verified": True,
            "name": "Pending Dashboard",
        })
        self.assertIn("/auth/pending", response.headers.get("Location", ""))

        dashboard_response = self.client.get("/", follow_redirects=False)

        self.assertEqual(dashboard_response.status_code, 302)
        self.assertIn("/auth/pending", dashboard_response.headers.get("Location", ""))

    def test_google_callback_same_sub_does_not_create_duplicate(self):
        identity = {
            "sub": "google-sub-repeat",
            "email": "repeat.google@example.com",
            "email_verified": True,
            "name": "Repeat Google",
        }

        first_response = self.run_google_callback_with_identity(identity)
        second_response = self.run_google_callback_with_identity(identity)

        self.assertIn("/auth/pending", first_response.headers.get("Location", ""))
        self.assertIn("/auth/pending", second_response.headers.get("Location", ""))
        self.assertEqual(User.query.filter_by(auth_provider="google", oauth_id="google-sub-repeat").count(), 1)

    def test_google_callback_existing_rejected_or_disabled_user_is_denied(self):
        for approval_status in (User.APPROVAL_REJECTED, User.APPROVAL_DISABLED):
            user = self.create_user(
                f"google-{approval_status}",
                full_name=f"Google {approval_status}",
                is_active=False,
                approval_status=approval_status,
            )
            user.email = f"{approval_status}.google@example.com"
            user.email_verified = True
            user.auth_provider = "google"
            user.oauth_id = f"google-sub-{approval_status}"
            db.session.commit()

            response = self.run_google_callback_with_identity({
                "sub": user.oauth_id,
                "email": user.email,
                "email_verified": True,
                "name": user.full_name,
            })

            self.assertEqual(response.status_code, 302)
            self.assertIn("/login", response.headers.get("Location", ""))
            self.assertNotEqual(self.get_session_user_id(), user.id)

    def test_google_callback_existing_active_linked_user_can_login(self):
        user = self.create_user("google-active", full_name="Google Active", approval_status=User.APPROVAL_ACTIVE)
        user.email = "active.google@example.com"
        user.email_verified = True
        user.auth_provider = "google"
        user.oauth_id = "google-sub-active"
        db.session.commit()
        self.assertIsNone(user.last_login)

        response = self.run_google_callback_with_identity({
            "sub": "google-sub-active",
            "email": "active.google@example.com",
            "email_verified": True,
            "name": "Google Active",
        })

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers.get("Location"), "/")
        self.assertEqual(self.get_session_user_id(), user.id)
        db.session.refresh(user)
        self.assertIsNotNone(user.last_login)

        dashboard_response = self.client.get("/", follow_redirects=False)
        self.assertEqual(dashboard_response.status_code, 200)

        csrf_token = self.get_csrf_token("/")
        logout_response = self.client.post("/logout", headers={"X-CSRFToken": csrf_token})
        self.assertEqual(logout_response.status_code, 302)
        self.assertIsNone(self.get_session_user_id())

    def test_google_callback_existing_local_email_is_not_auto_linked(self):
        local_user = self.create_user("local-email", full_name="Local Email")
        local_user.email = "same.email@example.com"
        local_user.email_verified = True
        local_user.auth_provider = "local"
        db.session.commit()

        response = self.run_google_callback_with_identity({
            "sub": "google-sub-local-collision",
            "email": "same.email@example.com",
            "email_verified": True,
            "name": "Collision",
        })

        self.assertEqual(response.status_code, 302)
        self.assertIn("/login", response.headers.get("Location", ""))
        self.assertIsNone(User.query.filter_by(oauth_id="google-sub-local-collision").first())
        self.assertEqual(local_user.auth_provider, "local")
        self.assertIsNone(self.get_session_user_id())

    def test_google_callback_rejects_unverified_missing_or_wrong_domain_identity(self):
        invalid_identities = [
            {"sub": "google-sub-unverified", "email": "unverified@example.com", "email_verified": False},
            {"email": "missing.sub@example.com", "email_verified": True},
            {"sub": "google-sub-missing-email", "email_verified": True},
            {"sub": "google-sub-domain", "email": "wrong@other.test", "email_verified": True},
        ]

        for identity in invalid_identities:
            before_count = User.query.count()
            response = self.run_google_callback_with_identity(identity)

            self.assertEqual(response.status_code, 302)
            self.assertIn("/login", response.headers.get("Location", ""))
            self.assertEqual(User.query.count(), before_count)
            self.assertIsNone(self.get_session_user_id())

    def test_google_auth_local_e2e_smoke_flow(self):
        # 1. Setup Approval Owner
        approval_owner = self.create_user("e2e-approval-owner", password="owner-pass", full_name="E2E Approval Owner", role="APPROVAL_OWNER")
        db.session.commit()

        # 2. Callback with new identity: creates user pending
        response = self.run_google_callback_with_identity({
            "sub": "google-sub-e2e",
            "email": "e2e.google@example.com",
            "email_verified": True,
            "name": "E2E User",
        })

        self.assertEqual(response.status_code, 302)
        self.assertIn("/auth/pending", response.headers.get("Location", ""))

        user = User.query.filter_by(auth_provider="google", oauth_id="google-sub-e2e").first()
        self.assertIsNotNone(user)
        self.assertEqual(user.approval_status, User.APPROVAL_PENDING)
        self.assertFalse(user.is_active)
        self.assertEqual(self.get_session_user_id(), user.id)

        # 3. Pending user cannot access dashboard
        dashboard_response = self.client.get("/", follow_redirects=False)
        self.assertEqual(dashboard_response.status_code, 302)
        self.assertIn("/auth/pending", dashboard_response.headers.get("Location", ""))

        # 4. Logout pending user session
        self.client.post("/logout", headers={"X-CSRFToken": self.get_csrf_token("/auth/pending")}, follow_redirects=False)
        self.assertIsNone(self.get_session_user_id())

        # 5. APPROVAL_OWNER login (local flow)
        self.login_as(approval_owner)
        self.assertEqual(self.get_session_user_id(), approval_owner.id)

        # 6. APPROVAL_OWNER opens /approval/pending and sees user
        pending_page = self.client.get("/approval/pending")
        self.assertEqual(pending_page.status_code, 200)
        html = pending_page.get_data(as_text=True)
        self.assertIn("e2e.google@example.com", html)

        # 7. APPROVAL_OWNER approves user
        approve_response = self.post_with_csrf(
            f"/approval/users/{user.id}/approve",
            path="/approval/pending",
            data={},
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(approve_response.status_code, 200)
        self.assertTrue(approve_response.get_json()["success"])

        db.session.refresh(user)
        self.assertEqual(user.approval_status, User.APPROVAL_ACTIVE)
        self.assertTrue(user.is_active)
        self.assertEqual(user.approved_by_id, approval_owner.id)
        self.assertIsNotNone(user.approved_at)

        # 8. Logout APPROVAL_OWNER
        self.client.post("/logout", headers={"X-CSRFToken": self.get_csrf_token("/approval/pending")}, follow_redirects=False)
        self.assertIsNone(self.get_session_user_id())

        # 9. Google callback with same sub logs user in
        self.assertIsNone(user.last_login)
        login_response = self.run_google_callback_with_identity({
            "sub": "google-sub-e2e",
            "email": "e2e.google@example.com",
            "email_verified": True,
            "name": "E2E User",
        })
        self.assertEqual(login_response.status_code, 302)
        self.assertEqual(login_response.headers.get("Location"), "/")
        self.assertEqual(self.get_session_user_id(), user.id)

        # last_login updated
        db.session.refresh(user)
        self.assertIsNotNone(user.last_login)

        # dashboard loads successfully (200)
        dashboard_load = self.client.get("/", follow_redirects=False)
        self.assertEqual(dashboard_load.status_code, 200)

        # 10. Logout Google user
        self.client.post("/logout", headers={"X-CSRFToken": self.get_csrf_token("/")}, follow_redirects=False)
        self.assertIsNone(self.get_session_user_id())

    def test_login_post_requires_csrf_token(self):
        self.create_user("login-csrf", password="login-pass", full_name="Login CSRF", role="STAFF")

        response = self.client.post(
            "/login",
            json={
                "username": "login-csrf",
                "password": "login-pass",
                "remember": False,
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False
        )

        self.assertEqual(response.status_code, 400)
        self.assertTrue(response.is_json)
        payload = response.get_json()
        self.assertEqual(payload["error"], "csrf_failed")
        self.assertNotIn("Location", response.headers)

    def test_login_post_succeeds_with_csrf_token(self):
        self.create_user("login-ok", password="login-pass", full_name="Login OK", role="STAFF")
        csrf_token = self.get_csrf_token("/login")

        response = self.client.post(
            "/login",
            json={
                "username": "login-ok",
                "password": "login-pass",
                "remember": False,
            },
            headers={
                "X-Requested-With": "XMLHttpRequest",
                "X-CSRFToken": csrf_token,
            },
            follow_redirects=False
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.is_json)
        payload = response.get_json()
        self.assertTrue(payload["success"])

    def test_user_can_access_app_only_when_active_and_approved(self):
        active_user = self.create_user("approval-active", password="approval-pass", full_name="Approval Active", role="STAFF")
        pending_user = self.create_user("approval-pending", password="approval-pass", full_name="Approval Pending", role="STAFF", is_active=False, approval_status="pending")
        rejected_user = self.create_user("approval-rejected", password="approval-pass", full_name="Approval Rejected", role="STAFF", is_active=False, approval_status="rejected")
        disabled_user = self.create_user("approval-disabled", password="approval-pass", full_name="Approval Disabled", role="STAFF", is_active=False, approval_status="disabled")

        self.assertTrue(active_user.can_access_app)
        self.assertTrue(active_user.is_approval_active)
        self.assertFalse(pending_user.can_access_app)
        self.assertTrue(pending_user.is_pending_approval)
        self.assertFalse(rejected_user.can_access_app)
        self.assertTrue(rejected_user.is_rejected_approval)
        self.assertFalse(disabled_user.can_access_app)
        self.assertTrue(disabled_user.is_disabled_approval)

    def test_login_blocks_pending_rejected_and_disabled_users(self):
        blocked_users = [
            ("login-pending", False, "pending", "T\u00e0i kho\u1ea3n c\u1ee7a b\u1ea1n \u0111ang ch\u1edd ch\u1ee7 spa duy\u1ec7t.", True, "/auth/pending"),
            ("login-rejected", False, "rejected", "T\u00e0i kho\u1ea3n kh\u00f4ng \u0111\u01b0\u1ee3c ph\u00e9p \u0111\u0103ng nh\u1eadp.", False, None),
            ("login-disabled", False, "disabled", "T\u00e0i kho\u1ea3n kh\u00f4ng \u0111\u01b0\u1ee3c ph\u00e9p \u0111\u0103ng nh\u1eadp.", False, None),
        ]

        for username, is_active, approval_status, expected_message, expected_pending, expected_redirect in blocked_users:
            with self.subTest(username=username):
                with self.client.session_transaction() as sess:
                    sess.clear()
                user = self.create_user(
                    username,
                    password="login-pass",
                    full_name=username.replace("-", " ").title(),
                    role="STAFF",
                    is_active=is_active,
                    approval_status=approval_status,
                )
                csrf_token = self.get_csrf_token("/login")
                response = self.client.post(
                    "/login",
                    json={
                        "username": username,
                        "password": "login-pass",
                        "remember": False,
                    },
                    headers={
                        "X-Requested-With": "XMLHttpRequest",
                        "X-CSRFToken": csrf_token,
                    },
                    follow_redirects=False,
                )

                self.assertEqual(response.status_code, 401)
                self.assertTrue(response.is_json)
                payload = response.get_json()
                self.assertFalse(payload["success"])
                self.assertEqual(payload["message"], expected_message)
                self.assertEqual(bool(payload.get("pending")), expected_pending)
                if expected_redirect:
                    self.assertEqual(payload.get("redirect"), expected_redirect)
                    with self.client.session_transaction() as sess:
                        self.assertEqual(sess[AUTH_SESSION_KEY], user.id)
                else:
                    self.assertNotIn("redirect", payload)

    def test_pending_login_sets_session_and_redirects_to_pending_page(self):
        pending_user = self.create_user(
            "login-pending-session",
            password="session-pass",
            full_name="Login Pending Session",
            role="STAFF",
            is_active=False,
            approval_status="pending",
        )
        csrf_token = self.get_csrf_token("/login")

        response = self.client.post(
            "/login",
            json={
                "username": "login-pending-session",
                "password": "session-pass",
                "remember": False,
            },
            headers={
                "X-Requested-With": "XMLHttpRequest",
                "X-CSRFToken": csrf_token,
            },
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 401)
        payload = response.get_json()
        self.assertFalse(payload["success"])
        self.assertTrue(payload["pending"])
        self.assertEqual(payload["redirect"], "/auth/pending")
        with self.client.session_transaction() as sess:
            self.assertEqual(sess[AUTH_SESSION_KEY], pending_user.id)

        pending_page = self.client.get("/auth/pending", follow_redirects=False)
        self.assertEqual(pending_page.status_code, 200)
        self.assertIn("Tài khoản của bạn đang chờ chủ spa duyệt.", pending_page.get_data(as_text=True))
        self.assertIn("Đăng xuất", pending_page.get_data(as_text=True))

    def test_pending_user_session_is_redirected_to_pending_page_from_main_app(self):
        pending_user = self.create_user(
            "session-pending",
            password="session-pass",
            full_name="Session Pending",
            role="STAFF",
            is_active=False,
            approval_status="pending",
        )
        self.login_as(pending_user)

        response = self.client.get("/", follow_redirects=False)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/auth/pending")

        login_page = self.client.get("/login", follow_redirects=False)
        self.assertEqual(login_page.status_code, 302)
        self.assertEqual(login_page.headers["Location"], "/auth/pending")

    def test_rejected_user_session_is_redirected_to_login_denial(self):
        rejected_user = self.create_user(
            "session-rejected",
            password="session-pass",
            full_name="Session Rejected",
            role="STAFF",
            is_active=False,
            approval_status="rejected",
        )
        self.login_as(rejected_user)

        response = self.client.get("/", follow_redirects=False)

        self.assertEqual(response.status_code, 302)
        self.assertIn("/login?denied=1", response.headers["Location"])

    def test_pending_page_requires_pending_session(self):
        with self.client.session_transaction() as sess:
            sess.clear()

        response = self.client.get("/auth/pending", follow_redirects=False)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/login")

    def test_login_rejects_external_next_redirect(self):
        self.create_user("login-next", password="login-pass", full_name="Login Next", role="STAFF")
        csrf_token = self.get_csrf_token("/login")

        response = self.client.post(
            "/login?next=https://evil.example",
            json={
                "username": "login-next",
                "password": "login-pass",
                "remember": False,
            },
            headers={
                "X-Requested-With": "XMLHttpRequest",
                "X-CSRFToken": csrf_token,
            },
            follow_redirects=False
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["redirect"], "/")

    def test_login_failed_attempt_is_logged_without_sensitive_data(self):
        self.create_user("login-telemetry", password="login-pass", full_name="Login Telemetry", role="STAFF")
        csrf_token = self.get_csrf_token("/login")
        request_ip = "203.0.113.10"

        with self.assertLogs("spamanager_security", level="INFO") as captured_logs:
            response = self.client.post(
                "/login",
                json={
                    "username": "login-telemetry",
                    "password": "wrong-pass",
                    "remember": False,
                },
                headers={
                    "X-Requested-With": "XMLHttpRequest",
                    "X-CSRFToken": csrf_token,
                    "X-Forwarded-For": request_ip,
                },
                follow_redirects=False
            )

        self.assertEqual(response.status_code, 401)
        self.assertTrue(response.is_json)
        payload = response.get_json()
        self.assertFalse(payload["success"])
        self.assertEqual(payload["message"], "Sai tên đăng nhập hoặc mật khẩu.")

        failed_log = ActivityLog.query.filter_by(action="AUTH_LOGIN_FAILED").order_by(ActivityLog.id.desc()).first()
        self.assertIsNotNone(failed_log)
        self.assertIn("login-telemetry", failed_log.description)
        self.assertIn(request_ip, failed_log.description)
        self.assertNotIn("wrong-pass", failed_log.description)
        self.assertTrue(any("login-telemetry" in log_entry for log_entry in captured_logs.output))
        self.assertTrue(any(request_ip in log_entry for log_entry in captured_logs.output))
        self.assertFalse(any("wrong-pass" in log_entry for log_entry in captured_logs.output))

    def test_login_rate_limit_blocks_after_threshold_and_scopes_by_username_and_ip(self):
        self.create_user("login-limit-a", password="login-pass-a", full_name="Login Limit A", role="STAFF")
        self.create_user("login-limit-b", password="login-pass-b", full_name="Login Limit B", role="STAFF")
        csrf_token = self.get_csrf_token("/login")
        blocked_ip = "203.0.113.20"
        other_ip = "203.0.113.21"

        for attempt in range(5):
            response = self.client.post(
                "/login",
                json={
                    "username": "login-limit-a",
                    "password": f"wrong-{attempt}",
                    "remember": False,
                },
                headers={
                    "X-Requested-With": "XMLHttpRequest",
                    "X-CSRFToken": csrf_token,
                    "X-Forwarded-For": blocked_ip,
                },
                follow_redirects=False
            )
            self.assertEqual(response.status_code, 401)
            self.assertEqual(response.get_json()["message"], "Sai tên đăng nhập hoặc mật khẩu.")

        rate_limited_response = self.client.post(
            "/login",
            json={
                "username": "login-limit-a",
                "password": "login-pass-a",
                "remember": False,
            },
            headers={
                "X-Requested-With": "XMLHttpRequest",
                "X-CSRFToken": csrf_token,
                "X-Forwarded-For": blocked_ip,
            },
            follow_redirects=False
        )
        self.assertEqual(rate_limited_response.status_code, 429)
        self.assertEqual(
            rate_limited_response.get_json()["message"],
            "Bạn đã đăng nhập sai quá nhiều lần. Vui lòng thử lại sau ít phút."
        )

        limited_log = ActivityLog.query.filter_by(action="AUTH_LOGIN_RATE_LIMITED").order_by(ActivityLog.id.desc()).first()
        self.assertIsNotNone(limited_log)
        self.assertIn("login-limit-a", limited_log.description)
        self.assertIn(blocked_ip, limited_log.description)

        other_user_client = app.test_client()
        other_user_token = re.search(
            r'name="csrf-token" content="([^"]+)"',
            other_user_client.get("/login").get_data(as_text=True)
        ).group(1)
        other_user_response = other_user_client.post(
            "/login",
            json={
                "username": "login-limit-b",
                "password": "login-pass-b",
                "remember": False,
            },
            headers={
                "X-Requested-With": "XMLHttpRequest",
                "X-CSRFToken": other_user_token,
                "X-Forwarded-For": blocked_ip,
            },
            follow_redirects=False
        )
        self.assertEqual(other_user_response.status_code, 200)
        self.assertTrue(other_user_response.get_json()["success"])

        different_ip_client = app.test_client()
        different_ip_token = re.search(
            r'name="csrf-token" content="([^"]+)"',
            different_ip_client.get("/login").get_data(as_text=True)
        ).group(1)
        different_ip_response = different_ip_client.post(
            "/login",
            json={
                "username": "login-limit-a",
                "password": "login-pass-a",
                "remember": False,
            },
            headers={
                "X-Requested-With": "XMLHttpRequest",
                "X-CSRFToken": different_ip_token,
                "X-Forwarded-For": other_ip,
            },
            follow_redirects=False
        )
        self.assertEqual(different_ip_response.status_code, 200)
        self.assertTrue(different_ip_response.get_json()["success"])

    def test_csrf_token_uses_compare_digest(self):
        source = inspect.getsource(csrf_module.validate_csrf_request)
        self.assertIn("compare_digest", source)

    def test_csrf_token_is_session_bound(self):
        self.create_user("bound-user", password="bound-pass", full_name="Bound User", role="STAFF")
        client_a = app.test_client()
        client_b = app.test_client()

        login_page = client_a.get("/login")
        token_a = re.search(r'name="csrf-token" content="([^"]+)"', login_page.get_data(as_text=True)).group(1)

        response = client_b.post(
            "/login",
            json={
                "username": "bound-user",
                "password": "bound-pass",
                "remember": False,
            },
            headers={
                "X-Requested-With": "XMLHttpRequest",
                "X-CSRFToken": token_a,
            },
            follow_redirects=False
        )

        self.assertEqual(response.status_code, 400)
        self.assertTrue(response.is_json)
        self.assertEqual(response.get_json()["error"], "csrf_failed")

    def test_csrf_fetch_only_adds_token_for_same_origin_unsafe_methods(self):
        source = Path("static/js/csrf.js").read_text(encoding="utf-8")
        self.assertIn("isUnsafeMethod", source)
        self.assertIn("isSameOrigin", source)
        self.assertIn("credentials = 'same-origin'", source)
        self.assertNotIn("window.fetch =", source)
        self.assertNotIn("Content-Type: 'multipart/form-data'", source)

    def test_global_csrf_protects_all_unsafe_routes(self):
        self.assertEqual(
            set(csrf_module.CSRF_SAFE_PATH_PREFIXES),
            {"/health", "/static/", "/media/"}
        )
        unsafe_methods = {"POST", "PUT", "PATCH", "DELETE"}
        for rule in app.url_map.iter_rules():
            rule_unsafe_methods = unsafe_methods.intersection(rule.methods or set())
            if not rule_unsafe_methods:
                continue
            concrete_rule = re.sub(r"<[^>]+>", "1", rule.rule)
            for method in rule_unsafe_methods:
                with app.test_request_context(concrete_rule, method=method):
                    if rule.rule.startswith(csrf_module.CSRF_SAFE_PATH_PREFIXES):
                        self.assertFalse(csrf_module.requires_csrf_protection(), rule.rule)
                    else:
                        self.assertTrue(csrf_module.requires_csrf_protection(), rule.rule)

    def test_no_state_changing_get_routes(self):
        allowed_read_only_get_routes = {
            "/settings/backup/download/<string:backup_id>",
            "/settings/restore-wizard/validate/<string:backup_id>",
            "/settings/template/customers",
            "/settings/template/services",
            "/settings/import/errors/download/<string:filename>",
            "/customers/<int:id>/can-delete",
        }
        banned_keywords = ("logout", "delete", "restore", "permanent", "import", "backup", "update", "upload", "status")

        for rule in app.url_map.iter_rules():
            if "GET" not in (rule.methods or set()):
                continue
            lower_rule = rule.rule.lower()
            if any(keyword in lower_rule for keyword in banned_keywords):
                self.assertIn(rule.rule, allowed_read_only_get_routes)

    def test_csrf_token_rotates_after_login_and_old_token_stops_working(self):
        self.create_user("rotate-user", password="rotate-pass", full_name="Rotate User", role="STAFF")
        old_token = self.get_csrf_token("/login")

        login_response = self.client.post(
            "/login",
            json={
                "username": "rotate-user",
                "password": "rotate-pass",
                "remember": False,
            },
            headers={
                "X-Requested-With": "XMLHttpRequest",
                "X-CSRFToken": old_token,
            },
            follow_redirects=False
        )
        self.assertEqual(login_response.status_code, 200)

        stale_token_response = self.client.post(
            "/customers/create",
            data={
                "name": f"Rotate Reject {uuid.uuid4().hex[:8]}",
                "phone": f"09{uuid.uuid4().int % 100000000:08d}",
                "email": f"rotate-reject-{uuid.uuid4().hex[:8]}@example.com",
                "address": "HCMC",
            },
            headers={
                "X-Requested-With": "XMLHttpRequest",
                "X-CSRFToken": old_token,
            },
            follow_redirects=False
        )
        self.assertEqual(stale_token_response.status_code, 400)
        self.assertEqual(stale_token_response.get_json()["error"], "csrf_failed")

        fresh_token = self.get_csrf_token("/customers")
        success_response = self.client.post(
            "/customers/create",
            data={
                "name": f"Rotate Accept {uuid.uuid4().hex[:8]}",
                "phone": f"09{uuid.uuid4().int % 100000000:08d}",
                "email": f"rotate-accept-{uuid.uuid4().hex[:8]}@example.com",
                "address": "HCMC",
            },
            headers={"X-CSRFToken": fresh_token},
            follow_redirects=False
        )
        self.assertEqual(success_response.status_code, 302)

    def test_logout_clears_session_and_csrf_token(self):
        self.create_user("logout-user", password="logout-pass", full_name="Logout User", role="STAFF")
        login_token = self.get_csrf_token("/login")

        login_response = self.client.post(
            "/login",
            json={
                "username": "logout-user",
                "password": "logout-pass",
                "remember": False,
            },
            headers={
                "X-Requested-With": "XMLHttpRequest",
                "X-CSRFToken": login_token,
            },
            follow_redirects=False
        )
        self.assertEqual(login_response.status_code, 200)

        logout_response = self.client.post(
            "/logout",
            headers={"X-CSRFToken": self.get_csrf_token("/customers")},
            follow_redirects=False
        )
        self.assertEqual(logout_response.status_code, 302)

        relogin_attempt = self.client.post(
            "/login",
            json={
                "username": "logout-user",
                "password": "logout-pass",
                "remember": False,
            },
            headers={
                "X-Requested-With": "XMLHttpRequest",
                "X-CSRFToken": login_token,
            },
            follow_redirects=False
        )
        self.assertEqual(relogin_attempt.status_code, 400)
        self.assertEqual(relogin_attempt.get_json()["error"], "csrf_failed")

        new_login_token = self.get_csrf_token("/login")
        self.assertNotEqual(login_token, new_login_token)

    def test_html_form_post_requires_csrf_token(self):
        owner = AuthService.seed_owner_if_empty()
        self.login_as(owner)

        response = self.client.post(
            "/customers/create",
            data={
                "name": "CSRF Customer",
                "phone": "0900000000",
                "email": "csrf@example.com",
                "address": "HCMC",
            },
            follow_redirects=False
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.content_type.split(";")[0], "text/html")
        self.assertIn("Yêu cầu không hợp lệ", response.get_data(as_text=True))
        self.assertEqual(Customer.query.filter_by(email="csrf@example.com").count(), 0)

        success_response = self.post_with_csrf(
            "/customers/create",
            path="/customers",
            data={
                "name": f"CSRF Customer {uuid.uuid4().hex[:8]}",
                "phone": f"09{uuid.uuid4().int % 100000000:08d}",
                "email": f"csrf-{uuid.uuid4().hex[:8]}@example.com",
                "address": "HCMC",
            },
            follow_redirects=False
        )

        self.assertEqual(success_response.status_code, 302)
        self.assertEqual(Customer.query.filter(Customer.name.like("CSRF Customer %")).count(), 1)

    def test_multipart_profile_update_requires_csrf_token(self):
        owner = AuthService.seed_owner_if_empty()
        self.login_as(owner)

        response = self.client.post(
            "/profile",
            data={
                "full_name": "Owner Updated",
                "avatar": (BytesIO(b"avatar-bytes"), "avatar.png"),
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            content_type="multipart/form-data",
            follow_redirects=False
        )

        self.assertEqual(response.status_code, 400)
        self.assertTrue(response.is_json)
        self.assertEqual(response.get_json()["error"], "csrf_failed")
        refreshed_owner = User.query.filter_by(username="owner").first()
        self.assertEqual(refreshed_owner.full_name, owner.full_name)

        success_response = self.post_with_csrf(
            "/profile",
            path="/profile",
            data={
                "full_name": "Owner Updated",
                "avatar": (BytesIO(b"avatar-bytes"), "avatar.png"),
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            content_type="multipart/form-data",
            follow_redirects=False
        )

        self.assertEqual(success_response.status_code, 200)
        self.assertTrue(success_response.is_json)
        self.assertTrue(success_response.get_json()["success"])

    def test_logout_get_is_safe_and_post_requires_csrf(self):
        owner = AuthService.seed_owner_if_empty()
        self.login_as(owner)

        get_response = self.client.get("/logout", follow_redirects=False)
        self.assertEqual(get_response.status_code, 405)
        still_authenticated = self.client.get("/customers", follow_redirects=False)
        self.assertEqual(still_authenticated.status_code, 200)

        csrf_token = self.get_csrf_token("/customers")
        post_response = self.client.post(
            "/logout",
            headers={"X-CSRFToken": csrf_token},
            follow_redirects=False
        )

        self.assertEqual(post_response.status_code, 302)
        self.assertIn("/login", post_response.headers.get("Location", ""))
        post_logout_access = self.client.get("/customers", follow_redirects=False)
        self.assertEqual(post_logout_access.status_code, 302)

    def test_health_check_returns_json_without_login(self):
        response = self.client.get("/health")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.headers.get("Cache-Control"), "no-store")
        self.assertTrue(response.is_json)
        self.assertEqual(response.get_json(), {
            "status": "ok",
            "app": "SpaManager",
            "database": "connected",
        })

    def test_health_check_executes_database_probe(self):
        with patch("app.db.session.execute") as mocked_execute:
            response = self.client.get("/health")

        self.assertEqual(response.status_code, 200)
        mocked_execute.assert_called_once()
        self.assertEqual(str(mocked_execute.call_args.args[0]), "SELECT 1")

    def test_health_check_returns_503_on_database_error_and_rolls_back(self):
        self.create_user("health-check-user", password="health-pass", full_name="Health Check", role="STAFF")
        original_rollback = db.session.rollback

        with patch("app.db.session.execute", side_effect=SQLAlchemyError("database unavailable")) as mocked_execute:
            with patch("app.db.session.rollback", wraps=original_rollback) as mocked_rollback:
                response = self.client.get("/health")

        self.assertEqual(response.status_code, 503)
        self.assertEqual(response.headers.get("Cache-Control"), "no-store")
        self.assertTrue(response.is_json)
        payload = response.get_json()
        self.assertEqual(payload["status"], "unhealthy")
        self.assertEqual(payload["database"], "unavailable")
        self.assertNotIn("password", response.get_data(as_text=True).lower())
        self.assertNotIn("database unavailable", response.get_data(as_text=True).lower())
        mocked_execute.assert_called_once()
        mocked_rollback.assert_called_once()
        self.assertEqual(User.query.filter_by(username="health-check-user").count(), 1)

    def test_health_check_post_method_is_not_allowed(self):
        response = self.client.post("/health")

        self.assertEqual(response.status_code, 405)
        self.assertNotIn("Location", response.headers)

    def test_health_check_session_remains_usable_after_rollback(self):
        self.create_user("session-safe", password="session-pass", full_name="Session Safe", role="STAFF")

        with patch("app.db.session.execute", side_effect=SQLAlchemyError("database unavailable")):
            response = self.client.get("/health")

        self.assertEqual(response.status_code, 503)
        query_result = User.query.filter_by(username="session-safe").first()
        self.assertIsNotNone(query_result)
        self.assertEqual(query_result.username, "session-safe")

    def test_production_requires_secret_key(self):
        with patch.dict(
            os.environ,
            {
                "DATABASE_URL": "sqlite:///prod.sqlite",
                "DEFAULT_OWNER_PASSWORD": "prod-owner-pass",
                "SECRET_KEY": "",
            },
            clear=True,
        ):
            with self.assertRaises(RuntimeError):
                ProductionConfig()

    def test_production_requires_database_url(self):
        with patch.dict(
            os.environ,
            {
                "SECRET_KEY": "prod-secret",
                "DEFAULT_OWNER_PASSWORD": "prod-owner-pass",
                "DATABASE_URL": "",
            },
            clear=True,
        ):
            with self.assertRaises(RuntimeError):
                ProductionConfig()

    def test_media_route_serves_logo_from_persistent_folder(self):
        self.create_media_file(Path("uploads") / "logos" / "sample-logo.png", b"\x89PNG\r\n\x1a\n")

        response = self.client.get("/media/logos/sample-logo.png")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "image/png")
        response.close()

    def test_missing_media_returns_404_without_redirect(self):
        response = self.client.get("/media/logos/missing.png", follow_redirects=False)

        self.assertEqual(response.status_code, 404)
        self.assertNotIn("Location", response.headers)

    def test_missing_route_returns_html_404_without_redirect_when_not_logged_in(self):
        response = self.client.get("/khong-ton-tai", follow_redirects=False)

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.content_type.split(";")[0], "text/html")
        self.assertNotIn("Location", response.headers)
        self.assertIn("404", response.get_data(as_text=True))

    def test_missing_route_returns_json_404_when_requested(self):
        response = self.client.get(
            "/khong-ton-tai-json",
            headers={"Accept": "application/json"},
            follow_redirects=False
        )

        self.assertEqual(response.status_code, 404)
        self.assertTrue(response.is_json)
        payload = response.get_json()
        self.assertEqual(payload["status"], "error")
        self.assertEqual(payload["error"], "not_found")
        self.assertNotIn("Location", response.headers)

    def test_static_missing_returns_404_without_redirect(self):
        response = self.client.get("/static/missing.css", follow_redirects=False)

        self.assertEqual(response.status_code, 404)
        self.assertNotIn("Location", response.headers)

    def test_protected_html_route_redirects_login_when_not_authenticated(self):
        response = self.client.get("/customers", follow_redirects=False)

        self.assertEqual(response.status_code, 302)
        self.assertIn("/login", response.headers.get("Location", ""))

    def test_json_auth_error_returns_401_without_redirect(self):
        response = self.client.post(
            "/change-password",
            json={
                "current_password": "x",
                "new_password": "y",
                "confirm_password": "y",
            },
            follow_redirects=False
        )

        self.assertEqual(response.status_code, 401)
        self.assertTrue(response.is_json)
        payload = response.get_json()
        self.assertEqual(payload["status"], "error")
        self.assertEqual(payload["error"], "unauthorized")
        self.assertNotIn("Location", response.headers)

    def test_password_policy_rules_are_centralized(self):
        empty_result = PasswordPolicy.validate_password("", require_confirm=False)
        whitespace_result = PasswordPolicy.validate_password("        ", require_confirm=False)
        short_result = PasswordPolicy.validate_password("short7", require_confirm=False)
        valid_result = PasswordPolicy.validate_password("validpass", require_confirm=False)
        mismatch_result = PasswordPolicy.validate_password(
            "validpass",
            confirm_password="different",
            require_confirm=True,
        )
        reuse_result = PasswordPolicy.validate_password(
            "validpass",
            current_password="validpass",
            require_confirm=False,
            prevent_reuse=True,
        )

        self.assertFalse(empty_result.valid)
        self.assertEqual(empty_result.message, "Mật khẩu không được để trống.")
        self.assertFalse(whitespace_result.valid)
        self.assertEqual(whitespace_result.message, "Mật khẩu không được để trống.")
        self.assertFalse(short_result.valid)
        self.assertEqual(short_result.message, "Mật khẩu mới phải có ít nhất 8 ký tự.")
        self.assertTrue(valid_result.valid)
        self.assertFalse(mismatch_result.valid)
        self.assertEqual(mismatch_result.message, "Xác nhận mật khẩu không khớp.")
        self.assertFalse(reuse_result.valid)
        self.assertEqual(reuse_result.message, "Mật khẩu mới không được giống mật khẩu hiện tại.")

    def test_change_password_route_enforces_shared_policy(self):
        self.create_user("policy-change", password="old-pass-123", full_name="Policy Change", role="STAFF")
        login_token = self.get_csrf_token("/login")
        login_response = self.client.post(
            "/login",
            json={
                "username": "policy-change",
                "password": "old-pass-123",
                "remember": False,
            },
            headers={
                "X-Requested-With": "XMLHttpRequest",
                "X-CSRFToken": login_token,
            },
            follow_redirects=False
        )
        self.assertEqual(login_response.status_code, 200)

        short_response = self.post_with_csrf(
            "/change-password",
            path="/customers",
            json={
                "current_password": "old-pass-123",
                "new_password": "short",
                "confirm_password": "short",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(short_response.status_code, 400)
        self.assertEqual(short_response.get_json()["message"], "Mật khẩu mới phải có ít nhất 8 ký tự.")

        mismatch_response = self.post_with_csrf(
            "/change-password",
            path="/customers",
            json={
                "current_password": "old-pass-123",
                "new_password": "new-pass-456",
                "confirm_password": "new-pass-789",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(mismatch_response.status_code, 400)
        self.assertEqual(mismatch_response.get_json()["message"], "Xác nhận mật khẩu không khớp.")

        same_password_response = self.post_with_csrf(
            "/change-password",
            path="/customers",
            json={
                "current_password": "old-pass-123",
                "new_password": "old-pass-123",
                "confirm_password": "old-pass-123",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(same_password_response.status_code, 400)
        self.assertEqual(same_password_response.get_json()["message"], "Mật khẩu mới không được giống mật khẩu hiện tại.")

        success_response = self.post_with_csrf(
            "/change-password",
            path="/customers",
            json={
                "current_password": "old-pass-123",
                "new_password": "new-pass-456",
                "confirm_password": "new-pass-456",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(success_response.status_code, 200)
        self.assertTrue(success_response.get_json()["success"])

        self.client.post("/logout", headers={"X-CSRFToken": self.get_csrf_token("/customers")}, follow_redirects=False)
        relogin_token = self.get_csrf_token("/login")
        old_login = self.client.post(
            "/login",
            json={"username": "policy-change", "password": "old-pass-123", "remember": False},
            headers={"X-Requested-With": "XMLHttpRequest", "X-CSRFToken": relogin_token},
            follow_redirects=False,
        )
        self.assertEqual(old_login.status_code, 401)
        new_login = self.client.post(
            "/login",
            json={"username": "policy-change", "password": "new-pass-456", "remember": False},
            headers={"X-Requested-With": "XMLHttpRequest", "X-CSRFToken": relogin_token},
            follow_redirects=False,
        )
        self.assertEqual(new_login.status_code, 200)
        self.assertTrue(new_login.get_json()["success"])

        change_log = ActivityLog.query.filter_by(action="CHANGE_PASSWORD").order_by(ActivityLog.id.desc()).first()
        self.assertIsNotNone(change_log)
        self.assertNotIn("old-pass-123", change_log.description)
        self.assertNotIn("new-pass-456", change_log.description)

    def test_admin_reset_password_route_uses_shared_policy(self):
        owner = self.create_user("policy-owner", password="owner-pass", full_name="Policy Owner", role="OWNER")
        target = self.create_user("policy-target", password="target-pass", full_name="Policy Target", role="STAFF")
        self.login_as(owner)

        short_response = self.post_with_csrf(
            f"/users/{target.id}/reset-password",
            path="/users",
            data={"new_password": "short", "confirm_password": "short"},
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(short_response.status_code, 400)
        self.assertEqual(short_response.get_json()["message"], "Mật khẩu mới phải có ít nhất 8 ký tự.")

        mismatch_response = self.post_with_csrf(
            f"/users/{target.id}/reset-password",
            path="/users",
            data={"new_password": "target-new-pass", "confirm_password": "different-pass"},
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(mismatch_response.status_code, 400)
        self.assertEqual(mismatch_response.get_json()["message"], "Xác nhận mật khẩu không khớp.")

        success_response = self.post_with_csrf(
            f"/users/{target.id}/reset-password",
            path="/users",
            data={"new_password": "target-new-pass", "confirm_password": "target-new-pass"},
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(success_response.status_code, 200)
        self.assertTrue(success_response.get_json()["success"])

        reset_log = ActivityLog.query.filter_by(action="RESET_USER_PASSWORD").order_by(ActivityLog.id.desc()).first()
        self.assertIsNotNone(reset_log)
        self.assertEqual(reset_log.user_id, owner.id)
        self.assertNotIn("target-new-pass", reset_log.description)

        self.client.post("/logout", headers={"X-CSRFToken": self.get_csrf_token("/users")}, follow_redirects=False)
        login_token = self.get_csrf_token("/login")
        target_login = self.client.post(
            "/login",
            json={"username": "policy-target", "password": "target-new-pass", "remember": False},
            headers={"X-Requested-With": "XMLHttpRequest", "X-CSRFToken": login_token},
            follow_redirects=False,
        )
        self.assertEqual(target_login.status_code, 200)
        self.assertTrue(target_login.get_json()["success"])

    def test_create_user_route_uses_shared_policy(self):
        owner = self.create_user("policy-create-owner", password="owner-pass", full_name="Policy Create Owner", role="OWNER")
        self.login_as(owner)

        short_response = self.post_with_csrf(
            "/users/create",
            path="/users",
            data={
                "username": "policy-short",
                "full_name": "Policy Short",
                "email": "policy-short@example.com",
                "role": "STAFF",
                "is_active": "1",
                "password": "short",
                "confirm_password": "short",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(short_response.status_code, 400)
        self.assertEqual(short_response.get_json()["message"], "Mật khẩu mới phải có ít nhất 8 ký tự.")

        success_response = self.post_with_csrf(
            "/users/create",
            path="/users",
            data={
                "username": "policy-create",
                "full_name": "Policy Create",
                "email": "policy-create@example.com",
                "role": "STAFF",
                "is_active": "1",
                "password": "create-pass-123",
                "confirm_password": "create-pass-123",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(success_response.status_code, 200)
        self.assertTrue(success_response.get_json()["success"])

        created_user = User.query.filter_by(username="policy-create").first()
        self.assertIsNotNone(created_user)
        create_log = ActivityLog.query.filter_by(action="CREATE_USER").order_by(ActivityLog.id.desc()).first()
        self.assertIsNotNone(create_log)
        self.assertNotIn("create-pass-123", create_log.description)

    def test_user_management_access_control_and_staff_blocked_routes(self):
        owner = self.create_user("users-access-owner", password="owner-pass", full_name="Users Access Owner", role="OWNER")
        admin = self.create_user("users-access-admin", password="admin-pass", full_name="Users Access Admin", role="ADMIN")
        staff = self.create_user("users-access-staff", password="staff-pass", full_name="Users Access Staff", role="STAFF")
        target = self.create_user("users-access-target", password="target-pass", full_name="Users Access Target", role="STAFF")

        for manager in (owner, admin):
            self.login_as(manager)
            response = self.client.get("/users")
            self.assertEqual(response.status_code, 200)
            self.assertIn("Người dùng", response.get_data(as_text=True))
            self.client.post("/logout", headers={"X-CSRFToken": self.get_csrf_token("/users")}, follow_redirects=False)

        self.login_as(staff)
        self.assertEqual(self.client.get("/users").status_code, 403)

        create_response = self.post_with_csrf(
            "/users/create",
            path="/users",
            data={
                "username": "users-access-new",
                "full_name": "Users Access New",
                "email": "users-access-new@example.com",
                "role": "STAFF",
                "is_active": "1",
                "password": "users-access-pass",
                "confirm_password": "users-access-pass",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(create_response.status_code, 403)
        self.assertIsNone(User.query.filter_by(username="users-access-new").first())

        edit_response = self.post_with_csrf(
            f"/users/{target.id}/edit",
            path="/users",
            data={
                "username": "users-access-target",
                "full_name": "Users Access Target",
                "email": "users-access-target@example.com",
                "role": "STAFF",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(edit_response.status_code, 403)

        reset_response = self.post_with_csrf(
            f"/users/{target.id}/reset-password",
            path="/users",
            data={"new_password": "users-access-new-pass", "confirm_password": "users-access-new-pass"},
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(reset_response.status_code, 403)

        toggle_response = self.post_with_csrf(
            f"/users/{target.id}/toggle-active",
            path="/users",
            data={"is_active": "0"},
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(toggle_response.status_code, 403)
        self.assertTrue(User.query.get(target.id).is_active)

    def test_approval_owner_pending_users_page_and_approval_actions_work(self):
        approval_owner = self.create_user("pending-approval-owner", password="owner-pass", full_name="Pending Approval Owner", role="APPROVAL_OWNER")
        admin = self.create_user("pending-admin", password="admin-pass", full_name="Pending Admin", role="ADMIN")
        pending_approve = self.create_user(
            "pending-approve",
            password="pending-pass",
            full_name="Pending Approve",
            role="STAFF",
            is_active=False,
            approval_status="pending",
        )
        pending_reject = self.create_user(
            "pending-reject",
            password="pending-pass",
            full_name="Pending Reject",
            role="STAFF",
            is_active=False,
            approval_status="pending",
        )
        active_user = self.create_user(
            "active-user",
            password="active-pass",
            full_name="Active User",
            role="STAFF",
            is_active=True,
            approval_status="active",
        )

        self.login_as(approval_owner)
        page = self.client.get("/approval/pending")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn("Danh sách tài khoản chờ duyệt", html)
        self.assertIn("pending-approve", html)
        self.assertIn("pending-reject", html)
        self.assertNotIn("active-user", html)

        # OWNER/ADMIN list should not mix in pending approval UI
        self.assertNotIn("pending-approval-owner", html)

        # Standard users list (/users) should not show "Tài khoản chờ duyệt" button
        self.login_as(admin)
        self.assertNotIn("Tài khoản chờ duyệt", self.client.get("/users").get_data(as_text=True))

        self.login_as(approval_owner)
        approve_response = self.post_with_csrf(
            f"/approval/users/{pending_approve.id}/approve",
            path="/approval/pending",
            data={},
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(approve_response.status_code, 200)
        approve_payload = approve_response.get_json()
        self.assertTrue(approve_payload["success"])
        approved_user = User.query.get(pending_approve.id)
        self.assertEqual(approved_user.approval_status, "active")
        self.assertTrue(approved_user.is_active)
        self.assertEqual(approved_user.approved_by_id, approval_owner.id)
        self.assertIsNotNone(approved_user.approved_at)

        self.client.post("/logout", headers={"X-CSRFToken": self.get_csrf_token("/approval/pending")}, follow_redirects=False)
        login_token = self.get_csrf_token("/login")
        approved_login = self.client.post(
            "/login",
            json={"username": "pending-approve", "password": "pending-pass", "remember": False},
            headers={"X-Requested-With": "XMLHttpRequest", "X-CSRFToken": login_token},
            follow_redirects=False,
        )
        self.assertEqual(approved_login.status_code, 200)
        self.assertTrue(approved_login.get_json()["success"])

        self.login_as(approval_owner)
        reject_response = self.post_with_csrf(
            f"/approval/users/{pending_reject.id}/reject",
            path="/approval/pending",
            data={},
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(reject_response.status_code, 200)
        reject_payload = reject_response.get_json()
        self.assertTrue(reject_payload["success"])
        rejected_user = User.query.get(pending_reject.id)
        self.assertEqual(rejected_user.approval_status, "rejected")
        self.assertFalse(rejected_user.is_active)
        self.assertIsNone(rejected_user.approved_by_id)
        self.assertIsNone(rejected_user.approved_at)

        self.client.post("/logout", headers={"X-CSRFToken": self.get_csrf_token("/approval/pending")}, follow_redirects=False)
        login_token = self.get_csrf_token("/login")
        rejected_login = self.client.post(
            "/login",
            json={"username": "pending-reject", "password": "pending-pass", "remember": False},
            headers={"X-Requested-With": "XMLHttpRequest", "X-CSRFToken": login_token},
            follow_redirects=False,
        )
        self.assertEqual(rejected_login.status_code, 401)
        self.assertEqual(rejected_login.get_json()["message"], "Tài khoản không được phép đăng nhập.")

    def test_non_approval_owners_blocked_from_approval_portal_and_approval_owner_blocked_from_spa(self):
        owner = self.create_user("pending-owner-guard", password="owner-pass", full_name="Pending Owner Guard", role="OWNER")
        admin = self.create_user("pending-admin-guard", password="admin-pass", full_name="Pending Admin Guard", role="ADMIN")
        staff = self.create_user("pending-staff-guard", password="staff-pass", full_name="Pending Staff Guard", role="STAFF")
        approval_owner = self.create_user("pending-approval-guard", password="owner-pass", full_name="Pending Approval Guard", role="APPROVAL_OWNER")

        # 1. Non-approval owners blocked from /approval/pending (403)
        self.login_as(owner)
        self.assertEqual(self.client.get("/approval/pending").status_code, 403)
        self.client.post("/logout", headers={"X-CSRFToken": self.get_csrf_token("/users")}, follow_redirects=False)

        self.login_as(admin)
        self.assertEqual(self.client.get("/approval/pending").status_code, 403)
        self.client.post("/logout", headers={"X-CSRFToken": self.get_csrf_token("/users")}, follow_redirects=False)

        self.login_as(staff)
        self.assertEqual(self.client.get("/approval/pending").status_code, 403)
        self.client.post("/logout", headers={"X-CSRFToken": self.get_csrf_token("/users")}, follow_redirects=False)

        # 2. Approval Owner is blocked from accessing SpaManager routes (like / or /users/pending)
        # and redirected back to /approval/pending
        self.login_as(approval_owner)
        self.assertEqual(self.client.get("/users/pending").status_code, 302)
        self.assertEqual(self.client.get("/users/pending").headers.get("Location"), "/approval/pending")

        self.assertEqual(self.client.get("/").status_code, 302)
        self.assertEqual(self.client.get("/").headers.get("Location"), "/approval/pending")

    def test_user_create_route_rejects_invalid_password_and_duplicate_identity(self):
        owner = self.create_user("users-create-owner", password="owner-pass", full_name="Users Create Owner", role="OWNER")
        existing_user = self.create_user("users-create-existing", password="existing-pass", full_name="Users Create Existing", role="STAFF",)
        existing_user.email = "users-create-existing@example.com"
        db.session.commit()
        self.login_as(owner)

        short_response = self.post_with_csrf(
            "/users/create",
            path="/users",
            data={
                "username": "users-create-short",
                "full_name": "Users Create Short",
                "email": "users-create-short@example.com",
                "role": "STAFF",
                "is_active": "1",
                "password": "short",
                "confirm_password": "short",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(short_response.status_code, 400)
        self.assertFalse(User.query.filter_by(username="users-create-short").first())

        mismatch_response = self.post_with_csrf(
            "/users/create",
            path="/users",
            data={
                "username": "users-create-mismatch",
                "full_name": "Users Create Mismatch",
                "email": "users-create-mismatch@example.com",
                "role": "STAFF",
                "is_active": "1",
                "password": "users-create-pass-123",
                "confirm_password": "different-pass",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(mismatch_response.status_code, 400)
        self.assertFalse(User.query.filter_by(username="users-create-mismatch").first())

        duplicate_username_response = self.post_with_csrf(
            "/users/create",
            path="/users",
            data={
                "username": "users-create-existing",
                "full_name": "Users Create Duplicate Username",
                "email": "users-create-dup-user@example.com",
                "role": "STAFF",
                "is_active": "1",
                "password": "users-create-pass-123",
                "confirm_password": "users-create-pass-123",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(duplicate_username_response.status_code, 400)
        self.assertFalse(User.query.filter_by(email="users-create-dup-user@example.com").first())

        duplicate_email_response = self.post_with_csrf(
            "/users/create",
            path="/users",
            data={
                "username": "users-create-dup-email",
                "full_name": "Users Create Duplicate Email",
                "email": existing_user.email,
                "role": "STAFF",
                "is_active": "1",
                "password": "users-create-pass-123",
                "confirm_password": "users-create-pass-123",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(duplicate_email_response.status_code, 400)
        self.assertFalse(User.query.filter_by(username="users-create-dup-email").first())

    def test_user_edit_route_updates_profile_and_rejects_invalid_role(self):
        owner = self.create_user("users-edit-owner", password="owner-pass", full_name="Users Edit Owner", role="OWNER")
        target = self.create_user("users-edit-target", password="target-pass", full_name="Users Edit Target", role="STAFF")
        original_password_hash = target.password_hash
        self.login_as(owner)

        success_response = self.post_with_csrf(
            f"/users/{target.id}/edit",
            path="/users",
            data={
                "username": "users-edit-target",
                "full_name": "Users Edit Target Updated",
                "email": "users-edit-target@example.com",
                "role": "ADMIN",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(success_response.status_code, 200)
        self.assertTrue(success_response.get_json()["success"])

        updated_user = User.query.get(target.id)
        self.assertEqual(updated_user.full_name, "Users Edit Target Updated")
        self.assertEqual(updated_user.email, "users-edit-target@example.com")
        self.assertEqual(updated_user.role, "ADMIN")
        self.assertEqual(updated_user.password_hash, original_password_hash)

        update_log = ActivityLog.query.filter_by(action="UPDATE_USER").order_by(ActivityLog.id.desc()).first()
        self.assertIsNotNone(update_log)
        self.assertEqual(update_log.user_id, owner.id)
        self.assertNotIn("target-pass", update_log.description)
        self.assertNotIn("password", update_log.description.lower())

        invalid_role_response = self.post_with_csrf(
            f"/users/{target.id}/edit",
            path="/users",
            data={
                "username": "users-edit-target",
                "full_name": "Users Edit Target Updated Again",
                "email": "users-edit-target@example.com",
                "role": "SUPERADMIN",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(invalid_role_response.status_code, 400)
        self.assertEqual(invalid_role_response.get_json()["message"], "Vai trò người dùng không hợp lệ.")
        self.assertEqual(User.query.get(target.id).role, "ADMIN")

    def test_user_edit_route_blocks_self_demote_from_manager_role(self):
        owner = self.create_user("users-edit-self-owner", password="owner-pass", full_name="Users Edit Self Owner", role="OWNER")
        self.login_as(owner)

        response = self.post_with_csrf(
            f"/users/{owner.id}/edit",
            path="/users",
            data={
                "username": "users-edit-self-owner",
                "full_name": "Users Edit Self Owner",
                "email": "users-edit-self-owner@example.com",
                "role": "STAFF",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.get_json()["message"], "Không thể tự hạ quyền quản trị của chính mình.")
        self.assertEqual(User.query.get(owner.id).role, "OWNER")

    def test_user_toggle_active_route_enforces_state_changes_and_blocks_self_disable(self):
        owner = self.create_user("users-toggle-owner", password="owner-pass", full_name="Users Toggle Owner", role="OWNER")
        target = self.create_user("users-toggle-target", password="target-pass", full_name="Users Toggle Target", role="STAFF")
        self.login_as(owner)

        deactivate_response = self.post_with_csrf(
            f"/users/{target.id}/toggle-active",
            path="/users",
            data={"is_active": "0"},
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(deactivate_response.status_code, 200)
        self.assertFalse(User.query.get(target.id).is_active)

        self.client.post("/logout", headers={"X-CSRFToken": self.get_csrf_token("/users")}, follow_redirects=False)
        login_token = self.get_csrf_token("/login")
        inactive_login = self.client.post(
            "/login",
            json={"username": "users-toggle-target", "password": "target-pass", "remember": False},
            headers={"X-Requested-With": "XMLHttpRequest", "X-CSRFToken": login_token},
            follow_redirects=False,
        )
        self.assertEqual(inactive_login.status_code, 401)
        self.assertFalse(inactive_login.is_json and inactive_login.get_json().get("success", False))

        self.login_as(owner)
        reactivate_response = self.post_with_csrf(
            f"/users/{target.id}/toggle-active",
            path="/users",
            data={"is_active": "1"},
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(reactivate_response.status_code, 200)
        self.assertTrue(User.query.get(target.id).is_active)

        self.client.post("/logout", headers={"X-CSRFToken": self.get_csrf_token("/users")}, follow_redirects=False)
        self.login_as(owner)
        self_disable_response = self.post_with_csrf(
            f"/users/{owner.id}/toggle-active",
            path="/users",
            data={"is_active": "0"},
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(self_disable_response.status_code, 400)
        self.assertEqual(self_disable_response.get_json()["message"], "Không thể vô hiệu hóa chính mình.")
        self.assertTrue(User.query.get(owner.id).is_active)

    def test_html_500_renders_template_and_rolls_back(self):
        owner = AuthService.seed_owner_if_empty()
        self.login_as(owner)

        def explode_dashboard():
            raise RuntimeError("boom")

        original_view = app.view_functions["dashboard.index"]
        app.view_functions["dashboard.index"] = explode_dashboard
        try:
            response = self.client.get("/", follow_redirects=False)
        finally:
            app.view_functions["dashboard.index"] = original_view

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.content_type.split(";")[0], "text/html")
        self.assertIn("500", response.get_data(as_text=True))
        self.assertNotIn("boom", response.get_data(as_text=True))
        self.assertNotIn("Traceback", response.get_data(as_text=True))

    def test_json_500_returns_json_and_rolls_back(self):
        owner = AuthService.seed_owner_if_empty()
        self.login_as(owner)

        def explode_dashboard():
            raise RuntimeError("boom-json")

        original_view = app.view_functions["dashboard.index"]
        app.view_functions["dashboard.index"] = explode_dashboard
        try:
            response = self.client.get("/", headers={"Accept": "application/json"}, follow_redirects=False)
        finally:
            app.view_functions["dashboard.index"] = original_view

        self.assertEqual(response.status_code, 500)
        self.assertTrue(response.is_json)
        payload = response.get_json()
        self.assertEqual(payload["status"], "error")
        self.assertEqual(payload["error"], "internal_server_error")
        self.assertNotIn("boom-json", response.get_data(as_text=True))
        self.assertNotIn("Traceback", response.get_data(as_text=True))

    def test_dashboard_today_appointments_render_expected_classes(self):
        owner = AuthService.seed_owner_if_empty()
        self.login_as(owner)

        response = self.client.get("/")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('id="dashboard-schedule-list"', html)
        self.assertIn('class="schedule-list"', html)
        self.assertIn('class="appointment-item schedule-item"', html)
        self.assertIn('id="appt-footer"', html)

    def test_dashboard_manager_sees_admin_summary_and_backup_status(self):
        owner = AuthService.seed_owner_if_empty()
        self.login_as(owner)

        response = self.client.get("/")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("Tổng quan quản trị", html)
        self.assertIn("Backup gần nhất", html)
        self.assertIn("Người dùng", html)
        self.assertIn("Cảnh báo gần đây", html)
        self.assertIn("Phím tắt quản trị", html)
        self.assertNotIn(os.path.abspath("backup"), html)

        api_response = self.client.get("/api/dashboard/data")
        payload = api_response.get_json()
        self.assertIn("admin_summary", payload)
        self.assertIn("backup", payload["admin_summary"])
        self.assertIn("users", payload["admin_summary"])
        self.assertIn("activity", payload["admin_summary"])

    def test_dashboard_manager_shows_latest_backup_summary_when_available(self):
        owner = self.create_user("dashboard-backup-owner", password="owner-pass", full_name="Dashboard Backup Owner", role="OWNER")
        self.login_as(owner)
        backup_id, backup_meta, backup_path = self.create_settings_backup_via_route(owner, notes="Dashboard summary backup")

        try:
            response = self.client.get("/")
            html = response.get_data(as_text=True)

            self.assertEqual(response.status_code, 200)
            self.assertIn("Backup gần nhất", html)
            self.assertNotIn(str(backup_path), html)
        finally:
            if backup_path.exists():
                backup_path.unlink()
            BackupRepository.delete(app, backup_id)

    def test_dashboard_staff_hides_admin_summary_and_api_does_not_leak_sensitive_blocks(self):
        owner = AuthService.seed_owner_if_empty()
        staff = self.create_user("dashboard-staff", password="staff-pass", full_name="Dashboard Staff", role="STAFF")
        self.login_as(owner)
        backup_id, backup_meta, backup_path = self.create_settings_backup_via_route(owner, notes="Staff visibility backup")

        try:
            self.login_as(staff)
            response = self.client.get("/")
            html = response.get_data(as_text=True)
            self.assertEqual(response.status_code, 200)
            self.assertIn("Lịch hẹn hôm nay", html)
            self.assertIn("Doanh thu hôm nay", html)
            self.assertNotIn("Tổng quan quản trị", html)
            self.assertNotIn("Phím tắt quản trị", html)
            self.assertNotIn("Backup gần nhất", html)
            self.assertNotIn("Người dùng", html)

            api_response = self.client.get("/api/dashboard/data")
            payload = api_response.get_json()
            self.assertNotIn("admin_summary", payload)
            self.assertNotIn("recent_activities", payload)
        finally:
            if backup_path.exists():
                backup_path.unlink()
            BackupRepository.delete(app, backup_id)

    def test_topbar_exposes_command_palette_hint(self):
        owner = AuthService.seed_owner_if_empty()
        self.login_as(owner)

        html = self.client.get("/").get_data(as_text=True)
        source = Path("static/js/command-palette.js").read_text(encoding="utf-8")
        palette_html = Path("templates/layout/command_palette.html").read_text(encoding="utf-8")

        self.assertIn('data-command-palette-open', html)
        self.assertIn('type="button"', html)
        self.assertIn('Tìm nhanh', html)
        self.assertIn('Ctrl K', html)
        self.assertIn('data-command-palette-close', palette_html)
        self.assertIn('aria-label="Đóng tìm nhanh"', palette_html)
        self.assertIn('command-palette-close', palette_html)
        self.assertIn('window.openCommandPalette = function (triggerEl)', source)
        self.assertIn('window.closeCommandPalette = function ()', source)
        self.assertIn('[data-command-palette-open]', source)
        self.assertIn('[data-command-palette-close]', source)

    def test_page_size_handler_is_scoped_to_explicit_controls(self):
        source = Path("static/js/shared-table.js").read_text(encoding="utf-8")
        macro = Path("templates/layout/table_macros.html").read_text(encoding="utf-8")
        invoice_template = Path("templates/invoice/index.html").read_text(encoding="utf-8")
        appointment_template = Path("templates/appointment/index.html").read_text(encoding="utf-8")
        activity_log_template = Path("templates/activity_log/index.html").read_text(encoding="utf-8")
        recycle_bin_template = Path("templates/recycle_bin/index.html").read_text(encoding="utf-8")

        self.assertIn("[data-stf-filter]", source)
        self.assertIn("[data-stf-per-page]", source)
        self.assertIn("fetchAndSwapPageSize", source)
        self.assertIn("window.history.replaceState", source)
        self.assertIn("data-stf-per-page-param", macro)
        self.assertNotIn("reloadWithParams(p, pageParam)", source)
        self.assertNotIn("window.location.href = u.toString()", invoice_template)
        self.assertNotIn("window.location.href = u.toString()", appointment_template)
        self.assertNotIn("window.location.href = u.toString()", activity_log_template)
        self.assertNotIn("window.location.href = u.toString()", recycle_bin_template)

    def test_seed_owner_creates_owner_when_database_is_empty(self):
        owner = AuthService.seed_owner_if_empty()

        self.assertIsNotNone(owner)
        self.assertEqual(owner.username, "owner")
        self.assertEqual(User.query.filter_by(username="owner").count(), 1)
        self.assertTrue(User.query.filter_by(username="owner").first().check_password("owner123"))

    def test_new_user_defaults_to_active_approval_state(self):
        user = self.create_user("approval-default", password="approval-pass", full_name="Approval Default", role="STAFF")

        refreshed_user = User.query.get(user.id)
        self.assertIsNotNone(refreshed_user)
        self.assertEqual(refreshed_user.approval_status, "active")
        self.assertIsNone(refreshed_user.approved_by_id)
        self.assertIsNone(refreshed_user.approved_at)
        self.assertTrue(refreshed_user.is_active)

    def test_seed_owner_does_not_change_existing_owner(self):
        existing_owner = self.create_user("owner", password="old-password", full_name="Chá»§ Spa", role="OWNER")
        existing_hash = existing_owner.password_hash

        result = AuthService.seed_owner_if_empty()

        refreshed_owner = User.query.filter_by(username="owner").first()
        self.assertEqual(result.id, existing_owner.id)
        self.assertEqual(User.query.filter_by(username="owner").count(), 1)
        self.assertEqual(refreshed_owner.password_hash, existing_hash)
        self.assertTrue(refreshed_owner.check_password("old-password"))

    def test_seed_owner_creates_owner_when_other_user_exists(self):
        self.create_user("customer-1", password="customer-pass", full_name="Customer 1", role="STAFF")

        owner = AuthService.seed_owner_if_empty()

        self.assertIsNotNone(owner)
        self.assertEqual(owner.username, "owner")
        self.assertEqual(User.query.filter_by(username="owner").count(), 1)
        self.assertEqual(User.query.count(), 2)

    def test_seed_owner_recovers_from_integrity_error_when_owner_appears(self):
        self.create_user("observer", password="observer-pass", full_name="Observer", role="STAFF")

        def inject_owner_before_commit(session):
            self.insert_owner_row()

        event.listen(db.session, "before_commit", inject_owner_before_commit)
        try:
            result = AuthService.seed_owner_if_empty()
        finally:
            event.remove(db.session, "before_commit", inject_owner_before_commit)

        self.assertEqual(result.username, "owner")
        self.assertEqual(User.query.filter_by(username="owner").count(), 1)
        self.assertEqual(User.query.filter_by(username="observer").first().username, "observer")
        self.assertEqual(User.query.count(), 2)

    def test_seed_owner_raises_when_integrity_error_and_owner_still_missing(self):
        self.create_user("observer-2", password="observer-pass", full_name="Observer 2", role="STAFF")
        original_add = db.session.add
        duplicate_added = False

        def add_duplicate_owner(obj):
            nonlocal duplicate_added
            original_add(obj)
            if isinstance(obj, User) and obj.username == "owner" and not duplicate_added:
                duplicate_added = True
                duplicate_owner = User(
                    username="owner",
                    full_name="Chá»§ Spa",
                    role="OWNER",
                    is_active=True,
                )
                duplicate_owner.set_password("duplicate-pass")
                original_add(duplicate_owner)

        with patch("services.auth_service.db.session.add", side_effect=add_duplicate_owner):
            with self.assertRaises(IntegrityError):
                AuthService.seed_owner_if_empty()

        self.assertIsNone(User.query.filter_by(username="owner").first())
        self.assertEqual(User.query.filter_by(username="observer-2").first().username, "observer-2")
        self.assertEqual(User.query.count(), 1)

    def test_seed_owner_session_still_queryable_after_rollback(self):
        self.create_user("session-check", password="session-pass", full_name="Session Check", role="STAFF")
        original_add = db.session.add
        duplicate_added = False

        def add_duplicate_owner(obj):
            nonlocal duplicate_added
            original_add(obj)
            if isinstance(obj, User) and obj.username == "owner" and not duplicate_added:
                duplicate_added = True
                duplicate_owner = User(
                    username="owner",
                    full_name="Chá»§ Spa",
                    role="OWNER",
                    is_active=True,
                )
                duplicate_owner.set_password("duplicate-pass")
                original_add(duplicate_owner)

        with patch("services.auth_service.db.session.add", side_effect=add_duplicate_owner):
            with self.assertRaises(IntegrityError):
                AuthService.seed_owner_if_empty()

        query_result = User.query.filter_by(username="session-check").first()
        self.assertIsNotNone(query_result)
        self.assertEqual(query_result.username, "session-check")

    def test_seed_owner_multiple_calls_keep_single_owner(self):
        AuthService.seed_owner_if_empty()
        AuthService.seed_owner_if_empty()
        AuthService.seed_owner_if_empty()

        self.assertEqual(User.query.filter_by(username="owner").count(), 1)
        self.assertEqual(User.query.count(), 1)

    def test_customer_delete_records_deleted_by(self):
        owner = AuthService.seed_owner_if_empty()
        customer = self.create_customer_record("Audit Customer")
        self.login_as(owner)

        response = self.post_with_csrf(f"/customers/{customer.id}/delete", path="/customers")

        self.assertEqual(response.status_code, 302)
        deleted_customer = Customer.query.get(customer.id)
        self.assertIsNotNone(deleted_customer)
        self.assertEqual(deleted_customer.deleted_by, "owner")

    def test_customer_restore_clears_deleted_by(self):
        owner = AuthService.seed_owner_if_empty()
        customer = self.create_customer_record("Restore Customer")
        self.login_as(owner)
        self.post_with_csrf(f"/customers/{customer.id}/delete", path="/customers")

        response = self.post_with_csrf(f"/recycle-bin/restore/Customer/{customer.id}", path="/recycle-bin")

        self.assertEqual(response.status_code, 200)
        restored_customer = Customer.query.get(customer.id)
        self.assertIsNotNone(restored_customer)
        self.assertIsNone(restored_customer.deleted_at)
        self.assertIsNone(restored_customer.deleted_by)

    def test_customer_create_blocks_duplicate_active_phone_and_email_normalization(self):
        CustomerService.create(name="Customer A", phone="0901234567", email="Test@Email.com", address="Address A")

        errors = CustomerService.check_duplicate(phone=" 0901234567 ", email="test@email.com", include_deleted=True)
        self.assertIn("phone", errors)
        self.assertIn("email", errors)

        with self.assertRaises(ConflictException) as context:
            CustomerService.create(name="Customer B", phone=" 0901234567 ", email="test@email.com", address="Address B")

        self.assertIn("Email", str(context.exception))
        self.assertEqual(Customer.query.filter(Customer.deleted_at.is_(None)).count(), 1)

    def test_customer_create_blocks_soft_deleted_duplicate_reference(self):
        customer = Customer(name="Soft Deleted Customer", phone="0908888888", email="trash@example.com", deleted_at=datetime.utcnow(), deleted_by="owner")
        db.session.add(customer)
        db.session.commit()

        errors = CustomerService.check_duplicate(phone=" 0908888888 ", email="TRASH@example.com", include_deleted=True)
        self.assertIn("phone", errors)
        self.assertIn("email", errors)

        with self.assertRaises(ConflictException) as context:
            CustomerService.create(name="Customer B", phone=" 0908888888 ", email="TRASH@example.com", address="Address B")

        self.assertIn("Email", str(context.exception))
        self.assertEqual(Customer.query.count(), 1)

    def test_customer_update_excludes_self_but_blocks_other_duplicate(self):
        customer_a = CustomerService.create(name="Customer A", phone="0902000000", email="a@example.com", address="Address A")
        customer_b = CustomerService.create(name="Customer B", phone="0902000001", email="b@example.com", address="Address B")

        updated_customer = CustomerService.update(
            customer_a.id,
            name="Customer A",
            phone="0902000000",
            email="a@example.com",
            address="Address A",
        )
        self.assertEqual(updated_customer.id, customer_a.id)

        errors = CustomerService.check_duplicate(
            phone="0902000000",
            email="A@example.com",
            exclude_customer_id=customer_b.id,
            include_deleted=True,
        )
        self.assertIn("phone", errors)
        self.assertIn("email", errors)

        with self.assertRaises(ConflictException) as context:
            CustomerService.update(
                customer_b.id,
                name="Customer B",
                phone="0902000000",
                email="A@example.com",
                address="Address B",
            )

        self.assertIn("Email", str(context.exception))
        self.assertEqual(Customer.query.get(customer_b.id).phone, "0902000001")

    def test_customer_restore_blocks_active_duplicate_conflict(self):
        customer_deleted = Customer(name="Deleted Customer", phone="0903000000", email="deleted@example.com", deleted_at=datetime.utcnow(), deleted_by="owner")
        customer_active = Customer(name="Active Customer", phone="0903000000", email="active@example.com")
        db.session.add_all([customer_deleted, customer_active])
        db.session.commit()
        customer_deleted_id = customer_deleted.id

        errors = CustomerService.check_duplicate(
            phone=customer_deleted.phone,
            email=customer_deleted.email,
            exclude_customer_id=customer_deleted_id,
            include_deleted=False,
        )
        self.assertIn("phone", errors)

        with self.assertRaises(ConflictException):
            CustomerService.restore(customer_deleted_id, actor="owner")

        self.assertIsNotNone(Customer.query.get(customer_deleted_id).deleted_at)

    def test_customer_import_blocks_duplicate_against_db_and_same_file(self):
        owner = AuthService.seed_owner_if_empty()
        self.login_as(owner)
        CustomerService.create(name="Existing Customer", phone="0904000000", email="existing@example.com", address="Address")
        import_file = self.create_customer_import_xlsx(
            rows=[
                ["Import One", "0904000000", "duplicate@example.com", "Street 1"],
                ["Import Two", "0904000000", "duplicate@example.com", "Street 2"],
            ]
        )

        try:
            with app.test_request_context("/settings"):
                report = ImportService.execute_import(app, str(import_file), "customers", "skip", False)
        finally:
            if import_file.exists():
                import_file.unlink()

        self.assertEqual(report["success"], 0)
        self.assertEqual(report["skipped"], 2)
        self.assertEqual(Customer.query.filter(Customer.deleted_at.is_(None)).count(), 1)
        self.assertEqual(len(report["errors"]), 2)

    def test_recycle_bin_inline_script_parses_cleanly(self):
        node_binary = shutil.which("node")
        if not node_binary:
            self.skipTest("Node.js is required to validate the inline recycle bin script.")

        recycle_bin_template = Path("templates/recycle_bin/index.html").read_text(encoding="utf-8")
        script_match = re.search(r"<script>(.*?)</script>\s*{% endblock %}", recycle_bin_template, re.S)
        self.assertIsNotNone(script_match)

        script_path = Path(tempfile.gettempdir()) / "spamanager-recycle-bin-inline.js"
        script_path.write_text(script_match.group(1), encoding="utf-8")

        result = subprocess.run(
            [node_binary, "--check", str(script_path)],
            capture_output=True,
            text=True,
            check=False,
        )

        self.assertEqual(result.returncode, 0, msg=result.stderr)

    def test_production_requires_owner_password(self):
        with patch.dict(
            os.environ,
            {
                "SECRET_KEY": "prod-secret",
                "DATABASE_URL": "sqlite:///prod.sqlite",
                "DEFAULT_OWNER_PASSWORD": "",
            },
            clear=True,
        ):
            with self.assertRaises(RuntimeError):
                ProductionConfig()

    def test_production_reads_database_url_and_persistent_paths(self):
        with patch.dict(
            os.environ,
            {
                "SECRET_KEY": "prod-secret",
                "DATABASE_URL": "sqlite:////app/database/spa.db",
                "DEFAULT_OWNER_PASSWORD": "prod-owner-pass",
                "PERSISTENT_ROOT": "/app/database",
            },
            clear=True,
        ):
            config = ProductionConfig()

        self.assertEqual(config.SQLALCHEMY_DATABASE_URI, "sqlite:////app/database/spa.db")
        self.assertEqual(config.PERSISTENT_ROOT, "/app/database")
        self.assertEqual(config.UPLOAD_ROOT.replace("\\", "/"), "/app/database/uploads")
        self.assertEqual(config.LOGO_UPLOAD_FOLDER.replace("\\", "/"), "/app/database/uploads/logos")
        self.assertEqual(config.AVATAR_UPLOAD_FOLDER.replace("\\", "/"), "/app/database/uploads/avatars")
        self.assertFalse(config.DEBUG)

    def test_legacy_postgres_url_is_normalized(self):
        with patch.dict(
            os.environ,
            {
                "SECRET_KEY": "prod-secret",
                "DATABASE_URL": "postgres://user:pass@localhost:5432/spamanager",
                "DEFAULT_OWNER_PASSWORD": "prod-owner-pass",
            },
            clear=True,
        ):
            config = ProductionConfig()

        self.assertEqual(config.SQLALCHEMY_DATABASE_URI, "postgresql://user:pass@localhost:5432/spamanager")

    def test_testing_config_uses_postgres_test_url_when_provided(self):
        with patch.dict(
            os.environ,
            {
                "TEST_DATABASE_URL": "postgresql://user:pass@localhost:5432/spamanager_test",
            },
            clear=True,
        ):
            config = TestingConfig()

        self.assertEqual(config.SQLALCHEMY_DATABASE_URI, "postgresql://user:pass@localhost:5432/spamanager_test")

    def test_local_config_defaults_to_postgresql_docker_profile(self):
        with patch.dict(os.environ, {}, clear=True):
            config = DevelopmentConfig()

        self.assertTrue(config.DEBUG)
        self.assertEqual(
            config.SQLALCHEMY_DATABASE_URI,
            "postgresql://spamanager:spamanager_dev_password@localhost:5433/spamanager_dev",
        )
        self.assertEqual(config.DEFAULT_OWNER_USERNAME, "owner")
        self.assertEqual(config.DEFAULT_OWNER_PASSWORD, "owner123")
        self.assertEqual(config.APP_VERSION, "5.9.0")

    def test_local_config_can_use_explicit_legacy_sqlite_fallback(self):
        with patch.dict(
            os.environ,
            {"SPA_ENABLE_SQLITE_LEGACY": "1"},
            clear=True,
        ):
            config = DevelopmentConfig()

        self.assertTrue(config.SQLALCHEMY_DATABASE_URI.startswith("sqlite:///"))

    def test_google_oauth_variables_remain_optional(self):
        with patch.dict(
            os.environ,
            {
                "SECRET_KEY": "prod-secret",
                "DATABASE_URL": "sqlite:///prod.sqlite",
                "DEFAULT_OWNER_PASSWORD": "prod-owner-pass",
            },
            clear=True,
        ):
            config = ProductionConfig()

        self.assertFalse(config.GOOGLE_AUTH_ENABLED)
        self.assertEqual(config.GOOGLE_CLIENT_ID, "")
        self.assertEqual(config.GOOGLE_CLIENT_SECRET, "")
        self.assertEqual(config.GOOGLE_REDIRECT_URI, "")
        self.assertEqual(config.GOOGLE_ALLOWED_DOMAIN, "")
        self.assertEqual(config.GOOGLE_SCOPES, ["openid", "email", "profile"])
        self.assertEqual(config.validate_google_oauth_config(), [])

    def test_google_oauth_flag_parser_handles_common_truthy_and_falsey_values(self):
        for value in (None, "", "0", "false", "False", "no", "off", "n", "  "):
            self.assertFalse(_parse_bool_env(value))

        for value in ("1", "true", "True", "yes", "on", "Y", "  yes  "):
            self.assertTrue(_parse_bool_env(value))

    def test_readme_and_env_template_match_current_production_setup(self):
        readme = Path("README.md").read_text(encoding="utf-8")
        env_example = Path(".env.example").read_text(encoding="utf-8")
        changelog = Path("CHANGELOG.md").read_text(encoding="utf-8")
        workflow = Path(".github/workflows/ci.yml").read_text(encoding="utf-8")

        self.assertIn("SpaManager is a Flask-based web app", readme)
        self.assertIn("PostgreSQL in production", readme)
        self.assertIn("Command Palette with `Ctrl+K`", readme)
        self.assertIn("badge.svg?branch=main", readme)
        self.assertIn("flask db upgrade", readme)
        self.assertIn("flask db stamp head", readme)
        self.assertIn("compileall .", readme)
        self.assertIn("DATABASE_URL=postgresql://<user>:<password>@localhost:5433/<db>", env_example)
        self.assertIn("TEST_DATABASE_URL=postgresql://<user>:<password>@localhost:5433/<test_db>", env_example)
        self.assertIn("SPA_ENABLE_SQLITE_LEGACY", env_example)
        self.assertIn("GOOGLE_AUTH_ENABLED=false", env_example)
        self.assertIn("GOOGLE_ALLOWED_DOMAIN=", env_example)
        self.assertIn("# DATABASE_URL=<Railway PostgreSQL reference variable>", env_example)
        self.assertIn("APP_VERSION=5.9.0", env_example)
        self.assertIn("v5.9.0", changelog)
        self.assertIn("v5.9.0", readme)
        self.assertIn("change-this-to-a-strong-password", env_example)
        self.assertIn("CSRF_ENABLED=1", env_example)
        self.assertIn("Google OAuth is disabled by default", readme)
        self.assertIn("GOOGLE_AUTH_ENABLED", readme)
        self.assertIn("python -m compileall .", workflow)
        self.assertNotIn("master", workflow)
        self.assertNotIn("owner123", readme)
        self.assertNotIn("owner123", env_example)
        self.assertNotIn("4.0.0", readme)

    def test_migration_commands_expose_baseline_revision(self):
        runner = app.test_cli_runner()

        current_before = runner.invoke(args=["db", "current"])
        self.assertEqual(current_before.exit_code, 0, current_before.output)
        self.assertIn("No revision stamp found", current_before.output)

        history = runner.invoke(args=["db", "history"])
        self.assertEqual(history.exit_code, 0, history.output)
        self.assertIn("0001_baseline", history.output)

    def test_workspace_migration_candidate_is_docs_only_tested_rehearsal_artifact(self):
        candidate_path = Path("docs/workspace/migration_candidates/0002_workspace_foundation.py.txt")
        self.assertTrue(candidate_path.exists())
        self.assertEqual(candidate_path.suffixes[-2:], [".py", ".txt"])

        candidate_text = candidate_path.read_text(encoding="utf-8")
        self.assertIn("TESTED PROJECT-STYLE REHEARSAL ARTIFACT ONLY", candidate_text)
        self.assertIn('revision = "0002_workspace_foundation"', candidate_text)
        self.assertIn('down_revision = "0001_baseline"', candidate_text)
        self.assertIn("def upgrade():", candidate_text)
        self.assertIn("def downgrade():", candidate_text)
        self.assertIn("do not copy into migrations/versions", candidate_text.lower())
        self.assertTrue(
            "temporary executable migration removed" in candidate_text.lower()
            or "dry-run" in candidate_text.lower()
            or "pass" in candidate_text.lower()
        )

        migration_files = [path.name for path in Path("migrations/versions").glob("*.py")]
        self.assertIn("0002_google_auth_approval.py", migration_files)
        self.assertFalse(Path("migrations/versions/0002_workspace_foundation.py").exists())
        self.assertFalse(Path("docs/workspace/WORKSPACE_MIGRATION_EXECUTION_APPROVAL.md").exists())

    def test_workspace_migration_rehearsal_plan_is_documented_and_non_executable(self):
        rehearsal_path = Path("docs/workspace/WORKSPACE_MIGRATION_REHEARSAL_PLAN.md")
        self.assertTrue(rehearsal_path.exists())

        rehearsal_text = rehearsal_path.read_text(encoding="utf-8")
        self.assertIn("Controlled Workspace Migration Rehearsal Plan", rehearsal_text)
        self.assertIn("Do not create or ship an executable migration", rehearsal_text)
        self.assertIn("rollback remains a controlled data plan", rehearsal_text)
        self.assertNotIn("migrations/versions/0002_workspace_foundation.py", rehearsal_text)

    def test_workspace_migration_execution_gate_is_documented_and_blocks_auto_deploy(self):
        gate_path = Path("docs/workspace/WORKSPACE_MIGRATION_EXECUTION_GATE.md")
        self.assertTrue(gate_path.exists())

        gate_text = gate_path.read_text(encoding="utf-8")
        self.assertIn("Workspace Migration Execution Gate and Deployment Control", gate_text)
        self.assertIn("Do not create or merge an executable workspace migration", gate_text)
        self.assertIn("Railway’s pre-deploy `db upgrade` can execute it automatically", gate_text)
        self.assertIn("before the file reaches `migrations/versions/`", gate_text)

    def test_workspace_executable_migration_approval_package_is_documented_and_non_executable(self):
        package_path = Path("docs/workspace/WORKSPACE_EXECUTABLE_MIGRATION_APPROVAL_PACKAGE.md")
        self.assertTrue(package_path.exists())

        package_text = package_path.read_text(encoding="utf-8")
        self.assertIn("Workspace Executable Migration Approval Package", package_text)
        self.assertIn("approve workspace migration deploy", package_text)
        self.assertIn("READY FOR OWNER APPROVAL PACKAGE REVIEW", package_text)
        self.assertIn("Does not create `migrations/versions/0002_workspace_foundation.py`", package_text)
        self.assertIn("Do not create `docs/workspace/WORKSPACE_MIGRATION_EXECUTION_APPROVAL.md`", package_text)
        self.assertNotIn("migrations/versions/0002_workspace_foundation.py\n", package_text)
        self.assertFalse(Path("migrations/versions/0002_workspace_foundation.py").exists())
        self.assertFalse(Path("docs/workspace/WORKSPACE_MIGRATION_EXECUTION_APPROVAL.md").exists())

    def test_workspace_migration_local_rehearsal_evidence_is_documented(self):
        evidence_path = Path("docs/workspace/WORKSPACE_MIGRATION_LOCAL_REHEARSAL_EVIDENCE.md")
        self.assertTrue(evidence_path.exists())

        evidence_text = evidence_path.read_text(encoding="utf-8")
        self.assertIn("Workspace Migration Local Rehearsal Evidence", evidence_text)
        self.assertIn("No executable migration was added to `migrations/versions/`", evidence_text)
        self.assertIn("python -m unittest discover -s tests -p \"test*.py\" -v", evidence_text)
        self.assertIn("python -m compileall .", evidence_text)

    def test_db_upgrade_creates_schema_and_stamps_head(self):
        self.clear_database_schema()
        self.assertEqual(sa_inspect(db.engine).get_table_names(), [])

        runner = app.test_cli_runner()
        result = runner.invoke(args=["db", "upgrade"])
        self.assertEqual(result.exit_code, 0, result.output)
        self.assertIn("Applied 0001_baseline", result.output)
        self.assertIn("Applied 0002_google_auth_approval", result.output)

        tables = sa_inspect(db.engine).get_table_names()
        self.assertIn("users", tables)
        self.assertIn("customers", tables)
        self.assertIn("alembic_version", tables)

        current_after = runner.invoke(args=["db", "current"])
        self.assertEqual(current_after.exit_code, 0, current_after.output)
        self.assertIn("0002_google_auth_approval", current_after.output)

    def test_version_is_rendered_from_config_in_setting_ui(self):
        owner = AuthService.seed_owner_if_empty()
        self.login_as(owner)
        response = self.client.get("/settings")
        html = response.get_data(as_text=True)
        self.assertIn("SpaManager v5.9.0", html)
        self.assertIn("v5.9.0", html)
        self.assertIn(">5.9.0<", html)

    def test_sidebar_footer_shows_current_version(self):
        owner = AuthService.seed_owner_if_empty()
        self.login_as(owner)
        response = self.client.get("/")
        html = response.get_data(as_text=True)
        self.assertIn("SpaManager v5.9.0", html)
        self.assertNotIn("SpaManager v4.0", html)

    def test_workspace_models_expose_expected_metadata(self):
        self.assertEqual(Workspace.__tablename__, "workspaces")
        self.assertEqual(WorkspaceMember.__tablename__, "workspace_members")

        workspace_columns = {column.name for column in Workspace.__table__.columns}
        member_columns = {column.name for column in WorkspaceMember.__table__.columns}

        self.assertTrue({
            "id",
            "name",
            "slug",
            "status",
            "created_by_id",
            "notes",
            "created_at",
            "updated_at",
        }.issubset(workspace_columns))

        self.assertTrue({
            "id",
            "workspace_id",
            "user_id",
            "role",
            "status",
            "invited_by_id",
            "joined_at",
            "created_at",
            "updated_at",
        }.issubset(member_columns))

        unique_names = {constraint.name for constraint in WorkspaceMember.__table__.constraints if constraint.__class__.__name__ == "UniqueConstraint"}
        self.assertIn("uq_workspace_members_workspace_user", unique_names)

    def test_workspace_models_expose_expected_status_and_role_constants(self):
        self.assertEqual(Workspace.WORKSPACE_STATUSES, ("active", "pending", "suspended", "archived"))
        self.assertEqual(WorkspaceMember.WORKSPACE_MEMBER_ROLES, ("owner", "admin", "staff"))
        self.assertEqual(WorkspaceMember.WORKSPACE_MEMBER_STATUSES, ("active", "invited", "disabled"))

    def test_workspace_model_smoke_create_and_relationships(self):
        owner = self.create_user("workspace-owner", password="owner-pass", full_name="Workspace Owner", role="OWNER")
        member_user = self.create_user("workspace-member", password="member-pass", full_name="Workspace Member", role="STAFF")

        workspace = Workspace(name="Test Workspace", slug="test-workspace", status="active", created_by=owner, notes="Smoke test")
        member = WorkspaceMember(workspace=workspace, user=member_user, role="staff", status="active", invited_by=owner)
        db.session.add(workspace)
        db.session.add(member)
        db.session.commit()

        loaded_workspace = Workspace.query.filter_by(slug="test-workspace").first()
        self.assertIsNotNone(loaded_workspace)
        self.assertEqual(loaded_workspace.created_by_id, owner.id)
        self.assertEqual(loaded_workspace.members[0].user_id, member_user.id)
        self.assertEqual(loaded_workspace.members[0].invited_by_id, owner.id)
        self.assertTrue(loaded_workspace.is_active())
        self.assertTrue(loaded_workspace.members[0].is_staff())

    def test_settings_template_includes_explicit_csrf_tokens_for_post_forms(self):
        template = Path("templates/setting/index.html").read_text(encoding="utf-8")
        self.assertIn('id="createBackupForm"', template)
        self.assertIn('id="editBackupNoteForm"', template)
        self.assertIn('id="uploadBackupForm"', template)
        self.assertIn('action="{{ url_for(\'setting.save_spa_info\') }}" method="POST"', template)
        self.assertGreaterEqual(template.count('name="csrf_token"'), 4)
        self.assertIn('Backup lưu lại trạng thái hiện tại của hệ thống.', template)
        self.assertIn('id="restoreConfirmCheck"', template)
        self.assertIn('id="deleteBackupConfirmCheck"', template)
        self.assertIn('id="wizardRestoreConfirmCheck"', template)
        self.assertIn('id="confirmRestoreBackupBtn" disabled', template)
        self.assertIn('id="confirmDeleteBackupBtn" disabled', template)
        self.assertIn('id="wizard-btn-confirm" disabled', template)
        self.assertIn('btn-restore-backup', template)
        self.assertIn('btn-delete-backup', template)
        self.assertNotIn('href="{{ url_for(\'setting.restore_from_backup\'', template)
        self.assertNotIn('href="{{ url_for(\'setting.delete_backup\'', template)

    def test_settings_backup_restore_ux_copy_and_js_guardrails_are_present(self):
        template = Path("templates/setting/index.html").read_text(encoding="utf-8")
        script = Path("static/js/setting.js").read_text(encoding="utf-8")

        self.assertIn('Bản sao lưu sẽ lưu lại toàn bộ dữ liệu hiện tại của hệ thống', template)
        self.assertIn('Hành động này không thể hoàn tác!', template)
        self.assertIn('Nên tạo một bản sao lưu hiện tại trước khi tiếp tục', template)
        self.assertIn('Tôi hiểu thao tác này sẽ thay thế dữ liệu hiện tại.', template)
        self.assertIn('Tôi hiểu thao tác xóa này là vĩnh viễn.', template)
        self.assertNotRegex(template, r'[A-Z]:\\\\')

        self.assertIn('deleteBackupConfirmCheck', script)
        self.assertIn('confirmRestoreBackupBtn.disabled = !this.checked', script)
        self.assertIn('wizardRestoreConfirmCheck', script)
        self.assertIn('confirmDeleteBackupBtn.disabled = !this.checked', script)
        self.assertIn('wizardBtnConfirm.disabled = !this.checked || isExecutingRestore', script)
        self.assertIn('requestSubmit', script)

    def test_database_engine_helpers_detect_sqlite_and_postgresql(self):
        self.assertEqual(get_database_engine("sqlite:///example.db"), "sqlite")
        self.assertEqual(get_database_engine("postgresql://user:pass@localhost:5432/spamanager"), "postgresql")
        self.assertTrue(is_sqlite_database("sqlite:///example.db"))
        self.assertTrue(is_postgresql_database("postgresql://user:pass@localhost:5432/spamanager"))
        self.assertIn("Backup Center", get_postgresql_backup_center_message())
        self.assertIn("PostgreSQL", get_postgresql_restore_guard_message())

    def test_settings_backup_center_shows_postgresql_guard_and_disables_actions(self):
        owner = self.create_user("settings-pg-guard-owner", password="owner-pass", full_name="PG Guard Owner", role="OWNER")
        self.login_as(owner)
        original_uri = app.config["SQLALCHEMY_DATABASE_URI"]
        app.config["SQLALCHEMY_DATABASE_URI"] = "postgresql://user:pass@localhost:5432/spamanager"
        try:
            response = self.client.get("/settings", follow_redirects=False)
            html = response.get_data(as_text=True)
        finally:
            app.config["SQLALCHEMY_DATABASE_URI"] = original_uri

        self.assertEqual(response.status_code, 200)
        self.assertIn("Backup Center", html)
        self.assertIn("PostgreSQL", html)
        self.assertIn("data-backup-engine=\"postgresql\"", html)
        self.assertIn("btn-create-backup", html)
        self.assertIn("disabled", html)

    def test_settings_postgresql_backup_guard_blocks_create_restore_upload_and_validate(self):
        owner = self.create_user("settings-pg-guard-actions-owner", password="owner-pass", full_name="PG Guard Actions Owner", role="OWNER")
        self.login_as(owner)
        backup_id, backup_meta, backup_path = self.create_settings_backup_via_route(owner, notes="SQLite backup before PG guard")

        temp_restore = Path(tempfile.gettempdir()) / f"pg-restore-{uuid.uuid4().hex}.sqlite"
        shutil.copy2(backup_path, temp_restore)

        try:
            original_uri = app.config["SQLALCHEMY_DATABASE_URI"]
            app.config["SQLALCHEMY_DATABASE_URI"] = "postgresql://user:pass@localhost:5432/spamanager"
            before_files = {path.name for path in Path(BackupService.get_backup_dir(app)).glob("*.sqlite")}
            before_total = Customer.query.count()

            create_response = self.post_with_csrf(
                "/settings/backup",
                path="/settings",
                data={"notes": "PG blocked", "backup_type": "Manual", "format": "json"},
                headers={"X-Requested-With": "XMLHttpRequest"},
            )
            self.assertEqual(create_response.status_code, 400)
            self.assertTrue(create_response.is_json)
            self.assertTrue(create_response.get_json()["blocked"])
            self.assertIn("PostgreSQL", create_response.get_json()["message"])

            with temp_restore.open("rb") as restore_handle:
                upload_response = self.post_with_csrf(
                    "/settings/backup/upload",
                    path="/settings",
                    data={"backup_file": (restore_handle, temp_restore.name), "notes": "PG blocked"},
                    headers={"X-Requested-With": "XMLHttpRequest"},
                )
            self.assertEqual(upload_response.status_code, 400)
            self.assertTrue(upload_response.is_json)
            self.assertTrue(upload_response.get_json()["blocked"])

            validate_response = self.client.get(f"/settings/restore-wizard/validate/{backup_id}")
            self.assertEqual(validate_response.status_code, 200)
            self.assertTrue(validate_response.is_json)
            self.assertTrue(validate_response.get_json()["blocked"])

            restore_response = self.post_with_csrf(
                f"/settings/backup/restore/{backup_id}",
                path="/settings",
                data={"backup_id": backup_id, "format": "json"},
                headers={"X-Requested-With": "XMLHttpRequest"},
            )
            self.assertEqual(restore_response.status_code, 400)
            self.assertTrue(restore_response.is_json)
            self.assertTrue(restore_response.get_json()["blocked"])

            confirm_response = self.post_with_csrf(
                "/settings/restore-wizard/confirm",
                path="/settings",
                json={"backup_id": backup_id},
                headers={"X-Requested-With": "XMLHttpRequest"},
            )
            self.assertEqual(confirm_response.status_code, 400)
            self.assertTrue(confirm_response.is_json)
            self.assertTrue(confirm_response.get_json()["blocked"])

            after_files = {path.name for path in Path(BackupService.get_backup_dir(app)).glob("*.sqlite")}
            self.assertEqual(before_files, after_files)
            self.assertEqual(Customer.query.count(), before_total)
        finally:
            app.config["SQLALCHEMY_DATABASE_URI"] = original_uri
            if temp_restore.exists():
                temp_restore.unlink()
            if backup_path.exists():
                backup_path.unlink()
            BackupRepository.delete(app, backup_id)

    def test_import_template_files_exist_and_routes_download_them(self):
        customers_template = Path("static/templates/import/customers_template.xlsx")
        services_template = Path("static/templates/import/services_template.xlsx")

        self.assertTrue(customers_template.exists())
        self.assertTrue(services_template.exists())
        self.assertGreater(customers_template.stat().st_size, 0)
        self.assertGreater(services_template.stat().st_size, 0)

        owner = AuthService.seed_owner_if_empty()
        self.login_as(owner)

        for url, filename in (
            ("/settings/template/customers", "customers_template.xlsx"),
            ("/settings/template/services", "services_template.xlsx"),
        ):
            with self.subTest(url=url):
                response = self.client.get(url, follow_redirects=False)
                try:
                    self.assertEqual(response.status_code, 200)
                    self.assertGreater(len(response.data), 0)
                    self.assertIn(filename, response.headers.get("Content-Disposition", ""))
                finally:
                    response.close()

    def test_settings_backup_center_uses_vietnamese_badges_and_status_labels(self):
        template = Path("templates/setting/index.html").read_text(encoding="utf-8")

        self.assertIn("ĐÃ NHẬP", template)
        self.assertIn("CỤC BỘ", template)
        self.assertIn("Thủ công", template)
        self.assertIn("Trước nhập", template)
        self.assertIn("Trước khôi phục", template)
        self.assertIn("Trước cập nhật", template)
        self.assertIn("Sao lưu thủ công", template)
        self.assertIn("Trước khi import", template)
        self.assertIn("Trước khi khôi phục", template)
        self.assertIn("Trước khi cập nhật", template)
        self.assertIn("Hợp lệ", template)
        self.assertIn("Thiếu tệp sao lưu", template)
        self.assertIn("Không hợp lệ", template)
        self.assertNotIn(">IMPORTED<", template)
        self.assertNotIn(">LOCAL<", template)
        self.assertNotIn(">Manual<", template)
        self.assertNotIn(">Import<", template)
        self.assertNotIn(">Restore<", template)
        self.assertNotIn(">Update<", template)

    def test_backup_service_metadata_uses_clean_vietnamese_text(self):
        service_source = Path("services/backup_service.py").read_text(encoding="utf-8")
        self.assertIn("Backup ngày", service_source)
        self.assertIn("Backup tạo lúc", service_source)
        self.assertIn("Hôm nay", service_source)
        self.assertIn("Hôm qua", service_source)
        for marker in ["Ã", "á»", "áº", "Æ", "Ä‘", "â€¢", "Â"]:
            self.assertNotIn(marker, service_source)

    def test_new_backup_metadata_and_list_render_clean_vietnamese_text(self):
        owner = self.create_user("settings-backup-vn-owner", password="owner-pass", full_name="Backup VN Owner", role="OWNER")
        self.login_as(owner)
        backup_id, backup_meta, backup_path = self.create_settings_backup_via_route(owner, notes="")
        try:
            self.assertEqual(backup_meta["app_version"], "SpaManager v5.9.0")
            self.assertIn("Backup ngày", backup_meta["display_name"])
            self.assertIn("Backup tạo lúc", backup_meta["notes"])

            response = self.client.get("/settings", follow_redirects=False)
            html = response.get_data(as_text=True)
            self.assertEqual(response.status_code, 200)
            self.assertIn("Backup ngày", html)
            self.assertIn("Hôm nay", html)
            self.assertNotIn("Backup ngÃ", html)
            self.assertNotIn("Backup táº", html)
            self.assertNotIn("Backup lÃºc", html)
        finally:
            if backup_path.exists():
                backup_path.unlink()
            BackupRepository.delete(app, backup_id)

    def test_settings_backup_list_reads_legacy_folder_read_only(self):
        owner = self.create_user("settings-backup-legacy-owner", password="owner-pass", full_name="Legacy Backup Owner", role="OWNER")
        self.login_as(owner)

        legacy_dir = Path(BackupService.get_legacy_backup_dir(app))
        legacy_dir.mkdir(parents=True, exist_ok=True)
        legacy_metadata_path = Path(BackupRepository.get_legacy_metadata_path(app))
        original_metadata = legacy_metadata_path.read_text(encoding="utf-8") if legacy_metadata_path.exists() else None

        legacy_backup_id = str(uuid.uuid4())
        legacy_filename = f"SpaManager_Backup_legacy_{uuid.uuid4().hex}.sqlite"
        legacy_file_path = legacy_dir / legacy_filename
        conn = sqlite3.connect(legacy_file_path)
        conn.execute("CREATE TABLE sample (id INTEGER PRIMARY KEY)")
        conn.commit()
        conn.close()

        legacy_metadata = {
            legacy_backup_id: {
                "id": legacy_backup_id,
                "filename": legacy_filename,
                "display_name": "Backup ngày 04/07/2026 15:24",
                "created_at": datetime.utcnow().isoformat(),
                "size": legacy_file_path.stat().st_size,
                "database_version": "v5.1.0",
                "app_version": "SpaManager v5.1.0",
                "notes": "Legacy backup should stay visible",
                "type": "Manual",
                "created_by": None,
                "status": "Valid",
            }
        }

        try:
            legacy_metadata_path.write_text(
                json.dumps(legacy_metadata, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

            response = self.client.get("/settings", follow_redirects=False)
            html = response.get_data(as_text=True)

            self.assertEqual(response.status_code, 200)
            self.assertIn("Legacy backup should stay visible", html)
            self.assertIn("Backup ngày 04/07/2026 15:24", html)
            self.assertIn("SpaManager v5.1.0", html)
        finally:
            if legacy_file_path.exists():
                legacy_file_path.unlink()
            if original_metadata is None:
                if legacy_metadata_path.exists():
                    legacy_metadata_path.unlink()
            else:
                legacy_metadata_path.write_text(original_metadata, encoding="utf-8")

    def test_settings_backup_corrupt_metadata_json_does_not_500(self):
        owner = self.create_user("settings-backup-corrupt-owner", password="owner-pass", full_name="Corrupt Backup Owner", role="OWNER")
        self.login_as(owner)
        backup_id, backup_meta, backup_path = self.create_settings_backup_via_route(owner, notes="Corrupt metadata fallback")
        metadata_path = Path(BackupRepository.get_metadata_path(app))
        original_metadata = metadata_path.read_text(encoding="utf-8") if metadata_path.exists() else None

        try:
            metadata_path.write_text("{ this is not valid json", encoding="utf-8")

            response = self.client.get("/settings", follow_redirects=False)
            html = response.get_data(as_text=True)

            self.assertEqual(response.status_code, 200)
            self.assertIn("Backup ngày", html)
            self.assertIn("Tự động đồng bộ từ đĩa cứng", html)
            self.assertNotIn("Không thể tải danh sách sao lưu", html)
        finally:
            if backup_path.exists():
                backup_path.unlink()
            if original_metadata is None:
                if metadata_path.exists():
                    metadata_path.unlink()
            else:
                metadata_path.write_text(original_metadata, encoding="utf-8")
            BackupRepository.delete(app, backup_id)

    def test_backup_directory_uses_persistent_root_volume(self):
        backup_dir = Path(BackupService.get_backup_dir(app)).resolve()
        self.assertTrue(str(backup_dir).startswith(str(TEST_MEDIA_ROOT.resolve())))

    def test_settings_backup_list_keeps_backup_versions_visible(self):
        owner = self.create_user("settings-backup-legacy-owner", password="owner-pass", full_name="Backup Legacy Owner", role="OWNER")
        self.login_as(owner)

        first_backup_id, first_backup_meta, first_backup_path = self.create_settings_backup_via_route(
            owner,
            notes="Backup version 5.1 should stay visible",
        )
        second_backup_id, second_backup_meta, second_backup_path = self.create_settings_backup_via_route(
            owner,
            notes="Backup version 5.3 should stay visible",
        )

        first_backup_meta["database_version"] = "v5.1.0"
        first_backup_meta["app_version"] = "SpaManager v5.1.0"
        self.assertEqual(second_backup_meta["app_version"], "SpaManager v5.9.0")
        second_backup_meta["database_version"] = "v5.9.0"
        second_backup_meta["app_version"] = "SpaManager v5.9.0"

        try:
            BackupRepository.save(app, first_backup_id, first_backup_meta)
            BackupRepository.save(app, second_backup_id, second_backup_meta)

            response = self.client.get("/settings", follow_redirects=False)
            html = response.get_data(as_text=True)

            self.assertEqual(response.status_code, 200)
            self.assertIn("Backup version 5.1 should stay visible", html)
            self.assertIn("Backup version 5.3 should stay visible", html)
            self.assertIn("SpaManager v5.1.0", html)
            self.assertIn("SpaManager v5.9.0", html)
            self.assertNotIn("/app/database", html)
        finally:
            if first_backup_path.exists():
                first_backup_path.unlink()
            if second_backup_path.exists():
                second_backup_path.unlink()
            BackupRepository.delete(app, first_backup_id)
            BackupRepository.delete(app, second_backup_id)

    def test_settings_backup_list_ignores_corrupted_metadata_entry(self):
        owner = self.create_user("settings-backup-corrupt-owner", password="owner-pass", full_name="Backup Corrupt Owner", role="OWNER")
        self.login_as(owner)
        original_metadata = BackupRepository.load_all(app).copy()
        backup_id, backup_meta, backup_path = self.create_settings_backup_via_route(owner, notes="Healthy backup remains visible")

        corrupted_id = str(uuid.uuid4())
        corrupted_metadata = BackupRepository.load_all(app)
        corrupted_metadata[corrupted_id] = {
            "id": corrupted_id,
            "notes": "Broken metadata entry without filename",
            "status": "Valid",
        }

        try:
            BackupRepository.save_all(app, corrupted_metadata)

            response = self.client.get("/settings", follow_redirects=False)
            html = response.get_data(as_text=True)

            self.assertEqual(response.status_code, 200)
            self.assertIn("Healthy backup remains visible", html)
            self.assertNotIn("Không thể tải danh sách sao lưu", html)
        finally:
            if backup_path.exists():
                backup_path.unlink()
            BackupRepository.save_all(app, original_metadata)
            if BackupRepository.get_by_id(app, backup_id):
                BackupRepository.delete(app, backup_id)

    def test_backup_file_path_helper_blocks_traversal_and_stays_inside_backup_dir(self):
        backup_dir = Path(BackupService.get_backup_dir(app)).resolve()
        resolved = Path(BackupService.get_backup_file_path(app, "../../evil.sqlite"))
        self.assertEqual(resolved.parent, backup_dir)
        self.assertEqual(resolved.name, "evil.sqlite")

    def test_settings_backup_upload_error_response_is_generic(self):
        owner = AuthService.seed_owner_if_empty()
        self.login_as(owner)

        temp_db = Path(tempfile.gettempdir()) / f"settings-upload-{uuid.uuid4().hex}.sqlite"
        conn = sqlite3.connect(temp_db)
        conn.execute("CREATE TABLE sample (id INTEGER PRIMARY KEY)")
        conn.commit()
        conn.close()

        try:
            with patch("routes.setting.BackupService.inspect_external_backup", side_effect=RuntimeError("secret-path-leak")):
                with temp_db.open("rb") as file_handle:
                    response = self.post_with_csrf(
                        "/settings/backup/upload",
                        path="/settings",
                        data={
                            "backup_file": (file_handle, temp_db.name),
                            "notes": "Test upload",
                        },
                        content_type="multipart/form-data",
                    )

            self.assertEqual(response.status_code, 500)
            payload = response.get_json()
            self.assertFalse(payload["success"])
            self.assertEqual(payload["message"], "Lỗi khi xử lý tệp tải lên.")
            self.assertNotIn("secret-path-leak", response.get_data(as_text=True))
        finally:
            if temp_db.exists():
                temp_db.unlink()

    def test_settings_staff_is_forbidden_on_sensitive_routes(self):
        owner = self.create_user("settings-owner", password="owner-pass", full_name="Settings Owner", role="OWNER")
        admin = self.create_user("settings-admin", password="admin-pass", full_name="Settings Admin", role="ADMIN")
        staff = self.create_user("settings-staff", password="staff-pass", full_name="Settings Staff", role="STAFF")

        backup_id, backup_meta, backup_path = self.create_settings_backup_via_route(owner, notes="Route access backup")
        import_file = self.create_customer_import_xlsx()
        logo_path = self.create_media_file(Path("uploads") / "logos" / "settings-route-logo.png", b"\x89PNG\r\n\x1a\n")
        logo_relative = "logos/settings-route-logo.png"
        Setting.set("spa_logo", logo_relative)

        staff_token = self.get_csrf_token("/customers")
        routes_to_block = [
            ("GET", "/settings", None, None),
            ("POST", "/settings/backup", {"notes": "x", "backup_type": "Manual", "format": "json"}, "form"),
            ("POST", f"/settings/backup/delete/{backup_id}", None, None),
            ("GET", f"/settings/backup/download/{backup_id}", None, None),
            ("POST", f"/settings/backup/restore/{backup_id}", {"backup_id": backup_id}, "json"),
            ("POST", "/settings/restore", {"restore_file": (logo_path.open("rb"), logo_path.name)}, "multipart"),
            ("GET", f"/settings/restore-wizard/validate/{backup_id}", None, None),
            ("POST", "/settings/restore-wizard/confirm", {"backup_id": backup_id}, "json"),
            ("POST", "/settings/import/analyze", {"import_file": (import_file.open("rb"), import_file.name), "import_type": "customers"}, "multipart"),
            ("POST", "/settings/import/execute", {"temp_file_id": "fake-temp", "import_type": "customers", "duplicate_action": "skip", "all_or_nothing": False}, "json"),
            ("POST", "/settings/delete-logo", None, None),
            ("POST", "/settings/save-spa-info", {
                "spa_name": "Blocked",
                "spa_owner": "Blocked",
                "spa_phone": "0900000000",
                "spa_email": "blocked@example.com",
                "spa_address": "Blocked",
                "spa_open_time": "08:00",
                "spa_close_time": "20:00",
            }, "form"),
        ]

        try:
            self.login_as(staff)
            for method, url, data, payload_type in routes_to_block:
                with self.subTest(route=url):
                    if method == "GET":
                        response = self.client.get(url, follow_redirects=False)
                    elif payload_type == "json":
                        response = self.client.post(url, json=data, headers={"X-CSRFToken": staff_token}, follow_redirects=False)
                    elif payload_type == "multipart":
                        response = self.client.post(
                            url,
                            data=data,
                            headers={"X-CSRFToken": staff_token},
                            content_type="multipart/form-data",
                            follow_redirects=False,
                        )
                        if url.endswith("/restore"):
                            data["restore_file"][0].close()
                        if url.endswith("/analyze"):
                            data["import_file"][0].close()
                    else:
                        response = self.client.post(url, data=data or {}, headers={"X-CSRFToken": staff_token}, follow_redirects=False)

                    self.assertIn(response.status_code, (403, 401), url)

            self.assertTrue(backup_path.exists())
            self.assertEqual(Setting.get("spa_logo"), logo_relative)
        finally:
            if import_file.exists():
                import_file.unlink()
            if logo_path.exists():
                logo_path.unlink()
            if backup_path.exists():
                backup_path.unlink()
            BackupRepository.delete(app, backup_id)

        self.login_as(owner)
        self.assertEqual(self.client.get("/settings", follow_redirects=False).status_code, 200)
        self.login_as(admin)
        self.assertEqual(self.client.get("/settings", follow_redirects=False).status_code, 200)

    def test_settings_backup_create_missing_csrf_has_no_side_effect(self):
        owner = self.create_user("settings-backup-owner", password="owner-pass", full_name="Backup Owner", role="OWNER")
        self.login_as(owner)
        backup_dir = Path(BackupService.get_backup_dir(app))
        before_files = {path.name for path in backup_dir.glob("*.sqlite")}

        response = self.client.post(
            "/settings/backup",
            data={"notes": "No CSRF", "backup_type": "Manual", "format": "json"},
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )

        after_files = {path.name for path in backup_dir.glob("*.sqlite")}
        self.assertEqual(response.status_code, 400)
        self.assertEqual(before_files, after_files)

    def test_settings_backup_delete_missing_csrf_keeps_backup_file(self):
        owner = self.create_user("settings-delete-owner", password="owner-pass", full_name="Delete Owner", role="OWNER")
        backup_id, backup_meta, backup_path = self.create_settings_backup_via_route(owner, notes="Delete CSRF backup")
        before_log_count = ActivityLog.query.count()

        try:
            response = self.client.post(f"/settings/backup/delete/{backup_id}", follow_redirects=False)
            self.assertEqual(response.status_code, 400)
            self.assertTrue(backup_path.exists())
            self.assertIsNotNone(BackupRepository.get_by_id(app, backup_id))
            self.assertEqual(ActivityLog.query.count(), before_log_count)
        finally:
            if backup_path.exists():
                backup_path.unlink()
            BackupRepository.delete(app, backup_id)

    def test_settings_restore_confirm_missing_csrf_keeps_database_unchanged(self):
        owner = self.create_user("settings-restore-owner", password="owner-pass", full_name="Restore Owner", role="OWNER")
        self.login_as(owner)
        customer_before = self.create_customer_record("Before Restore")
        backup_id, backup_meta, backup_path = self.create_settings_backup_via_route(owner, notes="Restore CSRF backup")
        customer_after = self.create_customer_record("After Backup")

        try:
            response = self.client.post(
                "/settings/restore-wizard/confirm",
                json={"backup_id": backup_id},
                follow_redirects=False,
            )
            self.assertEqual(response.status_code, 400)
            self.assertIsNotNone(Customer.query.get(customer_before.id))
            self.assertIsNotNone(Customer.query.get(customer_after.id))
            self.assertEqual(Customer.query.count(), 2)
        finally:
            if backup_path.exists():
                backup_path.unlink()
            BackupRepository.delete(app, backup_id)

    def test_settings_import_execute_missing_csrf_keeps_database_unchanged(self):
        owner = self.create_user("settings-import-owner", password="owner-pass", full_name="Import Owner", role="OWNER")
        self.login_as(owner)
        before_count = Customer.query.count()
        import_file = self.create_customer_import_xlsx()

        try:
            analyze_response = self.post_with_csrf(
                "/settings/import/analyze",
                path="/settings",
                data={
                    "import_file": (import_file.open("rb"), import_file.name),
                    "import_type": "customers",
                },
                content_type="multipart/form-data",
            )
            self.assertEqual(analyze_response.status_code, 200)
            temp_file_id = analyze_response.get_json()["temp_file_id"]

            execute_response = self.client.post(
                "/settings/import/execute",
                json={
                    "temp_file_id": temp_file_id,
                    "import_type": "customers",
                    "duplicate_action": "skip",
                    "all_or_nothing": False,
                },
                follow_redirects=False,
            )
            self.assertEqual(execute_response.status_code, 400)
            self.assertEqual(Customer.query.count(), before_count)
        finally:
            if import_file.exists():
                import_file.unlink()
            temp_path = Path(ImportService.get_upload_dir(app)) / temp_file_id if 'temp_file_id' in locals() else None
            if temp_path and temp_path.exists():
                temp_path.unlink()

    def test_settings_logo_upload_and_delete_require_csrf(self):
        owner = self.create_user("settings-logo-owner", password="owner-pass", full_name="Logo Owner", role="OWNER")
        self.login_as(owner)
        logo_before = Setting.get("spa_logo")
        before_files = {path.name for path in (TEST_MEDIA_ROOT / "uploads" / "logos").glob("*") if path.is_file()}

        logo_upload = BytesIO(b"\x89PNG\r\n\x1a\n")
        logo_upload.name = "logo-no-csrf.png"
        upload_response = self.client.post(
            "/settings/save-spa-info",
            data={
                "spa_name": "Spa Test",
                "spa_owner": "Owner",
                "spa_phone": "0901234567",
                "spa_email": "owner@example.com",
                "spa_address": "Address",
                "spa_open_time": "08:00",
                "spa_close_time": "20:00",
                "spa_logo": (logo_upload, logo_upload.name),
            },
            content_type="multipart/form-data",
            follow_redirects=False,
        )

        after_upload_files = {path.name for path in (TEST_MEDIA_ROOT / "uploads" / "logos").glob("*") if path.is_file()}
        self.assertEqual(upload_response.status_code, 400)
        self.assertEqual(logo_before, Setting.get("spa_logo"))
        self.assertEqual(before_files, after_upload_files)

        created_logo = self.create_media_file(Path("uploads") / "logos" / "delete-logo-no-csrf.png", b"\x89PNG\r\n\x1a\n")
        Setting.set("spa_logo", "logos/delete-logo-no-csrf.png")

        delete_response = self.client.post("/settings/delete-logo", follow_redirects=False)
        self.assertEqual(delete_response.status_code, 400)
        self.assertTrue(created_logo.exists())
        self.assertEqual(Setting.get("spa_logo"), "logos/delete-logo-no-csrf.png")

        if created_logo.exists():
            created_logo.unlink()
        Setting.set("spa_logo", logo_before or "")

    def test_settings_path_traversal_routes_reject_outside_paths(self):
        owner = self.create_user("settings-traversal-owner", password="owner-pass", full_name="Traversal Owner", role="OWNER")
        backup_id, backup_meta, backup_path = self.create_settings_backup_via_route(owner, notes="Traversal backup")
        self.login_as(owner)

        try:
            traversal_paths = [
                f"/settings/backup/download/..%2F{backup_id}",
                f"/settings/backup/delete/..%2F{backup_id}",
                f"/settings/backup/restore/..%2F{backup_id}",
            ]
            for url in traversal_paths:
                response = self.client.open(url, method="GET" if "download" in url else "POST", follow_redirects=False)
                self.assertIn(response.status_code, (400, 404), url)
                self.assertNotIn(str(TEST_MEDIA_ROOT).replace("\\", "/"), response.get_data(as_text=True))
            self.assertTrue(backup_path.exists())
        finally:
            if backup_path.exists():
                backup_path.unlink()
            BackupRepository.delete(app, backup_id)

    def test_settings_restore_validate_is_read_only_and_invalid_restore_file_is_blocked(self):
        owner = self.create_user("settings-validate-owner", password="owner-pass", full_name="Validate Owner", role="OWNER")
        backup_id, backup_meta, backup_path = self.create_settings_backup_via_route(owner, notes="Validate backup")
        before_customer_count = Customer.query.count()

        invalid_restore = Path(tempfile.gettempdir()) / f"invalid-restore-{uuid.uuid4().hex}.sqlite"
        invalid_restore.write_text("not a sqlite database", encoding="utf-8")

        try:
            self.login_as(owner)
            validate_response = self.client.get(f"/settings/restore-wizard/validate/{backup_id}", follow_redirects=False)
            self.assertEqual(validate_response.status_code, 200)
            self.assertEqual(Customer.query.count(), before_customer_count)

            with invalid_restore.open("rb") as file_handle:
                restore_response = self.post_with_csrf(
                    "/settings/restore",
                    path="/settings",
                    data={"restore_file": (file_handle, invalid_restore.name)},
                    content_type="multipart/form-data",
                )
            self.assertEqual(restore_response.status_code, 302)
            self.assertEqual(Customer.query.count(), before_customer_count)
        finally:
            if invalid_restore.exists():
                invalid_restore.unlink()
            if backup_path.exists():
                backup_path.unlink()
            BackupRepository.delete(app, backup_id)

    def test_settings_backup_logs_are_attributed_and_sanitized(self):
        owner = self.create_user("settings-log-owner", password="owner-pass", full_name="Log Owner", role="OWNER")
        self.login_as(owner)
        before_log_count = ActivityLog.query.count()
        backup_id, backup_meta, backup_path = self.create_settings_backup_via_route(owner, notes="Log backup")
        created_log = ActivityLog.query.order_by(ActivityLog.id.desc()).first()
        self.assertGreater(ActivityLog.query.count(), before_log_count)
        self.assertIsNotNone(created_log)
        self.assertEqual(created_log.user_id, owner.id)
        self.assertNotIn("password", created_log.description.lower())
        self.assertNotIn("token", created_log.description.lower())
        self.assertNotIn("database_url", created_log.description.lower())
        self.assertNotIn(str(TEST_MEDIA_ROOT).lower(), created_log.description.lower())

        delete_response = self.post_with_csrf(f"/settings/backup/delete/{backup_id}", path="/settings")
        self.assertEqual(delete_response.status_code, 200)
        delete_log = ActivityLog.query.order_by(ActivityLog.id.desc()).first()
        self.assertIsNotNone(delete_log)
        self.assertEqual(delete_log.user_id, owner.id)
        self.assertNotIn("password", delete_log.description.lower())
        self.assertNotIn("token", delete_log.description.lower())
        self.assertNotIn("database_url", delete_log.description.lower())
        self.assertNotIn(str(TEST_MEDIA_ROOT).lower(), delete_log.description.lower())

        if backup_path.exists():
            backup_path.unlink()
        BackupRepository.delete(app, backup_id)

    def test_sidebar_template_contains_clean_vietnamese_text(self):
        sidebar = Path("templates/layout/sidebar.html").read_text(encoding="utf-8")
        self.assertIn("SpaManager", sidebar)
        self.assertIn("Quản lý Spa • Nail • Makeup", sidebar)
        self.assertIn("© Văn Công Trường", sidebar)
        self.assertIn("Trang chủ", sidebar)
        self.assertIn("Người dùng", sidebar)
        self.assertNotIn("TIỆM NHÀ NHÍM", sidebar)
        for marker in ["Ã", "á»", "áº", "Æ", "Ä‘", "â€¢", "Â"]:
            self.assertNotIn(marker, sidebar)

    def test_login_template_contains_product_branding(self):
        login = Path("templates/auth/login.html").read_text(encoding="utf-8")
        self.assertIn("SpaManager", login)
        self.assertIn("Quản lý Spa • Nail • Makeup", login)
        self.assertIn("© Văn Công Trường", login)
        self.assertNotIn("TIỆM NHÀ NHÍM", login)

    def test_activity_log_action_badge_is_scoped_and_truncated(self):
        template = Path("templates/activity_log/index.html").read_text(encoding="utf-8")
        css = Path("static/css/pages/activity-log.css").read_text(encoding="utf-8")

        self.assertIn("activity-action-badge", template)
        self.assertIn("name=\"actor\"", template)
        self.assertIn("col-actor", template)
        self.assertIn(".activity-log-page .activity-action-badge", css)
        self.assertIn("text-overflow: ellipsis", css)
        self.assertIn("overflow: hidden", css)
        self.assertIn(".activity-log-page .activity-log-table .col-action", css)
        self.assertIn(".activity-log-page .activity-log-table .col-severity", css)
        self.assertIn(".activity-log-page .activity-log-table .col-actor", css)

    def test_public_routes_smoke_do_not_500(self):
        root_response = self.client.get("/", follow_redirects=False)
        self.assertNotEqual(root_response.status_code, 500)
        self.assertIn(root_response.status_code, (200, 302, 401))

        health_response = self.client.get("/health")
        self.assertEqual(health_response.status_code, 200)
        self.assertTrue(health_response.is_json)
        self.assertEqual(health_response.get_json()["status"], "ok")

        login_response = self.client.get("/login")
        self.assertEqual(login_response.status_code, 200)
        self.assertIn("csrf-token", login_response.get_data(as_text=True))

        missing_response = self.client.get("/route-smoke-missing", follow_redirects=False)
        self.assertEqual(missing_response.status_code, 404)
        self.assertNotIn("Traceback", missing_response.get_data(as_text=True))

    def test_authenticated_main_pages_smoke_render_key_markers(self):
        owner = AuthService.seed_owner_if_empty()
        self.login_as(owner)

        smoke_pages = [
            ("/", "Lịch hẹn hôm nay"),
            ("/customers", "Danh sách khách hàng"),
            ("/services", "Danh sách dịch vụ"),
            ("/appointments", "Danh sách lịch hẹn"),
            ("/invoices", "Danh sách hóa đơn"),
            ("/statistics", "Thống kê báo cáo"),
            ("/activity-logs", "Nhật ký hoạt động"),
            ("/recycle-bin", "Thùng rác"),
            ("/settings", "Cài đặt"),
            ("/profile", "Hồ sơ cá nhân"),
        ]

        for url, marker in smoke_pages:
            with self.subTest(url=url):
                response = self.client.get(url, follow_redirects=False)
                self.assertEqual(response.status_code, 200)
                self.assertIn(marker, response.get_data(as_text=True))

    def test_empty_state_routes_smoke_render_without_500(self):
        owner = AuthService.seed_owner_if_empty()
        self.login_as(owner)

        smoke_checks = [
            ("/customers", "Chưa có khách hàng nào"),
            ("/services", "Không tìm thấy dịch vụ nào"),
            ("/appointments", "Không tìm thấy lịch hẹn nào phù hợp."),
            ("/invoices", "Đang hiển thị: 0 hóa đơn"),
            ("/statistics", "statistics-page"),
            ("/statistics", "statistics-report-card"),
            ("/activity-logs", "Không tìm thấy nhật ký phù hợp"),
            ("/recycle-bin", "Thùng rác đang trống"),
        ]

        for url, marker in smoke_checks:
            with self.subTest(url=url, marker=marker):
                response = self.client.get(url, follow_redirects=False)
                self.assertEqual(response.status_code, 200)
                self.assertIn(marker, response.get_data(as_text=True))

    def test_permission_matrix_smoke_for_admin_pages(self):
        owner = self.create_user("perm-owner", password="owner-pass", full_name="Perm Owner", role="OWNER")
        admin = self.create_user("perm-admin", password="admin-pass", full_name="Perm Admin", role="ADMIN")
        staff = self.create_user("perm-staff", password="staff-pass", full_name="Perm Staff", role="STAFF")

        sensitive_pages = [
            ("/settings", "Cài đặt"),
            ("/users", "Người dùng"),
            ("/activity-logs", "Nhật ký hoạt động"),
            ("/recycle-bin", "Thùng rác"),
            ("/statistics", "Thống kê báo cáo"),
        ]

        for user, expected_status in ((owner, 200), (admin, 200)):
            self.login_as(user)
            for url, marker in sensitive_pages:
                with self.subTest(role=user.role, url=url):
                    response = self.client.get(url, follow_redirects=False)
                    self.assertEqual(response.status_code, expected_status)
                    self.assertIn(marker, response.get_data(as_text=True))

        self.login_as(staff)
        blocked_pages = [
            "/settings",
            "/users",
            "/activity-logs",
            "/recycle-bin",
            "/statistics",
        ]
        for url in blocked_pages:
            with self.subTest(role=staff.role, url=url):
                response = self.client.get(url, follow_redirects=False)
                self.assertIn(response.status_code, (401, 403))

    def test_docs_and_readme_include_qa_checklist_links(self):
        readme = Path("README.md").read_text(encoding="utf-8")
        docs_readme = Path("docs/README.md").read_text(encoding="utf-8")

        self.assertIn("docs/QA_CHECKLIST.md", readme)
        self.assertIn("QA_CHECKLIST.md", docs_readme)
        self.assertIn("RUNBOOK.md", docs_readme)
        self.assertIn("USER_GUIDE.md", docs_readme)
        self.assertIn("ADMIN_GUIDE.md", docs_readme)
        self.assertIn("DEMO_SCRIPT.md", docs_readme)
        self.assertNotIn("docs/AUDIT_REPORT_v3.7.md", docs_readme)
        self.assertNotIn("docs/AUTH_AUDIT_v3.9.md", docs_readme)

    def test_db_stamp_head_marks_existing_schema_without_rebuilding(self):
        tables_before = sorted(sa_inspect(db.engine).get_table_names())
        runner = app.test_cli_runner()

        result = runner.invoke(args=["db", "stamp", "head"])
        self.assertEqual(result.exit_code, 0, result.output)
        self.assertIn("Stamped 0002_google_auth_approval", result.output)

        tables_after = sorted(sa_inspect(db.engine).get_table_names())
        self.assertIn("alembic_version", tables_after)
        self.assertEqual(tables_before, [table for table in tables_after if table != "alembic_version"])

        current_after = runner.invoke(args=["db", "current"])
        self.assertEqual(current_after.exit_code, 0, current_after.output)
        self.assertIn("0002_google_auth_approval", current_after.output)

    def test_data_consistency_audit_passes_on_clean_database(self):
        report = run_data_consistency_audit()

        self.assertTrue(report.passed)
        self.assertEqual(report.total_errors, 0)
        self.assertEqual(report.total_warnings, 0)
        self.assertIn("Data consistency audit", report.to_text())
        self.assertIn("Status: PASS", report.to_text())

    def test_data_consistency_audit_detects_real_database_issues_without_writing(self):
        customer_a = Customer(name=" ", phone="0901234567", email="duplicate@example.com")
        customer_b = Customer(name="Customer B", phone="0901234567", email="duplicate@example.com", deleted_by="owner")
        customer_trim = Customer(name="  Trim Customer  ", phone=" 0905555555 ", email=" trimcase@example.com ")
        customer_c = Customer(name="Linked Customer", phone="0909999999", email="linked@example.com")
        service_a = Service(name=" ", price=-150000, duration=30, description="Bad service", category="other", deleted_at=datetime.utcnow())
        service_trim = Service(name="  Trim Service  ", price=150000, duration=30, description="Trim me", category="other")
        service_b = Service(name="Linked Service", price=200000, duration=45, description="Good service", category="other")
        db.session.add_all([customer_a, customer_b, customer_trim, customer_c, service_a, service_trim, service_b])
        db.session.commit()

        appointment = Appointment(
            customer_id=customer_c.id,
            service_id=service_a.id,
            appointment_time=datetime(2026, 7, 4, 10, 0),
            status="Bogus",
            deleted_at=datetime.utcnow(),
        )
        invoice = Invoice(
            customer_id=customer_c.id,
            invoice_date=None,
            subtotal=0,
            discount=0,
            total_amount=-50000,
            payment_method="UnknownMethod",
            deleted_by="owner",
        )
        db.session.add_all([appointment, invoice])
        db.session.commit()

        detail = InvoiceDetail(
            invoice_id=invoice.id,
            service_id=service_a.id,
            price=-1000,
            quantity=0,
        )
        db.session.add(detail)
        db.session.commit()

        self.execute_raw_sql(
            """
            INSERT INTO appointments (customer_id, service_id, appointment_time, status, notes, created_at, deleted_at, deleted_by)
            VALUES (:customer_id, :service_id, :appointment_time, :status, NULL, :created_at, NULL, NULL)
            """,
            customer_id=999999,
            service_id=888888,
            appointment_time=datetime(2026, 7, 5, 11, 30),
            status="Pending",
            created_at=datetime.utcnow(),
        )
        self.execute_raw_sql(
            """
            INSERT INTO invoices (customer_id, invoice_date, subtotal, discount, total_amount, payment_method, notes, created_at, deleted_at, deleted_by)
            VALUES (:customer_id, :invoice_date, :subtotal, :discount, :total_amount, :payment_method, NULL, :created_at, NULL, NULL)
            """,
            customer_id=999998,
            invoice_date=None,
            subtotal=0,
            discount=0,
            total_amount=1000,
            payment_method="Cash",
            created_at=datetime.utcnow(),
        )
        self.execute_raw_sql(
            """
            INSERT INTO invoice_details (invoice_id, service_id, price, quantity)
            VALUES (:invoice_id, :service_id, :price, :quantity)
            """,
            invoice_id=999997,
            service_id=888887,
            price=1000,
            quantity=1,
        )

        before_snapshot = (
            Customer.query.count(),
            Service.query.count(),
            Appointment.query.count(),
            Invoice.query.count(),
            InvoiceDetail.query.count(),
            ActivityLog.query.count(),
        )

        report = run_data_consistency_audit()

        after_snapshot = (
            Customer.query.count(),
            Service.query.count(),
            Appointment.query.count(),
            Invoice.query.count(),
            InvoiceDetail.query.count(),
            ActivityLog.query.count(),
        )

        self.assertEqual(before_snapshot, after_snapshot)
        self.assertFalse(report.passed)

        issue_codes = {issue.code for issue in report.issues}
        expected_codes = {
            "CUSTOMER_EMPTY_NAME",
            "CUSTOMER_DUPLICATE_PHONE",
            "CUSTOMER_DUPLICATE_EMAIL",
            "CUSTOMER_TRIM_NAME",
            "CUSTOMER_TRIM_PHONE",
            "CUSTOMER_TRIM_EMAIL",
            "CUSTOMER_SOFT_DELETE_MISMATCH",
            "SERVICE_EMPTY_NAME",
            "SERVICE_TRIM_NAME",
            "SERVICE_NEGATIVE_PRICE",
            "SERVICE_SOFT_DELETE_MISMATCH",
            "APPOINTMENT_MISSING_CUSTOMER",
            "APPOINTMENT_MISSING_SERVICE",
            "APPOINTMENT_INVALID_STATUS",
            "APPOINTMENT_SOFT_DELETE_MISMATCH",
            "APPOINTMENT_SOFT_DELETED_SERVICE",
            "INVOICE_MISSING_CUSTOMER",
            "INVOICE_EMPTY_DATE",
            "INVOICE_NEGATIVE_TOTAL",
            "INVOICE_INVALID_PAYMENT_METHOD",
            "INVOICE_SOFT_DELETE_MISMATCH",
            "INVOICE_DETAIL_MISSING_INVOICE",
            "INVOICE_DETAIL_MISSING_SERVICE",
            "INVOICE_DETAIL_INVALID_QUANTITY",
            "INVOICE_DETAIL_NEGATIVE_PRICE",
            "INVOICE_DETAIL_SOFT_DELETED_SERVICE",
        }
        self.assertTrue(expected_codes.issubset(issue_codes), issue_codes)

    def test_data_consistency_audit_detects_missing_appointment_time_from_fake_session(self):
        soft_deleted_at = datetime.utcnow()
        fake_customers = [
            SimpleNamespace(id=1, name=" ", phone="0901111111", email="dup@example.com", deleted_at=None, deleted_by=None),
            SimpleNamespace(id=2, name="Customer A", phone="0901111111", email="dup@example.com", deleted_at=None, deleted_by=None),
            SimpleNamespace(id=3, name="Soft Deleted Customer", phone="0902222222", email="customer3@example.com", deleted_at=soft_deleted_at, deleted_by=None),
        ]
        fake_services = [
            SimpleNamespace(id=1, name=" ", price=-10, duration=30, description=None, category="other", deleted_at=soft_deleted_at, deleted_by=None),
            SimpleNamespace(id=2, name="Service A", price=100000, duration=30, description=None, category="other", deleted_at=None, deleted_by=None),
        ]
        fake_appointments = [
            SimpleNamespace(id=1, customer_id=999, service_id=998, appointment_time=None, status="Bogus", deleted_at=soft_deleted_at, deleted_by=None),
            SimpleNamespace(id=2, customer_id=3, service_id=1, appointment_time=datetime.utcnow(), status="Pending", deleted_at=None, deleted_by=None),
        ]
        fake_invoices = [
            SimpleNamespace(id=1, customer_id=999, invoice_date=None, subtotal=0, discount=0, total_amount=-1, payment_method="Weird", deleted_at=soft_deleted_at, deleted_by=None),
            SimpleNamespace(id=2, customer_id=3, invoice_date=datetime(2026, 7, 4).date(), subtotal=0, discount=0, total_amount=1000, payment_method="Cash", deleted_at=None, deleted_by=None),
        ]
        fake_details = [
            SimpleNamespace(id=1, invoice_id=999, service_id=998, price=-100, quantity=0),
            SimpleNamespace(id=2, invoice_id=1, service_id=1, price=100, quantity=1),
        ]

        class FakeQuery:
            def __init__(self, rows):
                self.rows = rows

            def all(self):
                return list(self.rows)

        class FakeSession:
            def __init__(self, mapping):
                self.mapping = mapping
                self.no_autoflush = nullcontext()

            def query(self, model):
                return FakeQuery(self.mapping.get(model, []))

        fake_session = FakeSession({
            Customer: fake_customers,
            Service: fake_services,
            Appointment: fake_appointments,
            Invoice: fake_invoices,
            InvoiceDetail: fake_details,
        })

        report = run_data_consistency_audit(fake_session)
        issue_codes = {issue.code for issue in report.issues}

        self.assertFalse(report.passed)
        self.assertIn("APPOINTMENT_EMPTY_TIME", issue_codes)
        self.assertIn("APPOINTMENT_INVALID_STATUS", issue_codes)
        self.assertIn("CUSTOMER_DUPLICATE_PHONE", issue_codes)
        self.assertIn("INVOICE_DETAIL_NEGATIVE_PRICE", issue_codes)
        self.assertIn("Status: FAIL", report.to_text())

    def test_data_consistency_audit_cli_command_runs(self):
        runner = app.test_cli_runner()
        before_snapshot = (
            Customer.query.count(),
            Service.query.count(),
            Appointment.query.count(),
            Invoice.query.count(),
            InvoiceDetail.query.count(),
            ActivityLog.query.count(),
        )
        result = runner.invoke(args=["data", "audit"])
        after_snapshot = (
            Customer.query.count(),
            Service.query.count(),
            Appointment.query.count(),
            Invoice.query.count(),
            InvoiceDetail.query.count(),
            ActivityLog.query.count(),
        )

        self.assertEqual(result.exit_code, 0, result.output)
        self.assertIn("Data consistency audit", result.output)
        self.assertIn("Status: PASS", result.output)
        self.assertIn("Errors: 0", result.output)
        self.assertEqual(before_snapshot, after_snapshot)

    def test_performance_profile_service_reports_metrics_without_mutating_database(self):
        customer = self.create_customer_record(name="Perf Customer")
        service = self.create_service_record(name="Perf Service")
        self.create_appointment_record(customer=customer, service=service)
        self.create_invoice_record(customer=customer, service=service)

        before_snapshot = (
            Customer.query.count(),
            Service.query.count(),
            Appointment.query.count(),
            Invoice.query.count(),
            InvoiceDetail.query.count(),
            ActivityLog.query.count(),
        )

        report = run_performance_profile()

        after_snapshot = (
            Customer.query.count(),
            Service.query.count(),
            Appointment.query.count(),
            Invoice.query.count(),
            InvoiceDetail.query.count(),
            ActivityLog.query.count(),
        )

        self.assertGreaterEqual(report.total_duration_ms, 0)
        self.assertGreaterEqual(report.total_query_count, 0)
        self.assertGreater(len(report.metrics), 0)
        self.assertIn("Customers", report.dataset)
        self.assertIn("Invoices", report.dataset)
        self.assertEqual(before_snapshot, after_snapshot)

    def test_performance_profile_cli_command_runs_and_listener_cleanup_is_stable(self):
        runner = app.test_cli_runner()
        result = runner.invoke(args=["perf", "profile"])
        self.assertEqual(result.exit_code, 0, result.output)
        self.assertIn("Performance profile", result.output)
        self.assertIn("Total duration", result.output)
        self.assertIn("Total queries", result.output)
        self.assertIn("Dataset:", result.output)
        self.assertIn("Metrics:", result.output)

        metric_one = profile_block("listener.cleanup", lambda: Customer.query.count())
        metric_two = profile_block("listener.cleanup", lambda: Customer.query.count())
        self.assertEqual(metric_one.query_count, 1)
        self.assertEqual(metric_two.query_count, 1)
        self.assertEqual(metric_one.status, "OK")
        self.assertEqual(metric_two.status, "OK")

    def test_operational_diagnostics_service_reports_sections_without_mutating_database(self):
        customer = self.create_customer_record(name="Ops Customer")
        service = self.create_service_record(name="Ops Service")
        self.create_appointment_record(customer=customer, service=service)
        self.create_invoice_record(customer=customer, service=service)

        before_snapshot = (
            Customer.query.count(),
            Service.query.count(),
            Appointment.query.count(),
            Invoice.query.count(),
            InvoiceDetail.query.count(),
            ActivityLog.query.count(),
        )

        with patch("services.operational_diagnostics_service._load_backup_metadata", return_value={}):
            report = run_operational_diagnostics()

        after_snapshot = (
            Customer.query.count(),
            Service.query.count(),
            Appointment.query.count(),
            Invoice.query.count(),
            InvoiceDetail.query.count(),
            ActivityLog.query.count(),
        )

        self.assertIn(report.status, {"OK", "WARN", "FAIL"})
        self.assertIn("name", report.app)
        self.assertIn("type", report.database)
        self.assertEqual(report.backup["count"], 0)
        self.assertEqual(report.repair["mode"], "DRY-RUN")
        self.assertEqual(before_snapshot, after_snapshot)
        self.assertIn("Operational diagnostics", report.to_text())
        self.assertIn("App", report.to_text())
        self.assertIn("Database", report.to_text())
        self.assertIn("Backup", report.to_text())
        self.assertIn("Data audit", report.to_text())
        self.assertIn("Repair dry-run", report.to_text())
        self.assertIn("Performance", report.to_text())

    def test_operational_diagnostics_includes_security_account_summary(self):
        owner = self.create_user("ops-owner", password="owner-pass", full_name="Ops Owner", role="OWNER")
        admin = self.create_user("ops-admin", password="admin-pass", full_name="Ops Admin", role="ADMIN")
        staff_active = self.create_user("ops-staff", password="staff-pass", full_name="Ops Staff", role="STAFF")
        staff_inactive = self.create_user("ops-staff-inactive", password="staff-pass-2", full_name="Ops Staff Inactive", role="STAFF")
        staff_inactive.is_active = False
        db.session.commit()

        for action in ["AUTH_LOGIN_FAILED", "AUTH_LOGIN_FAILED", "AUTH_LOGIN_RATE_LIMITED", "CHANGE_PASSWORD", "RESET_USER_PASSWORD"]:
            log_entry = build_activity_log_entry(
                module="Auth",
                action=action,
                severity="WARNING",
                description=f"{action} telemetry",
                user_id=owner.id,
            )
            log_entry.created_at = datetime.utcnow()
            db.session.add(log_entry)
        db.session.commit()

        before_snapshot = (User.query.count(), ActivityLog.query.count())
        with patch("services.operational_diagnostics_service._load_backup_metadata", return_value={}):
            report = run_operational_diagnostics(include_performance=False, include_repair_plan=False)
        after_snapshot = (User.query.count(), ActivityLog.query.count())

        self.assertEqual(before_snapshot, after_snapshot)
        self.assertIn("security", report.to_dict())
        self.assertEqual(report.security["total_users"], 4)
        self.assertEqual(report.security["active_users"], 3)
        self.assertEqual(report.security["inactive_users"], 1)
        self.assertEqual(report.security["owners_total"], 1)
        self.assertEqual(report.security["owners_active"], 1)
        self.assertEqual(report.security["admins_total"], 1)
        self.assertEqual(report.security["admins_active"], 1)
        self.assertEqual(report.security["staff_total"], 2)
        self.assertEqual(report.security["staff_active"], 1)
        self.assertEqual(report.security["recent_login_failed_count"], 2)
        self.assertEqual(report.security["recent_login_rate_limited_count"], 1)
        self.assertEqual(report.security["recent_password_change_count"], 1)
        self.assertEqual(report.security["recent_password_reset_count"], 1)
        self.assertIn("Only one active OWNER account exists.", report.security["warnings"])
        self.assertIn("Inactive users exist in the database.", report.security["warnings"])
        self.assertIn("Login rate-limited events were detected in the last 24h.", report.security["warnings"])
        self.assertIn("Security / Accounts", report.to_text())
        self.assertIn("Login failed", report.to_text())
        self.assertIn("Login rate-limited", report.to_text())
        self.assertNotIn("owner-pass", report.to_text())
        self.assertNotIn("staff-pass", report.to_text())

    def test_operational_diagnostics_flags_invalid_roles_and_no_active_owner(self):
        owner = self.create_user("ops-owner-fail", password="owner-pass", full_name="Ops Owner Fail", role="OWNER")
        owner.is_active = False
        db.session.commit()
        admin = self.create_user("ops-admin-fail", password="admin-pass", full_name="Ops Admin Fail", role="ADMIN")
        invalid_user = self.create_user("ops-invalid-role", password="invalid-pass", full_name="Ops Invalid", role="STAFF")
        with db.engine.begin() as connection:
            connection.execute(
                text("UPDATE users SET role = :role WHERE id = :user_id"),
                {"role": "HACKER", "user_id": invalid_user.id},
            )

        before_snapshot = (User.query.count(), ActivityLog.query.count())
        with patch("services.operational_diagnostics_service._load_backup_metadata", return_value={}):
            report = run_operational_diagnostics(include_performance=False, include_repair_plan=False)
        after_snapshot = (User.query.count(), ActivityLog.query.count())

        self.assertEqual(before_snapshot, after_snapshot)
        self.assertEqual(report.status, "FAIL")
        self.assertEqual(report.security["owners_active"], 0)
        self.assertEqual(report.security["invalid_role_users"], 1)
        self.assertIn("No active OWNER account exists.", report.security["warnings"])
        self.assertIn("Invalid role values were detected.", report.security["warnings"])
        self.assertNotIn("HACKER", report.to_text())
        self.assertNotIn("owner-pass", report.to_text())

    def test_operational_diagnostics_cli_command_runs_and_skip_flags_work(self):
        runner = app.test_cli_runner()
        with patch("services.operational_diagnostics_service._load_backup_metadata", return_value={}):
            result = runner.invoke(args=["ops", "diagnostics", "--skip-performance"])

        self.assertEqual(result.exit_code, 0, result.output)
        self.assertIn("Operational diagnostics", result.output)
        self.assertIn("App", result.output)
        self.assertIn("Database", result.output)
        self.assertIn("Backup", result.output)
        self.assertIn("Data audit", result.output)
        self.assertIn("Repair dry-run", result.output)
        self.assertIn("Performance", result.output)
        self.assertIn("SKIPPED", result.output)

    def test_operational_diagnostics_cli_outputs_security_section_without_sensitive_data(self):
        owner = self.create_user("ops-cli-owner", password="owner-pass", full_name="Ops CLI Owner", role="OWNER")
        log_entry = build_activity_log_entry(
            module="Auth",
            action="AUTH_LOGIN_FAILED",
            severity="WARNING",
            description="AUTH_LOGIN_FAILED telemetry",
            user_id=owner.id,
        )
        log_entry.created_at = datetime.utcnow()
        db.session.add(log_entry)
        db.session.commit()

        runner = app.test_cli_runner()
        with patch("services.operational_diagnostics_service._load_backup_metadata", return_value={}):
            result = runner.invoke(args=["ops", "report", "--skip-performance", "--skip-repair-plan"])

        self.assertEqual(result.exit_code, 0, result.output)
        self.assertIn("Security / Accounts", result.output)
        self.assertIn("Login failed", result.output)
        self.assertIn("OWNER total", result.output)
        self.assertIn("ADMIN total", result.output)
        self.assertIn("STAFF total", result.output)
        self.assertNotIn("owner-pass", result.output)
        self.assertNotIn("password_hash", result.output)
        self.assertNotIn("SECRET_KEY", result.output)
        self.assertNotIn("token", result.output.lower())

    def test_operational_diagnostics_repair_plan_stays_dry_run(self):
        customer = Customer(name="  Ops Repair Customer  ", phone=" 0901234567 ", email=" ops-repair@example.com ")
        service = Service(name="  Ops Repair Service  ", price=100000, duration=30, description="Repair", category="other")
        db.session.add_all([customer, service])
        db.session.commit()

        before_snapshot = (
            Customer.query.get(customer.id).name,
            Customer.query.get(customer.id).phone,
            Customer.query.get(customer.id).email,
            Service.query.get(service.id).name,
        )

        with patch("services.operational_diagnostics_service._load_backup_metadata", return_value={}):
            report = run_operational_diagnostics(verbose=True)

        after_snapshot = (
            Customer.query.get(customer.id).name,
            Customer.query.get(customer.id).phone,
            Customer.query.get(customer.id).email,
            Service.query.get(service.id).name,
        )

        self.assertEqual(report.repair["mode"], "DRY-RUN")
        self.assertGreater(report.repair["repairable_actions"], 0)
        self.assertEqual(before_snapshot, after_snapshot)
        self.assertIn("CUSTOMER_TRIM_NAME", report.repair.get("top_action_codes", []))
        self.assertIn("SERVICE_TRIM_NAME", report.repair.get("top_action_codes", []))

    def test_data_repair_dry_run_reports_safe_actions_without_writing(self):
        customer = Customer(name="  Repair Customer  ", phone=" 0901234567 ", email="  repair@example.com  ", deleted_at=datetime.utcnow())
        service = Service(name="  Repair Service  ", price=100000, duration=30, description="Repair", category="other")
        db.session.add_all([customer, service])
        db.session.commit()

        before_snapshot = (
            Customer.query.get(customer.id).name,
            Customer.query.get(customer.id).phone,
            Customer.query.get(customer.id).email,
            Customer.query.get(customer.id).deleted_by,
            Service.query.get(service.id).name,
        )

        report = run_controlled_repair(dry_run=True)

        after_snapshot = (
            Customer.query.get(customer.id).name,
            Customer.query.get(customer.id).phone,
            Customer.query.get(customer.id).email,
            Customer.query.get(customer.id).deleted_by,
            Service.query.get(service.id).name,
        )

        self.assertTrue(report.dry_run)
        self.assertEqual(report.mode, "DRY-RUN")
        self.assertGreaterEqual(report.repairable_actions, 3)
        self.assertEqual(report.applied_count, 0)
        self.assertEqual(before_snapshot, after_snapshot)
        self.assertIn("CUSTOMER_TRIM_NAME", report.to_text())
        self.assertIn("SERVICE_TRIM_NAME", report.to_text())

    def test_data_repair_apply_requires_yes_on_cli(self):
        customer = Customer(name="  Apply Customer  ", phone=" 0911222333 ", email="  apply@example.com  ", deleted_at=datetime.utcnow())
        db.session.add(customer)
        db.session.commit()

        runner = app.test_cli_runner()
        before_name = Customer.query.get(customer.id).name
        before_phone = Customer.query.get(customer.id).phone
        before_email = Customer.query.get(customer.id).email
        before_deleted_by = Customer.query.get(customer.id).deleted_by

        result = runner.invoke(args=["data", "repair", "--apply"])

        after_customer = Customer.query.get(customer.id)
        self.assertEqual(result.exit_code, 0, result.output)
        self.assertIn("--yes", result.output)
        self.assertEqual(after_customer.name, before_name)
        self.assertEqual(after_customer.phone, before_phone)
        self.assertEqual(after_customer.email, before_email)
        self.assertEqual(after_customer.deleted_by, before_deleted_by)

    def test_data_repair_apply_safely_trims_and_fills_deleted_by(self):
        customer = Customer(name="  Trim Me  ", phone=" 0908888888 ", email="  trim@example.com  ", deleted_at=datetime.utcnow())
        service = Service(name="  Trim Service  ", price=120000, duration=45, description="Repair", category="other")
        db.session.add_all([customer, service])
        db.session.commit()

        report = run_controlled_repair(dry_run=False)

        repaired_customer = Customer.query.get(customer.id)
        repaired_service = Service.query.get(service.id)
        self.assertFalse(report.dry_run)
        self.assertEqual(report.mode, "APPLY")
        self.assertGreaterEqual(report.applied_count, 4)
        self.assertEqual(repaired_customer.name, "Trim Me")
        self.assertEqual(repaired_customer.phone, "0908888888")
        self.assertEqual(repaired_customer.email, "trim@example.com")
        self.assertEqual(repaired_customer.deleted_by, "Hệ thống")
        self.assertEqual(repaired_service.name, "Trim Service")

    def test_data_repair_cli_defaults_to_dry_run_and_apply_yes_changes_data(self):
        customer = Customer(name="  CLI Customer  ", phone=" 0906666666 ", email="  cli@example.com  ", deleted_at=datetime.utcnow())
        db.session.add(customer)
        db.session.commit()

        runner = app.test_cli_runner()
        dry_run_result = runner.invoke(args=["data", "repair"])
        self.assertEqual(dry_run_result.exit_code, 0, dry_run_result.output)
        self.assertIn("Mode: DRY-RUN", dry_run_result.output)
        self.assertEqual(Customer.query.get(customer.id).name, "  CLI Customer  ")

        apply_result = runner.invoke(args=["data", "repair", "--apply", "--yes"])
        self.assertEqual(apply_result.exit_code, 0, apply_result.output)
        self.assertIn("Mode: APPLY", apply_result.output)
        self.assertEqual(Customer.query.get(customer.id).name, "CLI Customer")
        self.assertEqual(Customer.query.get(customer.id).phone, "0906666666")
        self.assertEqual(Customer.query.get(customer.id).email, "cli@example.com")
        self.assertEqual(Customer.query.get(customer.id).deleted_by, "Hệ thống")

    def test_data_repair_skips_manual_issues_and_leaves_database_unchanged(self):
        duplicate_customer_a = Customer(name="Duplicate A", phone="0909999999", email="duplicate@example.com")
        duplicate_customer_b = Customer(name="Duplicate B", phone="0909999999", email="duplicate2@example.com")
        customer = Customer(name="Manual Customer", phone="0907777777", email="manual@example.com")
        service = Service(name="Manual Service", price=100000, duration=30, description="Repair", category="other")
        db.session.add_all([duplicate_customer_a, duplicate_customer_b, customer, service])
        db.session.commit()

        self.execute_raw_sql(
            """
            INSERT INTO appointments (customer_id, service_id, appointment_time, status, notes, created_at, deleted_at, deleted_by)
            VALUES (:customer_id, :service_id, :appointment_time, :status, NULL, :created_at, NULL, NULL)
            """,
            customer_id=999999,
            service_id=service.id,
            appointment_time=datetime(2026, 7, 5, 11, 30),
            status="Pending",
            created_at=datetime.utcnow(),
        )

        before_snapshot = (
            Customer.query.count(),
            Service.query.count(),
            Appointment.query.count(),
            Invoice.query.count(),
            InvoiceDetail.query.count(),
        )

        report = run_controlled_repair(dry_run=False)

        after_snapshot = (
            Customer.query.count(),
            Service.query.count(),
            Appointment.query.count(),
            Invoice.query.count(),
            InvoiceDetail.query.count(),
        )

        self.assertFalse(report.dry_run)
        self.assertEqual(before_snapshot, after_snapshot)
        self.assertTrue(any(item.code.startswith("CUSTOMER_DUPLICATE") for item in report.skipped))
        self.assertTrue(all(item.code.startswith("APPOINTMENT_") or item.code.startswith("INVOICE_") or item.code.startswith("CUSTOMER_DUPLICATE") or item.code.endswith("MISMATCH") for item in report.skipped))

    def extract_pdf_text(self, pdf_bytes):
        font_cmaps = {}
        font_resources = {}

        for obj_match in re.finditer(rb'(\d+)\s+0\s+obj(.*?)endobj', pdf_bytes, re.S):
            obj_num = int(obj_match.group(1))
            body = obj_match.group(2)
            if b'/Subtype /Type0' not in body and b'/Subtype /TrueType' not in body:
                continue

            name_match = re.search(rb'/Name\s*/([^\s/]+)', body)
            to_unicode_match = re.search(rb'/ToUnicode\s+(\d+)\s+0\s+R', body)
            if not name_match or not to_unicode_match:
                continue

            font_name = name_match.group(1).decode('latin1')
            cmap_obj_num = int(to_unicode_match.group(1))
            font_resources[font_name] = cmap_obj_num

        for cmap_match in re.finditer(rb'(\d+)\s+0\s+obj(.*?)endobj', pdf_bytes, re.S):
            obj_num = int(cmap_match.group(1))
            body = cmap_match.group(2)
            if obj_num not in font_resources.values():
                continue
            stream_match = re.search(rb'stream\r?\n(.*?)endstream', body, re.S)
            if not stream_match:
                continue
            decoded = zlib.decompress(stream_match.group(1).strip(b'\r\n'))
            cmap = {}
            for line in decoded.splitlines():
                line = line.strip()
                entry = re.match(rb'<([0-9A-Fa-f]+)>\s+<([0-9A-Fa-f]+)>', line)
                if not entry:
                    continue
                source_code = int(entry.group(1), 16)
                target_code = int(entry.group(2), 16)
                if target_code == 0:
                    continue
                cmap[source_code] = chr(target_code)
            font_cmaps[obj_num] = cmap

        extracted_parts = []
        for stream_match in re.finditer(rb'(\d+)\s+0\s+obj(.*?)endobj', pdf_bytes, re.S):
            body = stream_match.group(2)
            if b'stream' not in body:
                continue
            stream_match = re.search(rb'stream\r?\n(.*?)endstream', body, re.S)
            if not stream_match:
                continue
            stream_data = stream_match.group(1).strip(b'\r\n')
            try:
                decoded_stream = zlib.decompress(base64.a85decode(stream_data, adobe=True))
            except Exception:
                try:
                    decoded_stream = zlib.decompress(base64.a85decode(stream_data, adobe=False))
                except Exception:
                    try:
                        decoded_stream = zlib.decompress(stream_data)
                    except Exception:
                        continue

            if b'BT' not in decoded_stream or b'Tj' not in decoded_stream:
                continue

            current_font = None
            pos = 0
            while pos < len(decoded_stream):
                font_match = re.search(rb'/([A-Za-z0-9\+]+)\s+\d+(?:\.\d+)?\s+Tf', decoded_stream[pos:])
                text_match = re.search(rb'\((?:\\.|[^()])*\)\s+Tj', decoded_stream[pos:])

                next_pos = len(decoded_stream)
                next_kind = None
                if font_match:
                    next_pos = pos + font_match.start()
                    next_kind = "font"
                if text_match and pos + text_match.start() < next_pos:
                    next_pos = pos + text_match.start()
                    next_kind = "text"

                if next_kind is None:
                    break

                if next_kind == "font":
                    current_font = font_match.group(1).decode('latin1')
                    pos = pos + font_match.end()
                    continue

                literal = text_match.group(0)
                literal_text = literal[1:literal.rfind(b')')]
                literal_bytes = bytearray()
                idx = 0
                while idx < len(literal_text):
                    char = literal_text[idx:idx+1]
                    if char == b'\\':
                        idx += 1
                        if idx >= len(literal_text):
                            break
                        escaped = literal_text[idx:idx+1]
                        if escaped in {b'\\', b'(', b')'}:
                            literal_bytes.extend(escaped)
                            idx += 1
                            continue
                        if escaped in {b'n', b'r', b't', b'b', b'f'}:
                            translation = {b'n': b'\n', b'r': b'\r', b't': b'\t', b'b': b'\b', b'f': b'\f'}[escaped]
                            literal_bytes.extend(translation)
                            idx += 1
                            continue
                        octal = literal_text[idx:idx+3]
                        if re.fullmatch(rb'[0-7]{1,3}', octal):
                            literal_bytes.append(int(octal, 8))
                            idx += len(octal)
                            continue
                        literal_bytes.extend(escaped)
                        idx += 1
                        continue
                    literal_bytes.extend(char)
                    idx += 1

                if current_font:
                    cmap_obj_num = font_resources.get(current_font)
                    cmap = font_cmaps.get(cmap_obj_num, {})
                    extracted_parts.append(''.join(cmap.get(byte, chr(byte)) for byte in literal_bytes))
                pos = pos + text_match.end()

        return '\n'.join(extracted_parts)

    def test_pdf_font_helper_falls_back_without_crash(self):
        export_pdf_utils.reset_pdf_font_config_cache()
        missing_regular = Path(tempfile.gettempdir()) / f"missing-regular-{uuid.uuid4().hex}.ttf"
        missing_bold = Path(tempfile.gettempdir()) / f"missing-bold-{uuid.uuid4().hex}.ttf"

        with patch.object(export_pdf_utils, "_candidate_font_pairs", return_value=[(missing_regular, missing_bold)]):
            font_config = export_pdf_utils.get_pdf_font_config()

        self.assertTrue(font_config.fallback)
        self.assertEqual(font_config.regular, "Helvetica")
        self.assertEqual(font_config.bold, "Helvetica-Bold")
        export_pdf_utils.reset_pdf_font_config_cache()

    def test_pdf_font_helper_prefers_bundled_unicode_fonts(self):
        export_pdf_utils.reset_pdf_font_config_cache()
        font_config = export_pdf_utils.get_pdf_font_config()
        self.assertFalse(font_config.fallback)
        self.assertIn("assets\\fonts", font_config.regular_path.replace("/", "\\"))
        self.assertIn("assets\\fonts", font_config.bold_path.replace("/", "\\"))
        self.assertEqual(font_config.regular, "SpaUnicode")
        self.assertEqual(font_config.bold, "SpaUnicode-Bold")
        export_pdf_utils.reset_pdf_font_config_cache()

    def test_pdf_export_source_has_no_hardcoded_windows_font_path(self):
        source = Path(export_pdf_utils.__file__).read_text(encoding="utf-8")
        self.assertIn('Path(__file__).resolve().parents[1] / "assets" / "fonts"', source)
        self.assertNotIn("C:/Windows/Fonts", source)
        self.assertNotIn("C:\\Windows\\Fonts", source)

    def test_pdf_bundled_fonts_exist_and_have_content(self):
        regular_font = Path("assets/fonts/NotoSans-Regular.ttf")
        bold_font = Path("assets/fonts/NotoSans-Bold.ttf")
        self.assertTrue(regular_font.exists())
        self.assertTrue(bold_font.exists())
        self.assertGreater(regular_font.stat().st_size, 0)
        self.assertGreater(bold_font.stat().st_size, 0)
        git_tracked = subprocess.check_output(['git', 'ls-files', 'assets/fonts'], text=True)
        self.assertIn('assets/fonts/NotoSans-Regular.ttf', git_tracked)
        self.assertIn('assets/fonts/NotoSans-Bold.ttf', git_tracked)
        gitignore = Path('.gitignore').read_text(encoding='utf-8')
        self.assertNotIn('assets/', gitignore)
        self.assertNotIn('assets/fonts/', gitignore)
        self.assertNotIn('*.ttf', gitignore)

    def test_invoice_pdf_export_contains_vietnamese_text(self):
        owner = self.create_user("pdf-invoice-owner", password="owner-pass", full_name="PDF Invoice Owner", role="OWNER")
        self.login_as(owner)
        customer = self.create_customer_record(name="Khách hàng Trường")
        service = self.create_service_record(name="Dịch vụ Chăm sóc")
        invoice = self.create_invoice_record(customer=customer, service=service)
        invoice.display_payment_method = "Tiền mặt"

        response = self.client.get("/invoices/export/pdf", follow_redirects=False)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "application/pdf")
        self.assertIn("no-store", response.headers.get("Cache-Control", ""))
        self.assertIn("no-cache", response.headers.get("Cache-Control", ""))
        self.assertEqual(response.headers.get("Pragma"), "no-cache")
        self.assertEqual(response.headers.get("Expires"), "0")
        self.assertRegex(
            response.headers.get("Content-Disposition", ""),
            r'Danh_sach_hoa_don_\d{8}_\d{6}_\d{6}\.pdf',
        )
        self.assertIn(b"NotoSans", response.data)

        pdf_text = self.extract_pdf_text(response.data)
        self.assertIn("DANH SÁCH HÓA ĐƠN", pdf_text)
        self.assertIn("ĐIỀU KIỆN LỌC", pdf_text)
        self.assertIn("DANH SÁCH CHI TIẾT", pdf_text)
        self.assertIn("TỔNG CỘNG", pdf_text)
        self.assertIn("Khách hàng Trường", pdf_text)
        self.assertIn("Tiền mặt", pdf_text)

    def test_statistics_pdf_export_contains_vietnamese_text(self):
        owner = self.create_user("pdf-stat-owner", password="owner-pass", full_name="PDF Stat Owner", role="OWNER")
        self.login_as(owner)
        customer = self.create_customer_record(name="Khách hàng Thống kê")
        service = self.create_service_record(name="Dịch vụ Massage")
        invoice = self.create_invoice_record(customer=customer, service=service)
        invoice.display_payment_method = "Tiền mặt"

        response = self.client.get("/statistics/export/pdf", follow_redirects=False)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "application/pdf")
        self.assertIn("no-store", response.headers.get("Cache-Control", ""))
        self.assertIn("no-cache", response.headers.get("Cache-Control", ""))
        self.assertEqual(response.headers.get("Pragma"), "no-cache")
        self.assertEqual(response.headers.get("Expires"), "0")
        self.assertRegex(
            response.headers.get("Content-Disposition", ""),
            r'ThongKe_\d{8}_\d{6}_\d{6}\.pdf',
        )
        self.assertIn(b"NotoSans", response.data)

        pdf_text = self.extract_pdf_text(response.data)
        self.assertIn("BÁO CÁO THỐNG KÊ SPA", pdf_text)
        self.assertIn("THÔNG TIN TỔNG QUAN", pdf_text)
        self.assertIn("THỐNG KÊ KHÁCH HÀNG", pdf_text)
        self.assertIn("THỐNG KÊ DỊCH VỤ", pdf_text)
        self.assertIn("Khách hàng Thống kê", pdf_text)
        self.assertIn("Dịch vụ Massage", pdf_text)
        timestamp_match = re.search(r"Ngày xuất báo cáo:\s*(\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2})", pdf_text)
        self.assertIsNotNone(timestamp_match)
        pdf_timestamp = datetime.strptime(timestamp_match.group(1), "%d/%m/%Y %H:%M:%S")
        pdf_timestamp = pdf_timestamp.replace(tzinfo=local_now().tzinfo)
        self.assertLess(abs((local_now() - pdf_timestamp).total_seconds()), 60)

    def test_statistics_page_renders_clean_vietnamese_labels_and_filter_links(self):
        owner = self.create_user("stats-page-owner", password="owner-pass", full_name="Stats Page Owner", role="OWNER")
        self.login_as(owner)
        customer = self.create_customer_record(name="Khách hàng Báo cáo")
        service = self.create_service_record(name="Dịch vụ Báo cáo")
        self.create_invoice_record(customer=customer, service=service)

        response = self.client.get(
            "/statistics?from_date=2026-07-04&to_date=2026-07-04&cust_q=Báo&svc_q=Dịch",
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('statistics-page', html)
        self.assertIn('statistics-report-card', html)
        self.assertIn('statistics-report-heading', html)
        self.assertIn('statistics-report-title-row', html)
        self.assertIn('statistics-report-tools', html)
        self.assertIn('statistics-report-search', html)
        self.assertIn('statistics-report-actions', html)
        self.assertIn('statistics-table-scroll', html)
        self.assertIn('statistics-table-footer', html)
        self.assertIn('statistics-inline-search', html)
        self.assertIn("Thống kê báo cáo", html)
        self.assertIn("Bộ lọc thời gian", html)
        self.assertIn("Thống kê chi tiêu khách hàng", html)
        self.assertIn("Doanh thu theo dịch vụ", html)
        self.assertIn("Khách hàng Báo cáo", html)
        self.assertIn("Dịch vụ Báo cáo", html)
        self.assertIn("Top theo khoảng thời gian đã chọn", html)
        self.assertIn("Ẩn/Hiện cột", html)
        self.assertIn("Mỗi trang", html)
        self.assertNotIn("Ã", html)
        self.assertNotIn("á»", html)
        self.assertIn("/statistics/export/excel", html)
        self.assertIn("/statistics/export/pdf", html)
        self.assertIn("from_date=2026-07-04", html)
        self.assertIn("to_date=2026-07-04", html)

    def test_statistics_invalid_and_swapped_dates_do_not_break_report(self):
        owner = self.create_user("stats-date-owner", password="owner-pass", full_name="Stats Date Owner", role="OWNER")
        self.login_as(owner)
        customer = self.create_customer_record(name="Khách hàng Lọc ngày")
        service = self.create_service_record(name="Dịch vụ Lọc ngày")
        self.create_invoice_record(customer=customer, service=service)

        invalid_response = self.client.get("/statistics?from_date=bad-date&to_date=also-bad", follow_redirects=False)
        self.assertEqual(invalid_response.status_code, 200)

        swapped_response = self.client.get(
            "/statistics?from_date=2026-07-05&to_date=2026-07-04",
            follow_redirects=False,
        )
        self.assertEqual(swapped_response.status_code, 200)
        swapped_html = swapped_response.get_data(as_text=True)
        self.assertIn("Khách hàng Lọc ngày", swapped_html)
        self.assertIn("Dịch vụ Lọc ngày", swapped_html)
        self.assertIn("statistics-report-heading", swapped_html)
        self.assertIn("statistics-inline-search", swapped_html)

    def test_statistics_report_card_layout_is_simple_and_scoped(self):
        css = Path("static/css/pages/statistics.css").read_text(encoding="utf-8")

        self.assertIn("statistics-report-card", css)
        self.assertIn("statistics-report-heading", css)
        self.assertIn("statistics-report-title-row", css)
        self.assertIn("statistics-report-tools", css)
        self.assertIn("statistics-report-search", css)
        self.assertIn("statistics-report-actions", css)
        self.assertIn("statistics-table-scroll", css)
        self.assertIn("statistics-table-footer", css)
        self.assertNotIn("position: absolute", css)
        self.assertNotRegex(css, r"statistics-report-[^{]*margin-top:\s*-\d")
        self.assertNotRegex(css, r"statistics-table-[^{]*margin-top:\s*-\d")

    def test_statistics_export_excel_keeps_filters_and_cache_headers(self):
        owner = self.create_user("stats-excel-owner", password="owner-pass", full_name="Stats Excel Owner", role="OWNER")
        self.login_as(owner)
        customer = self.create_customer_record(name="Khách hàng Excel")
        service = self.create_service_record(name="Dịch vụ Excel")
        self.create_invoice_record(customer=customer, service=service)

        response = self.client.get(
            "/statistics/export/excel?from_date=2026-07-04&to_date=2026-07-04&cust_q=Excel&svc_q=Excel",
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.mimetype,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("no-store", response.headers.get("Cache-Control", ""))
        self.assertIn("no-cache", response.headers.get("Cache-Control", ""))
        self.assertEqual(response.headers.get("Pragma"), "no-cache")
        self.assertEqual(response.headers.get("Expires"), "0")
        self.assertRegex(
            response.headers.get("Content-Disposition", ""),
            r'ThongKe_\d{8}_\d{6}_\d{6}\.xlsx',
        )

        from openpyxl import load_workbook
        workbook = load_workbook(filename=BytesIO(response.data))
        sheet = workbook["Tổng quan"]
        self.assertEqual(sheet["A1"].value, "BÁO CÁO THỐNG KÊ SPA")
        self.assertIn("Khoảng thời gian: 04/07/2026 - 04/07/2026", sheet["A2"].value)
        self.assertIn("Ngày xuất báo cáo:", sheet["A3"].value)

    def test_approval_owner_lockdowns_and_bootstrap(self):
        owner = self.create_user("lockdown-owner", password="owner-pass", full_name="Lockdown Owner", role="OWNER")
        db.session.commit()

        self.login_as(owner)

        # 1. Role list in form creating user does not contain APPROVAL_OWNER
        create_page = self.client.get("/users/create")
        self.assertEqual(create_page.status_code, 200)
        self.assertNotIn("APPROVAL_OWNER", create_page.get_data(as_text=True))
        self.assertNotIn("Quản trị duyệt tài khoản", create_page.get_data(as_text=True))

        # 2. POST creating user with role APPROVAL_OWNER is rejected
        create_response = self.post_with_csrf(
            "/users/create",
            path="/users",
            data={
                "username": "illegal-approval-owner",
                "full_name": "Illegal Approval Owner",
                "email": "illegal.approval@example.com",
                "role": "APPROVAL_OWNER",
                "is_active": "1",
                "password": "valid_password_123",
                "confirm_password": "valid_password_123",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(create_response.status_code, 400)
        self.assertFalse(User.query.filter_by(username="illegal-approval-owner").first())

        # 3. POST editing user to APPROVAL_OWNER is rejected
        target_user = self.create_user("normal-staff-edit", password="staff-pass", role="STAFF")
        db.session.commit()

        edit_response = self.post_with_csrf(
            f"/users/{target_user.id}/edit",
            path=f"/users/{target_user.id}/edit",
            data={
                "username": "normal-staff-edit",
                "full_name": "Normal Staff Edit",
                "email": "normal.staff@example.com",
                "role": "APPROVAL_OWNER",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(edit_response.status_code, 400)
        db.session.refresh(target_user)
        self.assertEqual(target_user.role, "STAFF")

        # 4. /users listing does not display APPROVAL_OWNER
        approval_owner = self.create_user("real-approval-owner", password="owner-pass", role="APPROVAL_OWNER")
        db.session.commit()

        users_page = self.client.get("/users")
        self.assertEqual(users_page.status_code, 200)
        self.assertNotIn("real-approval-owner", users_page.get_data(as_text=True))

        # 5. /users/<id>/edit or /users/<id>/reset-password or /users/<id>/toggle-active returns 403
        self.assertEqual(self.client.get(f"/users/{approval_owner.id}/edit").status_code, 403)
        self.assertEqual(self.client.get(f"/users/{approval_owner.id}/reset-password").status_code, 403)

        toggle_response = self.post_with_csrf(
            f"/users/{approval_owner.id}/toggle-active",
            path="/users",
            data={"is_active": "0"},
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(toggle_response.status_code, 403)

        # 6. Bootstrap approval owner from environment variables works idempotently
        original_config = dict(app.config)
        try:
            with patch.dict(
                app.config,
                {
                    "APPROVAL_OWNER_USERNAME": "bootstrap-approval",
                    "APPROVAL_OWNER_PASSWORD": "bootstrap-password-123",
                    "APPROVAL_OWNER_EMAIL": "bootstrap@example.com",
                },
                clear=False,
            ):
                seeded = AuthService.seed_approval_owner_if_configured()
                self.assertIsNotNone(seeded)
                self.assertEqual(seeded.username, "bootstrap-approval")
                self.assertEqual(seeded.role, "APPROVAL_OWNER")
                self.assertTrue(seeded.is_active)
                self.assertEqual(seeded.approval_status, "active")

                # Test idempotency (doesn't duplicate or crash)
                seeded2 = AuthService.seed_approval_owner_if_configured()
                self.assertEqual(seeded2.id, seeded.id)

                # Test duplicate prevention / hijacking warning
                # Try to bootstrap with same username but conflicts with existing owner
                with patch.dict(
                    app.config,
                    {
                        "APPROVAL_OWNER_USERNAME": "lockdown-owner",  # Existing OWNER
                        "APPROVAL_OWNER_PASSWORD": "bootstrap-password-123",
                        "APPROVAL_OWNER_EMAIL": "newowneremail@example.com",
                    },
                    clear=False,
                ):
                    seeded_conflict = AuthService.seed_approval_owner_if_configured()
                    self.assertIsNone(seeded_conflict)
                    # Verify lockdown-owner's role was not changed to APPROVAL_OWNER
                    db.session.refresh(owner)
                    self.assertEqual(owner.role, "OWNER")

        finally:
            app.config.update(original_config)


if __name__ == "__main__":
    unittest.main()
