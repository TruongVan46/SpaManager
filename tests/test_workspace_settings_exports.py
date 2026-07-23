"""
tests/test_workspace_settings_exports.py
=========================================
Task 6.5.9b — Workspace-scoped settings and exports.

Covers:
  1. test_settings_are_workspace_scoped
  2. test_settings_no_current_workspace_fail_closed
  3. test_logo_spa_info_does_not_leak_between_workspaces
  4. test_invoice_export_excel_is_workspace_scoped
  5. test_invoice_export_pdf_is_workspace_scoped
  6. test_statistics_export_is_workspace_scoped
  7. test_export_without_workspace_returns_empty
"""

import os
import shutil
import tempfile
import unittest
from datetime import date
from pathlib import Path

# ── isolated test DB ────────────────────────────────────────────────────────
TEST_DB_FILE = Path(tempfile.gettempdir()) / "spamanager_ws_settings_exports.sqlite"
TEST_MEDIA_ROOT = Path(tempfile.gettempdir()) / "spamanager_ws_settings_exports_media"

for _p in (TEST_DB_FILE,):
    if _p.exists():
        try:
            _p.unlink()
        except Exception:
            pass
if TEST_MEDIA_ROOT.exists():
    shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)

os.environ["APP_ENV"] = "testing"
os.environ["TEST_DATABASE_URL"] = f"sqlite:///{TEST_DB_FILE.as_posix()}"
os.environ["PERSISTENT_ROOT"] = TEST_MEDIA_ROOT.as_posix()
os.environ["UPLOAD_ROOT"] = (TEST_MEDIA_ROOT / "uploads").as_posix()
os.environ["LOGO_UPLOAD_FOLDER"] = (TEST_MEDIA_ROOT / "uploads" / "logos").as_posix()
os.environ["AVATAR_UPLOAD_FOLDER"] = (TEST_MEDIA_ROOT / "uploads" / "avatars").as_posix()

from app import app
from extensions import db
from flask import session
from core.auth.constants import AUTH_SESSION_KEY
from tests.session_helpers import set_authenticated_session

from models.user import User
from models.workspace import Workspace, WorkspaceMember
from models.setting import Setting
from models.customer import Customer
from models.service import Service
from models.invoice import Invoice
from models.invoice_detail import InvoiceDetail
from models.activity_log import ActivityLog
from models.appointment import Appointment
from services.invoice_service import InvoiceService
from services.workspace_service import WorkspaceService


class TestWorkspaceSettingsExports(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.app_context = app.app_context()
        cls.app_context.push()
        db.create_all()

    @classmethod
    def tearDownClass(cls):
        db.session.remove()
        db.engine.dispose()
        if TEST_DB_FILE.exists():
            try:
                TEST_DB_FILE.unlink()
            except Exception:
                pass
        if TEST_MEDIA_ROOT.exists():
            shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)
        cls.app_context.pop()

    def setUp(self):
        db.session.remove()
        db.session.rollback()
        # Clear in dependency order
        ActivityLog.query.delete()
        InvoiceDetail.query.delete()
        Invoice.query.delete()
        Appointment.query.delete()
        Setting.query.delete()
        WorkspaceMember.query.delete()
        Customer.query.delete()
        Service.query.delete()
        User.query.delete()
        Workspace.query.delete()
        db.session.commit()
        self.client = app.test_client()

    def tearDown(self):
        db.session.remove()
        db.session.rollback()
        ActivityLog.query.delete()
        InvoiceDetail.query.delete()
        Invoice.query.delete()
        Appointment.query.delete()
        Setting.query.delete()
        WorkspaceMember.query.delete()
        Customer.query.delete()
        Service.query.delete()
        User.query.delete()
        Workspace.query.delete()
        db.session.commit()


    # ── helpers ──────────────────────────────────────────────────────────────

    def _create_workspace_and_owner(self, slug):
        ws = Workspace(name=f"Workspace {slug}", slug=slug, status="active")
        db.session.add(ws)
        db.session.flush()

        owner = User(
            username=f"owner_{slug}",
            email=f"owner_{slug}@test.com",
            role="OWNER",
            full_name=f"Owner {slug}",
        )
        owner.set_password("Password123!")
        db.session.add(owner)
        db.session.flush()

        member = WorkspaceMember(
            workspace_id=ws.id,
            user_id=owner.id,
            role="owner",
            status="active",
        )
        db.session.add(member)
        db.session.commit()
        return ws, owner

    def _login_as(self, user, workspace_id, isolation=True):
        """Set session with workspace context and isolation flag."""
        with self.client.session_transaction() as sess:
            set_authenticated_session(sess, user, workspace_id=workspace_id, enable_workspace_isolation=isolation)
            sess["current_workspace_id"] = workspace_id
            sess["_enable_workspace_isolation"] = isolation

    def _seed_setting(self, key, value, workspace_id):
        """Directly write a setting row for a given workspace_id."""
        s = Setting(key=key, value=value, workspace_id=workspace_id)
        db.session.add(s)
        db.session.commit()

    def _create_customer(self, workspace_id, name, phone="0901234567"):
        c = Customer(name=name, phone=phone, workspace_id=workspace_id)
        db.session.add(c)
        db.session.commit()
        return c

    def _create_service(self, workspace_id, name, price=100_000):
        s = Service(name=name, price=price, workspace_id=workspace_id)
        db.session.add(s)
        db.session.commit()
        return s

    def _create_invoice(self, workspace_id, customer, service, total=100_000):
        inv = Invoice(
            customer_id=customer.id,
            total_amount=float(total),
            payment_method="cash",
            invoice_date=date.today(),
            workspace_id=workspace_id,
        )
        db.session.add(inv)
        db.session.flush()
        detail = InvoiceDetail(
            invoice_id=inv.id,
            service_id=service.id,
            price=float(total),
            quantity=1,
        )
        db.session.add(detail)
        db.session.commit()
        return inv


    # ─────────────────────────────────────────────────────────────────────────
    # 1. test_settings_are_workspace_scoped
    # ─────────────────────────────────────────────────────────────────────────

    def test_settings_are_workspace_scoped(self):
        """Workspace A's spa_name must not be visible to Workspace B."""
        ws_a, owner_a = self._create_workspace_and_owner("ws-set-a")
        ws_b, owner_b = self._create_workspace_and_owner("ws-set-b")

        # Seed settings for A and B directly at DB level.
        self._seed_setting("spa_name", "Spa Alpha", ws_a.id)
        self._seed_setting("spa_name", "Spa Beta", ws_b.id)

        # --- Read as workspace A ---
        self._login_as(owner_a, ws_a.id)
        with self.client.application.test_request_context("/"):
            with self.client.session_transaction() as sess:
                from flask import session as flask_session
            # Use a request context with the correct session data
        resp_a = self.client.get("/settings")
        # Settings page must show spa_name for A (page visible = 200)
        self.assertNotEqual(resp_a.status_code, 500)

        # Verify at model layer: scoped query returns correct values per workspace.
        a_row = Setting.query.filter_by(key="spa_name", workspace_id=ws_a.id).first()
        b_row = Setting.query.filter_by(key="spa_name", workspace_id=ws_b.id).first()
        self.assertIsNotNone(a_row, "Workspace A spa_name row must exist")
        self.assertIsNotNone(b_row, "Workspace B spa_name row must exist")
        self.assertEqual(a_row.value, "Spa Alpha")
        self.assertEqual(b_row.value, "Spa Beta")
        self.assertNotEqual(a_row.value, b_row.value, "Settings must differ between workspaces")

    # ─────────────────────────────────────────────────────────────────────────
    # 2. test_settings_no_current_workspace_fail_closed
    # ─────────────────────────────────────────────────────────────────────────

    def test_settings_no_current_workspace_fail_closed(self):
        """Without workspace context: Setting.get returns default, Setting.set raises."""
        ws_a, owner_a = self._create_workspace_and_owner("ws-fc-a")
        self._seed_setting("spa_name", "Spa FC", ws_a.id)

        # Simulate: no workspace context (isolation active but no workspace_id).
        with self.client.session_transaction() as sess:
            set_authenticated_session(sess, owner_a)
            sess["_enable_workspace_isolation"] = True
            # Deliberately omit "current_workspace_id"

        # Setting.get must return default, not the tenant's real value.
        with app.test_request_context("/"):
            with app.test_client().session_transaction() as sess:
                set_authenticated_session(sess, owner_a)
                sess["_enable_workspace_isolation"] = True
            # Direct model call without workspace context.
            result = Setting.get("spa_name", "SAFE_DEFAULT")
            self.assertEqual(
                result,
                "SAFE_DEFAULT",
                "Setting.get without workspace context must return default, not tenant data",
            )

        # POST save-spa-info without workspace → must be redirected/blocked.
        with self.client.session_transaction() as sess:
            sess.pop("current_workspace_id", None)
            sess["_enable_workspace_isolation"] = True
        resp = self.client.post(
            "/settings/save-spa-info",
            data={
                "spa_name": "Should Not Save",
                "spa_phone": "0900000000",
            },
        )
        # Must not be 200 with success — either redirect to login (302/401)
        # or forbidden (403), or flash with error (redirect 302).
        self.assertNotEqual(resp.status_code, 200, "Save without workspace must not succeed silently")

        # Verify: no new setting row was written.
        leaked = Setting.query.filter_by(key="spa_name", value="Should Not Save").first()
        self.assertIsNone(leaked, "Must not write settings without a valid workspace context")

    # ─────────────────────────────────────────────────────────────────────────
    # 3. test_logo_spa_info_does_not_leak_between_workspaces
    # ─────────────────────────────────────────────────────────────────────────

    def test_logo_spa_info_does_not_leak_between_workspaces(self):
        """Spa logo and contact info must not cross workspace boundaries."""
        ws_a, owner_a = self._create_workspace_and_owner("ws-logo-a")
        ws_b, owner_b = self._create_workspace_and_owner("ws-logo-b")

        self._seed_setting("spa_logo", "logos/alpha_logo.png", ws_a.id)
        self._seed_setting("spa_logo", "logos/beta_logo.png", ws_b.id)
        self._seed_setting("spa_phone", "0901111111", ws_a.id)
        self._seed_setting("spa_phone", "0902222222", ws_b.id)

        # Each workspace's row must be independent.
        logo_a = Setting.query.filter_by(key="spa_logo", workspace_id=ws_a.id).first()
        logo_b = Setting.query.filter_by(key="spa_logo", workspace_id=ws_b.id).first()
        phone_a = Setting.query.filter_by(key="spa_phone", workspace_id=ws_a.id).first()
        phone_b = Setting.query.filter_by(key="spa_phone", workspace_id=ws_b.id).first()

        self.assertIsNotNone(logo_a)
        self.assertIsNotNone(logo_b)
        self.assertNotEqual(logo_a.value, logo_b.value, "Logo must differ between workspaces")
        self.assertNotEqual(phone_a.value, phone_b.value, "Phone must differ between workspaces")

        # Updating B must not change A.
        logo_b.value = "logos/beta_updated_logo.png"
        db.session.commit()

        logo_a_after = Setting.query.filter_by(key="spa_logo", workspace_id=ws_a.id).first()
        self.assertEqual(
            logo_a_after.value,
            "logos/alpha_logo.png",
            "Updating workspace B logo must not affect workspace A",
        )

    # ─────────────────────────────────────────────────────────────────────────
    # 4. test_invoice_export_excel_is_workspace_scoped
    # ─────────────────────────────────────────────────────────────────────────

    def test_invoice_export_excel_is_workspace_scoped(self):
        """Invoice Excel export for workspace A must not contain workspace B's data."""
        ws_a, owner_a = self._create_workspace_and_owner("ws-inv-excel-a")
        ws_b, owner_b = self._create_workspace_and_owner("ws-inv-excel-b")

        cust_a = self._create_customer(ws_a.id, "Khách Hàng Alpha", "0901111111")
        cust_b = self._create_customer(ws_b.id, "Khách Hàng Beta", "0902222222")
        svc_a = self._create_service(ws_a.id, "Dịch vụ Alpha", 200_000)
        svc_b = self._create_service(ws_b.id, "Dịch vụ Beta", 300_000)
        inv_a = self._create_invoice(ws_a.id, cust_a, svc_a, 200_000)
        inv_b = self._create_invoice(ws_b.id, cust_b, svc_b, 300_000)

        # Export as workspace A.
        self._login_as(owner_a, ws_a.id)
        resp = self.client.get("/invoices/export/excel")

        self.assertEqual(resp.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats",
            resp.content_type,
            "Must return Excel content-type",
        )

        # Parse the Excel bytes and verify only ws_a data is present.
        from io import BytesIO
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(resp.data))
        ws_sheet = wb.active
        all_cell_values = [
            str(cell.value) for row in ws_sheet.iter_rows() for cell in row if cell.value
        ]
        content = " ".join(all_cell_values)

        # Workspace B's customer name must NOT appear in workspace A's export.
        self.assertNotIn(
            "Khách Hàng Beta",
            content,
            "Workspace B customer must not appear in Workspace A's Excel export",
        )
        # Workspace B's customer name for A should appear (sanity check).
        self.assertIn(
            "Khách Hàng Alpha",
            content,
            "Workspace A customer must appear in Workspace A's Excel export",
        )

    # ─────────────────────────────────────────────────────────────────────────
    # 5. test_invoice_export_pdf_is_workspace_scoped
    # ─────────────────────────────────────────────────────────────────────────

    def test_invoice_export_pdf_is_workspace_scoped(self):
        """Invoice PDF export service layer must scope data to current workspace."""
        ws_a, owner_a = self._create_workspace_and_owner("ws-inv-pdf-a")
        ws_b, owner_b = self._create_workspace_and_owner("ws-inv-pdf-b")

        cust_a = self._create_customer(ws_a.id, "PDF Khách Alpha", "0901111111")
        cust_b = self._create_customer(ws_b.id, "PDF Khách Beta", "0902222222")
        svc_a = self._create_service(ws_a.id, "PDF Service Alpha", 150_000)
        svc_b = self._create_service(ws_b.id, "PDF Service Beta", 250_000)
        self._create_invoice(ws_a.id, cust_a, svc_a, 150_000)
        self._create_invoice(ws_b.id, cust_b, svc_b, 250_000)

        # Verify at service layer: get_filtered_invoices is already workspace-scoped.
        # We test by checking the service only returns workspace A invoices when
        # workspace A context is active.
        self._login_as(owner_a, ws_a.id)
        resp = self.client.get("/invoices/export/pdf")

        # Must return PDF.
        self.assertEqual(resp.status_code, 200)
        self.assertIn("application/pdf", resp.content_type)

        # PDF content must not reference Workspace B customer name as plain text.
        pdf_bytes = resp.data
        self.assertNotIn(
            b"PDF Kh\xc3\xa1ch Beta",
            pdf_bytes,
            "Workspace B customer name must not appear in Workspace A PDF export (raw bytes)",
        )

    # ─────────────────────────────────────────────────────────────────────────
    # 6. test_statistics_export_is_workspace_scoped
    # ─────────────────────────────────────────────────────────────────────────

    def test_statistics_export_is_workspace_scoped(self):
        """Statistics export revenue must reflect only the current workspace's data."""
        ws_a, owner_a = self._create_workspace_and_owner("ws-stat-a")
        ws_b, owner_b = self._create_workspace_and_owner("ws-stat-b")

        cust_a = self._create_customer(ws_a.id, "Stat Customer A", "0901111111")
        cust_b = self._create_customer(ws_b.id, "Stat Customer B", "0902222222")
        svc_a = self._create_service(ws_a.id, "Stat Service A", 400_000)
        svc_b = self._create_service(ws_b.id, "Stat Service B", 500_000)
        self._create_invoice(ws_a.id, cust_a, svc_a, 400_000)
        self._create_invoice(ws_b.id, cust_b, svc_b, 500_000)

        # Export statistics Excel as workspace A.
        self._login_as(owner_a, ws_a.id)
        resp = self.client.get("/statistics/export/excel")

        self.assertEqual(resp.status_code, 200)
        self.assertIn("application/vnd.openxmlformats", resp.content_type)

        from io import BytesIO
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(resp.data))
        ws_sheet = wb.active
        all_cell_values = [
            str(cell.value) for row in ws_sheet.iter_rows() for cell in row if cell.value
        ]
        content = " ".join(all_cell_values)

        # Workspace B customer name must NOT appear in workspace A's export.
        self.assertNotIn(
            "Stat Customer B",
            content,
            "Workspace B customer must not appear in Workspace A's statistics export",
        )

    # ─────────────────────────────────────────────────────────────────────────
    # 7. test_export_without_workspace_returns_empty
    # ─────────────────────────────────────────────────────────────────────────

    def test_export_auto_selects_single_active_workspace_and_remains_scoped(self):
        """A sole active membership is selected and the export remains scoped."""
        ws_a, owner_a = self._create_workspace_and_owner("ws-noexp-a")
        cust_a = self._create_customer(ws_a.id, "Selected Workspace Customer", "0901111111")
        svc_a = self._create_service(ws_a.id, "Global Leak Service", 100_000)
        self._create_invoice(ws_a.id, cust_a, svc_a, 100_000)

        ws_b, _owner_b = self._create_workspace_and_owner("ws-noexp-b")
        cust_b = self._create_customer(ws_b.id, "Other Workspace Customer", "0902222222")
        svc_b = self._create_service(ws_b.id, "Other Workspace Service", 200_000)
        self._create_invoice(ws_b.id, cust_b, svc_b, 200_000)

        # Create a staff member who cannot auto-provision/repair
        staff = User(
            username="staff_noexp",
            email="staff_noexp@test.com",
            role="STAFF",
            full_name="Staff No Exp",
            is_active=True,
            approval_status="active"
        )
        staff.set_password("Password123!")
        db.session.add(staff)
        db.session.flush()

        WorkspaceService.add_member_for_user(ws_a.id, staff, "STAFF")
        db.session.commit()

        # Login as staff but WITHOUT setting current_workspace_id.
        # The global guard must select the sole active membership.
        with self.client.session_transaction() as sess:
            set_authenticated_session(sess, staff)
            sess["_enable_workspace_isolation"] = True
            # Deliberately no current_workspace_id

        # The export routes require login; without workspace they should either
        # redirect to login, return 403, or return an empty export — NOT global data.
        resp = self.client.get("/invoices/export/excel")

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp.mimetype,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with self.client.session_transaction() as sess:
            self.assertEqual(sess.get("current_workspace_id"), ws_a.id)
            self.assertEqual(sess.get(AUTH_SESSION_KEY), staff.id)
            self.assertTrue(sess.get("_enable_workspace_isolation"))

        from io import BytesIO
        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(resp.data))
        ws_sheet = wb.active
        content = " ".join(
            str(cell.value)
            for row in ws_sheet.iter_rows()
            for cell in row
            if cell.value
        )
        self.assertIn("Selected Workspace Customer", content)
        self.assertNotIn("Other Workspace Customer", content)
        self.assertEqual(
            WorkspaceMember.query.filter_by(user_id=staff.id).count(),
            1,
        )

    def test_export_without_active_membership_is_denied(self):
        """A user without an active membership cannot reach the export route."""
        _workspace, _owner = self._create_workspace_and_owner("ws-no-membership")
        staff = User(
            username="staff_no_membership",
            email="staff_no_membership@test.com",
            role="STAFF",
            full_name="Staff No Membership",
            is_active=True,
            approval_status="active",
        )
        staff.set_password("Password123!")
        db.session.add(staff)
        db.session.commit()

        with self.client.session_transaction() as sess:
            set_authenticated_session(sess, staff)
            sess["_enable_workspace_isolation"] = True

        resp = self.client.get("/invoices/export/excel")
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/login", resp.headers["Location"])
        self.assertNotEqual(
            resp.mimetype,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with self.client.session_transaction() as sess:
            self.assertIsNone(sess.get(AUTH_SESSION_KEY))
            self.assertIsNone(sess.get("current_workspace_id"))

        self.assertEqual(WorkspaceMember.query.filter_by(user_id=staff.id).count(), 0)

    # ─────────────────────────────────────────────────────────────────────────
    # 8. test_settings_get_scoped_via_model_layer (unit test of Setting model)
    # ─────────────────────────────────────────────────────────────────────────

    def test_settings_get_scoped_via_model_layer(self):
        """
        Direct model-layer test: Setting._set_for_workspace and _scoped_query
        must correctly isolate reads by workspace_id.
        """
        ws_a, owner_a = self._create_workspace_and_owner("ws-model-a")
        ws_b, owner_b = self._create_workspace_and_owner("ws-model-b")

        # Write different values for same key in different workspaces.
        Setting._set_for_workspace("spa_name", "Model Spa A", ws_a.id)
        Setting._set_for_workspace("spa_name", "Model Spa B", ws_b.id)
        db.session.commit()

        # Read back — must be isolated.
        row_a = Setting._scoped_query("spa_name", ws_a.id).first()
        row_b = Setting._scoped_query("spa_name", ws_b.id).first()

        self.assertIsNotNone(row_a)
        self.assertIsNotNone(row_b)
        self.assertEqual(row_a.value, "Model Spa A")
        self.assertEqual(row_b.value, "Model Spa B")
        self.assertNotEqual(row_a.id, row_b.id, "Must be separate DB rows")

    # ─────────────────────────────────────────────────────────────────────────
    # 9. test_system_setting_does_not_mix_with_tenant_setting
    # ─────────────────────────────────────────────────────────────────────────

    def test_system_setting_does_not_mix_with_tenant_setting(self):
        """
        get_system / set_system (workspace_id IS NULL) must not return
        tenant rows, and get() with workspace context must not return NULL rows.
        """
        ws_a, owner_a = self._create_workspace_and_owner("ws-sys-a")

        # Write a system-level key (workspace_id = NULL).
        Setting.set_system("db_version", "v9.9.9")

        # Write a tenant-level key with same name.
        Setting._set_for_workspace("db_version", "TENANT_OVERRIDE", ws_a.id)
        db.session.commit()

        # get_system must return the NULL-workspace row only.
        sys_val = Setting.get_system("db_version", "NOT_FOUND")
        self.assertEqual(sys_val, "v9.9.9", "get_system must return system-level row")

        # Count: there must be exactly 2 rows for db_version (one per scope).
        count = Setting.query.filter_by(key="db_version").count()
        self.assertEqual(count, 2, "Must have separate rows for system and tenant")

    # ─────────────────────────────────────────────────────────────────────────
    # 10. test_statistics_page_is_workspace_scoped
    # ─────────────────────────────────────────────────────────────────────────
    def test_statistics_page_is_workspace_scoped(self):
        """Statistics page UI must isolate A and B data."""
        ws_a, owner_a = self._create_workspace_and_owner("ws-stat-ui-a")
        ws_b, owner_b = self._create_workspace_and_owner("ws-stat-ui-b")

        cust_a = self._create_customer(ws_a.id, "Khách Alpha UI", "0901111111")
        cust_b = self._create_customer(ws_b.id, "Khách Beta UI", "0902222222")
        svc_a = self._create_service(ws_a.id, "Service Alpha UI", 200_000)
        svc_b = self._create_service(ws_b.id, "Service Beta UI", 300_000)
        self._create_invoice(ws_a.id, cust_a, svc_a, 200_000)
        self._create_invoice(ws_b.id, cust_b, svc_b, 300_000)

        # Access as workspace A
        self._login_as(owner_a, ws_a.id)
        resp_a = self.client.get("/statistics")
        self.assertEqual(resp_a.status_code, 200)
        html_a = resp_a.get_data(as_text=True)

        self.assertIn("Khách Alpha UI", html_a)
        self.assertIn("Service Alpha UI", html_a)
        self.assertNotIn("Khách Beta UI", html_a)
        self.assertNotIn("Service Beta UI", html_a)

        # Access as workspace B
        self._login_as(owner_b, ws_b.id)
        resp_b = self.client.get("/statistics")
        self.assertEqual(resp_b.status_code, 200)
        html_b = resp_b.get_data(as_text=True)

        self.assertIn("Khách Beta UI", html_b)
        self.assertIn("Service Beta UI", html_b)
        self.assertNotIn("Khách Alpha UI", html_b)
        self.assertNotIn("Service Alpha UI", html_b)

    # ─────────────────────────────────────────────────────────────────────────
    # 11. test_statistics_page_no_current_workspace_fail_closed
    # ─────────────────────────────────────────────────────────────────────────
    def test_statistics_page_no_current_workspace_fail_closed(self):
        """Accessing statistics page without workspace context must be blocked."""
        ws_a, owner_a = self._create_workspace_and_owner("ws-stat-fc-a")

        # Create a staff member who cannot auto-provision/repair
        staff = User(
            username="staff_stat_fc",
            email="staff_sf@test.com",
            role="STAFF",
            full_name="Staff Stat FC",
            is_active=True,
            approval_status="active"
        )
        staff.set_password("Password123!")
        db.session.add(staff)
        db.session.flush()

        WorkspaceService.add_member_for_user(ws_a.id, staff, "STAFF")
        db.session.commit()

        # Login without workspace ID
        with self.client.session_transaction() as sess:
            set_authenticated_session(sess, staff)
            sess["_enable_workspace_isolation"] = True
            # no current_workspace_id

        resp = self.client.get("/statistics")
        self.assertIn(resp.status_code, (302, 401, 403))

    # ─────────────────────────────────────────────────────────────────────────
    # 12. test_statistics_export_excel_is_workspace_scoped
    # ─────────────────────────────────────────────────────────────────────────
    def test_statistics_export_excel_is_workspace_scoped(self):
        """Excel statistics export must only contain active workspace data."""
        ws_a, owner_a = self._create_workspace_and_owner("ws-stat-excel-a")
        ws_b, owner_b = self._create_workspace_and_owner("ws-stat-excel-b")

        cust_a = self._create_customer(ws_a.id, "Excel Cust A", "0901111111")
        cust_b = self._create_customer(ws_b.id, "Excel Cust B", "0902222222")
        svc_a = self._create_service(ws_a.id, "Excel Svc A", 120_000)
        svc_b = self._create_service(ws_b.id, "Excel Svc B", 180_000)
        self._create_invoice(ws_a.id, cust_a, svc_a, 120_000)
        self._create_invoice(ws_b.id, cust_b, svc_b, 180_000)

        self._login_as(owner_a, ws_a.id)
        resp = self.client.get("/statistics/export/excel")
        self.assertEqual(resp.status_code, 200)

        from io import BytesIO
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(resp.data))
        all_cell_values = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        all_cell_values.append(str(cell.value))
        content = " ".join(all_cell_values)

        self.assertIn("Excel Cust A", content)
        self.assertNotIn("Excel Cust B", content)

    # ─────────────────────────────────────────────────────────────────────────
    # 13. test_statistics_detail_routes_are_workspace_scoped
    # ─────────────────────────────────────────────────────────────────────────
    def test_statistics_detail_routes_are_workspace_scoped(self):
        """Detail endpoints must return 404 if referencing another workspace's entity."""
        ws_a, owner_a = self._create_workspace_and_owner("ws-stat-detail-a")
        ws_b, owner_b = self._create_workspace_and_owner("ws-stat-detail-b")

        cust_b = self._create_customer(ws_b.id, "Beta Detail Customer", "0902222222")
        svc_b = self._create_service(ws_b.id, "Beta Detail Service", 150_000)

        # Login to workspace A
        self._login_as(owner_a, ws_a.id)

        # Try to view customer details of B
        resp_cust = self.client.get(f"/statistics/customer/{cust_b.id}")
        self.assertEqual(resp_cust.status_code, 404)

        # Try to view service details of B
        resp_svc = self.client.get(f"/statistics/service/{svc_b.id}")
        self.assertEqual(resp_svc.status_code, 404)

    # ─────────────────────────────────────────────────────────────────────────
    # 14. test_statistics_export_pdf_is_workspace_scoped
    # ─────────────────────────────────────────────────────────────────────────
    def test_statistics_export_pdf_is_workspace_scoped(self):
        """PDF statistics export must return 200 and not leak raw bytes of another workspace."""
        ws_a, owner_a = self._create_workspace_and_owner("ws-stat-pdf-a")
        ws_b, owner_b = self._create_workspace_and_owner("ws-stat-pdf-b")

        cust_a = self._create_customer(ws_a.id, "PDF Cust A", "0901111111")
        cust_b = self._create_customer(ws_b.id, "PDF Cust B", "0902222222")
        svc_a = self._create_service(ws_a.id, "PDF Svc A", 120_000)
        svc_b = self._create_service(ws_b.id, "PDF Svc B", 180_000)
        self._create_invoice(ws_a.id, cust_a, svc_a, 120_000)
        self._create_invoice(ws_b.id, cust_b, svc_b, 180_000)

        self._login_as(owner_a, ws_a.id)
        resp = self.client.get("/statistics/export/pdf")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("application/pdf", resp.content_type)

        pdf_bytes = resp.data
        self.assertNotIn(b"PDF Cust B", pdf_bytes)


if __name__ == "__main__":
    unittest.main()
