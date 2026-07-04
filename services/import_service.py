import os
import re
import time
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy import func
from flask import url_for

from extensions import db
from models.customer import Customer
from models.service import Service
from core.logger import app_logger
from validators.import_validator import ImportValidator
from services.backup_service import BackupService
from services.customer_service import CustomerService
from services.activity_log_service import ActivityLogService
from utils.timezone_utils import local_now

# services/import_service.py


class ImportService:
    """Service for importing customers and services from Excel files using an advanced multi-step wizard."""

    # Template column definitions
    CUSTOMER_COLUMNS = ['Họ tên', 'Số điện thoại', 'Email', 'Địa chỉ']
    SERVICE_COLUMNS = ['Tên dịch vụ', 'Giá (VND)', 'Thời lượng (phút)', 'Mô tả', 'Nhóm dịch vụ']

    @staticmethod
    def get_template_dir(app):
        """Get the template directory path, creating it if needed."""
        template_dir = os.path.join(app.root_path, 'static', 'templates', 'import')
        os.makedirs(template_dir, exist_ok=True)
        return template_dir

    @staticmethod
    def get_upload_dir(app):
        """Get the upload directory for import files."""
        upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'import')
        os.makedirs(upload_dir, exist_ok=True)
        return upload_dir

    # ──────────────────────────────────────────────
    # Template Generation
    # ──────────────────────────────────────────────

    @staticmethod
    def generate_customer_template(app):
        """Generate customers_template.xlsx with headers and sample data."""
        template_dir = ImportService.get_template_dir(app)
        filepath = os.path.join(template_dir, 'customers_template.xlsx')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Khách hàng'

        # Header row
        for col_idx, header in enumerate(ImportService.CUSTOMER_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)

        # Sample data rows
        sample_data = [
            ['Nguyễn Văn A', '0901234567', 'nguyenvana@email.com', '123 Đường ABC, Quận 1, TP.HCM'],
            ['Trần Thị B', '0912345678', 'tranthib@email.com', '456 Đường XYZ, Quận 2, TP.HCM'],
        ]
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 45

        wb.save(filepath)
        wb.close()

        return filepath

    @staticmethod
    def generate_service_template(app):
        """Generate services_template.xlsx with headers and sample data."""
        template_dir = ImportService.get_template_dir(app)
        filepath = os.path.join(template_dir, 'services_template.xlsx')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Dịch vụ'

        # Header row
        for col_idx, header in enumerate(ImportService.SERVICE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)

        # Sample data rows
        sample_data = [
            ['Massage thư giãn', 300000, 60, 'Massage toàn thân giúp thư giãn cơ thể', 'Massage'],
            ['Chăm sóc da mặt', 500000, 90, 'Chăm sóc da chuyên sâu với sản phẩm cao cấp', 'Spa'],
            ['Sơn gel cao cấp', 150000, 45, 'Sơn gel bền đẹp với nhiều màu sắc lựa chọn', 'Nail'],
            ['Trang điểm cô dâu', 800000, 90, 'Trang điểm cô dâu ngày cưới chuyên nghiệp', 'Makeup'],
            ['Cắt và uốn tóc', 600000, 120, 'Cắt tạo kiểu và uốn tóc phục hồi Collagen', 'Làm tóc'],
            ['Dịch vụ khác VIP', 1000000, 120, 'Dịch vụ đặc biệt theo yêu cầu', 'Khác'],
        ]
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 20

        wb.save(filepath)
        wb.close()

        return filepath

    # ──────────────────────────────────────────────
    # Validation Helpers
    # ──────────────────────────────────────────────

    @staticmethod
    def _validate_file_format(filepath):
        """Validate that file is a valid Excel file."""
        if not filepath.lower().endswith(('.xlsx', '.xls')):
            return False, 'File phải có định dạng .xlsx hoặc .xls'
        try:
            wb = load_workbook(filepath, read_only=True)
            wb.close()
            return True, None
        except Exception:
            return False, 'File Excel không hợp lệ hoặc bị hỏng.'

    @staticmethod
    def _validate_columns(ws, expected_columns):
        """Validate that the worksheet has the expected column headers."""
        headers = []
        for col_idx in range(1, len(expected_columns) + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            headers.append(str(cell_value).strip() if cell_value else '')

        missing = []
        for idx, expected in enumerate(expected_columns):
            if idx >= len(headers) or headers[idx] != expected:
                missing.append(expected)

        if missing:
            return False, f'Thiếu hoặc sai tên cột: {", ".join(missing)}'
        return True, None

    @staticmethod
    def _is_valid_email(email):
        """Simple email validation."""
        if not email:
            return True  # email is optional
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, str(email).strip()) is not None

    @staticmethod
    def _is_valid_phone(phone):
        """Phone must be exactly 10 digits starting with 0 (not 00)."""
        if not phone:
            return True  # phone is optional
        cleaned = ''.join(c for c in str(phone) if c.isdigit())
        pattern = r'^0[1-9]\d{8}$'
        return re.match(pattern, cleaned) is not None

    # ──────────────────────────────────────────────
    # Wizard Step 2 & 3: File Analysis
    # ──────────────────────────────────────────────

    @staticmethod
    def analyze_file(app, filepath, import_type):
        """
        Analyze the Excel file and return preview rows, validation errors, and duplicate checks.
        Does not write to database.
        """
        validator = ImportValidator()
        validator.validate({'import_type': import_type, 'duplicate_action': 'skip'})
        validator.raise_if_invalid("Thông tin nhập khẩu không hợp lệ.")

        is_valid, error = ImportService._validate_file_format(filepath)
        if not is_valid:
            return {'success': False, 'message': error}

        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
        except Exception as e:
            return {'success': False, 'message': f'Không thể mở file Excel: {str(e)}'}

        # Validate columns based on import_type
        expected_cols = ImportService.CUSTOMER_COLUMNS if import_type == 'customers' else ImportService.SERVICE_COLUMNS
        is_valid, error = ImportService._validate_columns(ws, expected_cols)
        if not is_valid:
            wb.close()
            return {'success': False, 'message': error}

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        total_rows = len(rows)
        if total_rows == 0:
            return {'success': False, 'message': 'File Excel trống, không có dữ liệu.'}

        # Build preview data (up to 20 rows)
        preview_rows = []
        for idx, row in enumerate(rows[:20]):
            preview_rows.append({
                'row_index': idx + 2,
                'cells': [str(cell) if cell is not None else '' for cell in row[:len(expected_cols)]]
            })

        # Preload database records to check duplicates efficiently
        existing_service_names = {}
        customer_records = []

        if import_type == 'customers':
            customer_records = Customer.query.all()
        else:
            services = Service.query.filter(Service.deleted_at.is_(None)).all()
            for s in services:
                if s.name:
                    existing_service_names[str(s.name).strip().lower()] = s.id

        validation_results = []
        seen_phones = {}
        seen_emails = {}
        seen_service_names = {}

        for idx, row in enumerate(rows, start=2):
            row_errors = []
            is_duplicate = False
            dup_reasons = []

            if import_type == 'customers':
                name = str(row[0]).strip() if row[0] is not None else ''
                phone = str(row[1]).strip() if row[1] is not None else ''
                email = str(row[2]).strip() if row[2] is not None else ''
                address = str(row[3]).strip() if row[3] is not None else ''

                # Validate
                if not name:
                    row_errors.append('H? t?n kh?ng ???c ?? tr?ng')

                normalized_phone = CustomerService.normalize_customer_phone(phone)
                normalized_email = CustomerService.normalize_customer_email(email)

                if phone:
                    if not ImportService._is_valid_phone(phone):
                        row_errors.append('S? ?i?n tho?i ph?i g?m ??ng 10 ch? s? v? b?t ??u b?ng s? 0')
                    elif normalized_phone:
                        phone_conflicts = CustomerService.find_duplicate_conflicts(
                            phone=normalized_phone,
                            email=None,
                            include_deleted=True,
                            customer_records=customer_records,
                        )
                        if phone_conflicts['phone']:
                            is_duplicate = True
                            dup_reasons.extend(CustomerService._build_duplicate_messages(phone_conflicts))
                        if normalized_phone in seen_phones:
                            is_duplicate = True
                            dup_reasons.append(f'S? ?i?n tho?i tr?ng v?i d?ng {seen_phones[normalized_phone]}')
                        else:
                            seen_phones[normalized_phone] = idx

                if email:
                    if not ImportService._is_valid_email(email):
                        row_errors.append('Email kh?ng ??ng ??nh d?ng')
                    elif normalized_email:
                        email_conflicts = CustomerService.find_duplicate_conflicts(
                            phone=None,
                            email=normalized_email,
                            include_deleted=True,
                            customer_records=customer_records,
                        )
                        if email_conflicts['email']:
                            is_duplicate = True
                            dup_reasons.extend(CustomerService._build_duplicate_messages(email_conflicts))
                        if normalized_email in seen_emails:
                            is_duplicate = True
                            dup_reasons.append(f'Email tr?ng v?i d?ng {seen_emails[normalized_email]}')
                        else:
                            seen_emails[normalized_email] = idx

                duplicate_conflicts = {
                    'phone': [
                        {
                            'id': customer.id,
                            'deleted_at': customer.deleted_at is not None,
                        }
                        for customer in CustomerService.find_duplicate_conflicts(
                            phone=normalized_phone,
                            email=None,
                            include_deleted=True,
                            customer_records=customer_records,
                        )['phone']
                    ] if normalized_phone else [],
                    'email': [
                        {
                            'id': customer.id,
                            'deleted_at': customer.deleted_at is not None,
                        }
                        for customer in CustomerService.find_duplicate_conflicts(
                            phone=None,
                            email=normalized_email,
                            include_deleted=True,
                            customer_records=customer_records,
                        )['email']
                    ] if normalized_email else [],
                }

                validation_results.append({
                    'row_index': idx,
                    'data': {'name': name, 'phone': phone, 'email': email, 'address': address},
                    'errors': row_errors,
                    'is_duplicate': is_duplicate,
                    'duplicate_reason': '; '.join(dup_reasons),
                    'duplicate_conflicts': duplicate_conflicts,
                })
            else:  # services
                name = str(row[0]).strip() if row[0] is not None else ''
                price_raw = row[1]
                duration_raw = row[2]
                description = str(row[3]).strip() if row[3] is not None else ''
                category_raw = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''

                if not name:
                    row_errors.append('Tên dịch vụ không được để trống')
                else:
                    name_lower = name.lower()
                    # Check DB duplicate
                    if name_lower in existing_service_names:
                        is_duplicate = True
                        dup_reasons.append('Tên dịch vụ đã tồn tại trong hệ thống')
                    # Check batch duplicate
                    if name_lower in seen_service_names:
                        is_duplicate = True
                        dup_reasons.append(f'Tên dịch vụ trùng với dòng {seen_service_names[name_lower]}')
                    else:
                        seen_service_names[name_lower] = idx

                # Price validation
                price = 0.0
                if price_raw is None or str(price_raw).strip() == '':
                    row_errors.append('Giá dịch vụ không được để trống')
                else:
                    try:
                        price = float(price_raw)
                        if price < 0:
                            row_errors.append('Giá dịch vụ không được là số âm')
                    except (ValueError, TypeError):
                        row_errors.append('Giá dịch vụ phải là số hợp lệ')

                # Duration validation
                duration = None
                if duration_raw is not None and str(duration_raw).strip() != '':
                    try:
                        duration = int(float(duration_raw))
                        if duration < 0:
                            row_errors.append('Thời lượng dịch vụ không được là số âm')
                    except (ValueError, TypeError):
                        row_errors.append('Thời lượng phải là số nguyên phút')

                validation_results.append({
                    'row_index': idx,
                    'data': {
                        'name': name, 'price': price_raw, 'duration': duration_raw, 
                        'description': description, 'category': category_raw
                    },
                    'errors': row_errors,
                    'is_duplicate': is_duplicate,
                    'duplicate_reason': '; '.join(dup_reasons)
                })

        return {
            'success': True,
            'import_type': import_type,
            'total_rows': total_rows,
            'headers': expected_cols,
            'preview_rows': preview_rows,
            'validation_results': validation_results
        }

    # ──────────────────────────────────────────────
    # Wizard Step 6: Import Execution
    # ──────────────────────────────────────────────

    @staticmethod
    def execute_import(app, filepath, import_type, duplicate_action, all_or_nothing):
        """
        Execute the import with a database transaction, handle duplicates, and output a detailed report.
        """
        validator = ImportValidator()
        validator.validate({'import_type': import_type, 'duplicate_action': duplicate_action})
        validator.raise_if_invalid("Thông tin nhập khẩu không hợp lệ.")


        report = {
            'total': 0,
            'success': 0,      # newly inserted
            'overwritten': 0,  # updated
            'skipped': 0,      # ignored due to duplicate settings
            'failed': 0,       # validation error rows skipped
            'errors': [],
            'error_report_url': None
        }

        # 1. Automatically create a backup ("Before Import")
        try:
            BackupService.create_backup(app, notes=f"Tự động sao lưu trước khi import {import_type}", backup_type="Before Import")
        except Exception as e:
            app_logger.error(f"Failed to create backup before import: {e}", module="IMPORT", exc_info=True)

        # 2. Run analysis to get duplicate and validation details
        analysis = ImportService.analyze_file(app, filepath, import_type)
        if not analysis.get('success'):
            report['errors'].append({'row': 0, 'message': analysis.get('message', 'Không thể phân tích file.')})
            report['failed'] = 1
            return report

        total_rows = analysis['total_rows']
        report['total'] = total_rows
        results = analysis['validation_results']

        # 3. Check All-or-Nothing constraint
        has_any_error = any(len(r['errors']) > 0 for r in results)
        if all_or_nothing and has_any_error:
            report['failed'] = total_rows
            report['errors'] = [{'row': r['row_index'], 'message': '; '.join(r['errors'])} for r in results if r['errors']]
            
            # Log failure
            ActivityLogService.log_action(
                module=ActivityLogService.MODULE_SETTINGS,
                action=ActivityLogService.ACTION_IMPORT,
                description=f"Import {import_type} thất bại (Chế độ All-or-Nothing: Phát hiện {len(report['errors'])} dòng lỗi)",
                severity=ActivityLogService.SEVERITY_ERROR
            )
            return report

        # Write start activity log
        app_logger.info(f"Starting import of {total_rows} {import_type} records (mode: {duplicate_action})", module="IMPORT")
        ActivityLogService.log_action(
            module=ActivityLogService.MODULE_SETTINGS,
            action=ActivityLogService.ACTION_IMPORT,
            description=f"Bắt đầu quá trình import {total_rows} dòng {import_type} (Chế độ: {duplicate_action})",
            severity=ActivityLogService.SEVERITY_INFO
        )

        try:
            for r in results:
                row_idx = r['row_index']
                row_data = r['data']
                row_errors = r['errors']
                is_dup = r['is_duplicate']

                if row_errors:
                    # Line has validation errors
                    report['failed'] += 1
                    report['errors'].append({'row': row_idx, 'message': '; '.join(row_errors)})
                    continue

                if import_type == 'customers':
                    name = row_data['name']
                    phone = row_data['phone'] if row_data['phone'] else None
                    email = row_data['email'] if row_data['email'] else None
                    address = row_data['address'] if row_data['address'] else None
                    normalized_phone = CustomerService.normalize_customer_phone(phone)
                    normalized_email = CustomerService.normalize_customer_email(email)
                    duplicate_conflicts = row_data.get('duplicate_conflicts') or {'phone': [], 'email': []}
                    has_active_duplicate = any(not conflict.get('deleted_at') for conflict in duplicate_conflicts.get('phone', [])) or any(not conflict.get('deleted_at') for conflict in duplicate_conflicts.get('email', []))
                    has_deleted_duplicate = any(conflict.get('deleted_at') for conflict in duplicate_conflicts.get('phone', [])) or any(conflict.get('deleted_at') for conflict in duplicate_conflicts.get('email', []))

                    if is_dup:
                        if duplicate_action == 'skip' or duplicate_action == 'insert_only':
                            report['skipped'] += 1
                            report['errors'].append({'row': row_idx, 'message': f'B? qua d?ng do tr?ng th?ng tin (S?T: {phone}, Email: {email})'})
                            continue
                        elif duplicate_action == 'overwrite':
                            existing = None
                            if has_active_duplicate:
                                active_conflict_ids = [
                                    conflict.get('id')
                                    for conflict in duplicate_conflicts.get('phone', []) + duplicate_conflicts.get('email', [])
                                    if not conflict.get('deleted_at')
                                ]
                                for conflict_id in active_conflict_ids:
                                    existing = Customer.query.get(conflict_id)
                                    if existing and existing.deleted_at is None:
                                        break
                            if existing:
                                existing.name = name
                                if phone:
                                    existing.phone = phone
                                if email:
                                    existing.email = email
                                if address:
                                    existing.address = address
                                report['overwritten'] += 1
                            elif has_deleted_duplicate:
                                report['skipped'] += 1
                                report['errors'].append({'row': row_idx, 'message': f'B? qua d?ng do tr?ng v?i kh?ch h?ng trong th?ng r?c (S?T: {phone}, Email: {email})'})
                            else:
                                customer = Customer(name=name, phone=phone, email=email, address=address)
                                db.session.add(customer)
                                report['success'] += 1
                    else:
                        # New customer
                        customer = Customer(name=name, phone=phone, email=email, address=address)
                        db.session.add(customer)
                        report['success'] += 1
                else:  # services
                    name = row_data['name']
                    price = float(row_data['price']) if row_data['price'] is not None else 0.0
                    duration = int(float(row_data['duration'])) if row_data['duration'] is not None else None
                    description = row_data['description'] if row_data['description'] else None
                    category_raw = str(row_data['category']).lower() if row_data['category'] else ''

                    # Normalize category
                    category = 'other'
                    if 'nail' in category_raw: category = 'nail'
                    elif 'massage' in category_raw: category = 'massage'
                    elif 'spa' in category_raw: category = 'spa'
                    elif 'makeup' in category_raw or 'trang điểm' in category_raw or 'make' in category_raw: category = 'makeup'
                    elif 'tóc' in category_raw or 'hair' in category_raw: category = 'hair'

                    if is_dup:
                        if duplicate_action == 'skip' or duplicate_action == 'insert_only':
                            report['skipped'] += 1
                            report['errors'].append({'row': row_idx, 'message': f'Bỏ qua dòng do trùng tên dịch vụ: "{name}"'})
                            continue
                        elif duplicate_action == 'overwrite':
                            existing = Service.query.filter(func.lower(Service.name) == name.lower(), Service.deleted_at.is_(None)).first()
                            if existing:
                                existing.price = price
                                if duration is not None:
                                    existing.duration = duration
                                if description:
                                    existing.description = description
                                existing.category = category
                                report['overwritten'] += 1
                            else:
                                service = Service(name=name, price=price, duration=duration, description=description, category=category)
                                db.session.add(service)
                                report['success'] += 1
                    else:
                        # New service
                        service = Service(name=name, price=price, duration=duration, description=description, category=category)
                        db.session.add(service)
                        report['success'] += 1

            # Commit all DB operations
            db.session.commit()

            # 5. Generate error report file if there are failed/skipped rows
            if report['failed'] > 0 or report['skipped'] > 0:
                upload_dir = ImportService.get_upload_dir(app)
                report_filename = f"bao_cao_loi_import_{int(time.time())}.txt"
                report_filepath = os.path.join(upload_dir, report_filename)
                
                with open(report_filepath, 'w', encoding='utf-8') as f:
                    f.write(f"BÁO CÁO CHI TIẾT IMPORT {import_type.upper()}\n")
                    f.write(f"Thời gian: {local_now().strftime('%d/%m/%Y %H:%M:%S')}\n")
                    f.write(f"Tổng số dòng: {total_rows}\n")
                    f.write(f"Thành công (Thêm mới): {report['success']}\n")
                    f.write(f"Ghi đè (Cập nhật): {report['overwritten']}\n")
                    f.write(f"Bỏ qua: {report['skipped']}\n")
                    f.write(f"Thất bại (Dòng lỗi): {report['failed']}\n")
                    f.write("="*50 + "\n\n")
                    
                    for err in report['errors']:
                        f.write(f"Dòng {err['row']}: {err['message']}\n")
                
                # Assign download URL
                report['error_report_url'] = url_for('setting.download_import_errors', filename=report_filename)

            # Log success
            app_logger.info(f"Import {import_type} completed successfully (Success: {report['success']}, Overwritten: {report['overwritten']}, Skipped: {report['skipped']}, Failed: {report['failed']})", module="IMPORT")
            ActivityLogService.log_action(
                module=ActivityLogService.MODULE_SETTINGS,
                action=ActivityLogService.ACTION_IMPORT,
                description=f"Import {import_type} hoàn tất (Thành công: {report['success']}, Ghi đè: {report['overwritten']}, Bỏ qua: {report['skipped']}, Thất bại: {report['failed']})",
                severity=ActivityLogService.SEVERITY_SUCCESS
            )

        except Exception as e:
            db.session.rollback()
            report['success'] = 0
            report['overwritten'] = 0
            report['skipped'] = 0
            report['failed'] = total_rows
            report['errors'] = [{'row': 0, 'message': f'Lỗi hệ thống trong transaction: {str(e)}'}]
            
            # Log system error
            app_logger.error(f"Import {import_type} failed due to database transaction error: {str(e)}", module="IMPORT", exc_info=True)
            ActivityLogService.log_action(
                module=ActivityLogService.MODULE_SETTINGS,
                action=ActivityLogService.ACTION_IMPORT,
                description=f"Import {import_type} thất bại do lỗi hệ thống DB: {str(e)}",
                severity=ActivityLogService.SEVERITY_ERROR
            )

        finally:
            # Clean up uploaded temporary excel file
            if os.path.exists(filepath):
                try:
                    os.remove(filepath)
                except Exception:
                    pass

        return report
